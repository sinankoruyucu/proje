import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from database import CustomerDatabase
from datetime import datetime, timedelta
import os
from tkcalendar import DateEntry
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import sqlite3

# Excel import için gerekli kütüphaneler





class CustomerManagementGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Diyetisyen Türkmen KURT")
        self.root.geometry("1200x700")
        self.root.configure(bg='#f8f9fa')  # Modern açık gri arka plan
        
        # Veritabanı bağlantısı ve cache
        self.db = CustomerDatabase()
        self.customer_cache = {}
        self.order_cache = {}
        self.cache_timestamp = datetime.now()
        
        # Veritabanı yükseltme mesajını gösterme (sessiz yükseltme)
        # if self.db.upgrade_message:
        #     messagebox.showinfo("Veritabanı", self.db.upgrade_message)
        
        # Ayarlar değişkenleri
        self.app_title = "Diyetisyen Türkmen KURT"  # Varsayılan başlık
        self.settings_password = "11235"  # Ayarlar şifresi
        self.title_position = "center"  # Başlık konumu: "left", "center", "right"
        
        # Ayarları yükle
        self.load_settings()
        
        # Saat ve tarih değişkenleri
        self.clock_label = None
        self.clock_running = False
        
        # Ana frame - Modern card design
        self.main_frame = ttk.Frame(root, style='Modern.TFrame')
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=20)
        
        # Modern stil ayarları
        style = ttk.Style()
        style.theme_use('clam')  # Daha modern tema
        
        # Modern renk paleti - Gradient ve profesyonel renkler
        style.configure('Modern.TFrame', 
                       background='#ffffff',
                       borderwidth=1,
                       relief='flat')
        
        style.configure('Modern.TLabel', 
                       background='#ffffff', 
                       font=('Segoe UI', 10),
                       foreground='#2c3e50')
        
        # Primary Button - Modern mavi gradient
        style.configure('Modern.TButton', 
                       background='#3498db', 
                       foreground='white',
                       font=('Segoe UI', 13, 'bold'),
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(30, 18))
        style.map('Modern.TButton',
                 background=[('active', '#2980b9'), ('pressed', '#21618c')],
                 foreground=[('active', 'white'), ('pressed', 'white')],
                 relief=[('pressed', 'flat'), ('active', 'flat')])
                 
        # Success Button - Modern yeşil
        style.configure('Success.TButton', 
                       background='#27ae60', 
                       foreground='white',
                       font=('Segoe UI', 13, 'bold'),
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(30, 18))
        style.map('Success.TButton',
                 background=[('active', '#229954'), ('pressed', '#1e8449')],
                 foreground=[('active', 'white'), ('pressed', 'white')])
                 
        # Warning Button - Modern turuncu
        style.configure('Warning.TButton', 
                       background='#f39c12', 
                       foreground='white',
                       font=('Segoe UI', 13, 'bold'),
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(30, 18))
        style.map('Warning.TButton',
                 background=[('active', '#e67e22'), ('pressed', '#d35400')],
                 foreground=[('active', 'white'), ('pressed', 'white')])
                 
        # Danger Button - Modern kırmızı
        style.configure('Danger.TButton', 
                       background='#e74c3c', 
                       foreground='white',
                       font=('Segoe UI', 13, 'bold'),
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none',
                       padding=(30, 18))
        style.map('Danger.TButton',
                 background=[('active', '#c0392b'), ('pressed', '#a93226')],
                 foreground=[('active', 'white'), ('pressed', 'white')])
        
        # Modern Entry stili - Rounded görünüm
        style.configure('Modern.TEntry', 
                       fieldbackground='#ffffff',
                       borderwidth=2,
                       relief='flat',
                       insertcolor='#3498db',
                       font=('Segoe UI', 11),
                       padding=(10, 8))
        style.map('Modern.TEntry',
                 bordercolor=[('focus', '#3498db'), ('!focus', '#bdc3c7')],
                 lightcolor=[('focus', '#3498db'), ('!focus', '#ecf0f1')],
                 darkcolor=[('focus', '#3498db'), ('!focus', '#ecf0f1')])
        
        # Modern Combobox stili
        style.configure('Modern.TCombobox',
                       fieldbackground='#ffffff',
                       background='#ffffff',
                       borderwidth=2,
                       relief='flat',
                       font=('Segoe UI', 11),
                       padding=(10, 8))
        style.map('Modern.TCombobox',
                 bordercolor=[('focus', '#3498db'), ('!focus', '#bdc3c7')],
                 lightcolor=[('focus', '#3498db'), ('!focus', '#ecf0f1')],
                 darkcolor=[('focus', '#3498db'), ('!focus', '#ecf0f1')])
        
        # Modern Treeview stili - Alternating colors
        style.configure('Modern.Treeview',
                       background='#ffffff',
                       fieldbackground='#ffffff',
                       font=('Segoe UI', 10),
                       rowheight=30,
                       borderwidth=0)
        style.configure('Modern.Treeview.Heading',
                       font=('Segoe UI', 11, 'bold'),
                       background='#34495e',
                       foreground='white',
                       relief='flat',
                       borderwidth=0,
                       padding=(10, 10))
        style.map('Modern.Treeview',
                 background=[('selected', '#3498db')],
                 foreground=[('selected', 'white')])
        style.map('Modern.Treeview.Heading',
                 background=[('active', '#2c3e50')])
        
        # Modern Notebook stili - Rounded tabs
        style.configure('Modern.TNotebook',
                       background='#ffffff',
                       borderwidth=0,
                       tabposition='nw',
                       tabmargins=[0, 0, 0, 0])
        style.configure('Modern.TNotebook.Tab',
                       background='#ecf0f1',
                       foreground='#2c3e50',
                       font=('Segoe UI', 11, 'bold'),
                       padding=[15, 10],
                       borderwidth=0,
                       relief='flat',
                       focuscolor='none')
        style.map('Modern.TNotebook.Tab',
                 background=[('selected', '#3498db'), ('active', '#bdc3c7')],
                 foreground=[('selected', 'white'), ('active', '#2c3e50')],
                 padding=[('selected', [15, 10]), ('active', [15, 10]), ('!selected', [15, 10])])
        
        # Modern Scrollbar stili - Minimalist
        style.configure('Modern.Vertical.TScrollbar',
                       background='#bdc3c7',
                       troughcolor='#ecf0f1',
                       width=8,
                       borderwidth=0,
                       relief='flat',
                       darkcolor='#bdc3c7',
                       lightcolor='#bdc3c7')
        style.map('Modern.Vertical.TScrollbar',
                 background=[('active', '#95a5a6'), ('pressed', '#7f8c8d')])
                 
        # Horizontal scrollbar
        style.configure('Modern.Horizontal.TScrollbar',
                       background='#bdc3c7',
                       troughcolor='#ecf0f1',
                       width=8,
                       borderwidth=0,
                       relief='flat')
        
        # Beyaz Frame stili
        style.configure('White.TFrame',
                       background='white',
                       borderwidth=0)
        
        # Başlık frame (başlık ve ayarlar butonu için) - Oval geçişli
        title_frame = ttk.Frame(self.main_frame, style='Modern.TFrame')
        title_frame.pack(fill=tk.X, pady=(0, 30))
        
        # Grid ağırlıklarını ayarla
        title_frame.grid_columnconfigure(0, weight=1)  # Sol boşluk
        title_frame.grid_columnconfigure(1, weight=0)  # Başlık (sabit genişlik)
        title_frame.grid_columnconfigure(2, weight=1)  # Sağ boşluk
        title_frame.grid_columnconfigure(3, weight=0)  # Saat (sabit genişlik)
        title_frame.grid_columnconfigure(4, weight=0)  # Ayarlar (sabit genişlik)
        
        # Başlık (ortada) - Modern gradient style
        self.title_label = ttk.Label(title_frame, text=self.app_title, 
                                    font=('Segoe UI', 20, 'bold'),
                                    foreground='#2c3e50',
                                    style='Modern.TLabel')
        self.title_label.grid(row=0, column=1, sticky='ew', pady=(10, 0))
        
        # Saat ve tarih (en sağ)
        self.clock_label = ttk.Label(title_frame, text="", 
                                    font=('Segoe UI', 12, 'bold'),
                                    foreground='#34495e',
                                    style='Modern.TLabel')
        self.clock_label.grid(row=0, column=3, padx=(0, 15), sticky='e')
        
        # Ayarlar butonu (en sağ) - Modern tasarım
        settings_button = tk.Button(title_frame, text="⚙️ Ayarlar", 
                                   command=self.show_settings_dialog,
                                   bg='#34495e', fg='white', font=('Segoe UI', 10, 'bold'),
                                   relief='flat', borderwidth=0, padx=20, pady=10,
                                   cursor='hand2', activebackground='#2c3e50', activeforeground='white')
        settings_button.grid(row=0, column=4, padx=(0, 0), sticky='e', pady=(10, 0))
        
        # Ayarlar butonu hover efekti - Smooth transition
        def settings_on_enter(e):
            settings_button.configure(bg='#2c3e50', relief='flat')
        
        def settings_on_leave(e):
            settings_button.configure(bg='#34495e', relief='flat')
        
        settings_button.bind('<Enter>', settings_on_enter)
        settings_button.bind('<Leave>', settings_on_leave)
        
        # Saat ve tarihi başlat
        self.start_clock()
        
        # Başlık konumunu ayarla
        self.update_title_position()
        
        # Başlık metnini güncelle
        self.title_label.config(text=self.app_title)
        self.root.title(self.app_title)
        
        # Notebook container frame - Tab'ları sola çekmek için
        notebook_container = ttk.Frame(self.main_frame, style='Modern.TFrame')
        notebook_container.pack(fill=tk.BOTH, expand=True, pady=(20, 15))
        
        # Notebook (sekmeli arayüz) - Modern card style
        self.notebook = ttk.Notebook(notebook_container, style='Modern.TNotebook')
        self.notebook.pack(fill=tk.BOTH, expand=True, anchor='w')
        
        # Müşteriler sekmesi
        self.customers_frame = ttk.Frame(self.notebook, style='Modern.TFrame')
        self.notebook.add(self.customers_frame, text="Danışan Kayıt")
        
        # Siparişler sekmesi
        self.orders_frame = ttk.Frame(self.notebook, style='Modern.TFrame')
        self.notebook.add(self.orders_frame, text="Siparişler")
        
        # Gelir ve Danışanlar sekmesi
        self.stats_frame = ttk.Frame(self.notebook, style='Modern.TFrame')
        self.notebook.add(self.stats_frame, text="Gelir ve Danışanlar")
        
        # Bildirimler sekmesi
        self.notifications_frame = ttk.Frame(self.notebook, style='White.TFrame')
        self.notebook.add(self.notifications_frame, text="Ödeme Kontrol")
        
        self.setup_customers_tab()
        self.setup_orders_tab()
        self.setup_stats_tab()
        self.setup_notifications_tab()
        
        # Notebook tab değişim event'ini bağla
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
        
        # İlk yükleme
        self.load_customers()
        self.load_stats()
        
        # Sistem başlatıldı
        
        # Cache temizleme fonksiyonu
        def clear_cache():
            self.customer_cache = {}
            self.order_cache = {}
            self.cache_timestamp = datetime.now()
        
        self.clear_cache = clear_cache
    
    def refresh_orders_manual(self):
        """Manuel sipariş yenileme - cache'i bypass eder"""
        self.clear_cache()
        self.load_orders()
    
    def on_tab_changed(self, event):
        """Sekme değiştiğinde çağrılır - otomatik yenileme için"""
        selected_tab = event.widget.select()
        tab_name = event.widget.tab(selected_tab, "text")
        
        # Siparişler sekmesine geçildiğinde otomatik yenile
        if tab_name == "Siparişler":
            self.load_orders()
        # Ödeme Kontrol sekmesine geçildiğinde otomatik yenile
        elif tab_name == "Ödeme Kontrol":
            self.check_notifications()
        # Gelir ve Danışanlar sekmesine geçildiğinde otomatik yenile
        elif tab_name == "Gelir ve Danışanlar":
            self.load_stats()
    
    def show_settings_dialog(self):
        """Modern ayarlar dialog'unu gösterir"""
        # Modern şifre kontrolü penceresi
        password_dialog = tk.Toplevel(self.root)
        password_dialog.title("🔐 Güvenlik Kontrolü")
        password_dialog.geometry("550x450")
        password_dialog.transient(self.root)
        password_dialog.grab_set()
        password_dialog.configure(bg='#f8f9fa')
        password_dialog.resizable(False, False)
        
        # Pencereyi merkeze yerleştir
        password_dialog.update_idletasks()
        x = (password_dialog.winfo_screenwidth() // 2) - (550 // 2)
        y = (password_dialog.winfo_screenheight() // 2) - (450 // 2)
        password_dialog.geometry(f"550x450+{x}+{y}")
        
        # Ana container - Modern card tasarımı
        main_container = tk.Frame(password_dialog, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Header card
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 25))
        
        # Modern gölge efekti için border frame
        shadow_frame = tk.Frame(main_container, bg='#e0e0e0', height=2)
        shadow_frame.pack(fill=tk.X, pady=(0, 25))
        
        # İkon ve başlık
        icon_frame = tk.Frame(header_card, bg='white')
        icon_frame.pack(fill=tk.X, padx=25, pady=20)
        
        tk.Label(icon_frame, text="🔐", font=('Segoe UI', 32), bg='white', fg='#495057').pack()
        
        tk.Label(icon_frame, 
                text="Ayarlar Erişimi",
                font=('Segoe UI', 20, 'bold'),
                bg='white',
                fg='#2c3e50').pack(pady=(10, 5))
        
        tk.Label(icon_frame,
                text="Ayarlara erişmek için lütfen şifrenizi girin",
                font=('Segoe UI', 11),
                bg='white',
                fg='#7f8c8d').pack()
        
        # Şifre girişi kartı
        input_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        input_card.pack(fill=tk.X, pady=(0, 25))
        
        input_inner = tk.Frame(input_card, bg='white')
        input_inner.pack(fill=tk.X, padx=25, pady=20)
        
        tk.Label(input_inner, text="Şifre:", font=('Segoe UI', 12, 'bold'), 
                bg='white', fg='#495057').pack(anchor='w', pady=(0, 8))
        
        password_var = tk.StringVar()
        
        # Modern şifre girişi frame
        entry_frame = tk.Frame(input_inner, bg='white')
        entry_frame.pack(fill=tk.X, pady=(0, 15))
        
        password_entry = tk.Entry(entry_frame, textvariable=password_var, show="*", 
                                 font=('Segoe UI', 12), relief='solid', bd=1,
                                 highlightthickness=1, highlightcolor='#007bff',
                                 highlightbackground='#ced4da', bg='#ffffff', fg='#495057')
        password_entry.pack(fill=tk.X, ipady=8, ipadx=10)
        password_entry.focus()
        
        # Hata mesajı için label
        error_label = tk.Label(input_inner, text="", font=('Segoe UI', 10), 
                              bg='white', fg='#dc3545')
        error_label.pack(anchor='w')
        
        def check_password():
            if password_var.get() == self.settings_password:
                password_dialog.destroy()
                self.open_settings()
            else:
                error_label.config(text="❌ Yanlış şifre! Lütfen tekrar deneyin.")
                password_var.set("")
                password_entry.focus()
                # Hata mesajını 3 saniye sonra temizle
                password_dialog.after(3000, lambda: error_label.config(text=""))
        
        def on_enter(event):
            check_password()
        
        password_entry.bind('<Return>', on_enter)
        
        # Butonlar kartı
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X)
        
        button_inner = tk.Frame(button_card, bg='white')
        button_inner.pack(fill=tk.X, padx=25, pady=20)
        
        button_container = tk.Frame(button_inner, bg='white')
        button_container.pack()
        
        # Modern buton stilleri
        login_btn = tk.Button(button_container, text="🔓 Giriş Yap", command=check_password,
                             bg='#007bff', fg='white', font=('Segoe UI', 11, 'bold'),
                             relief='flat', borderwidth=0, padx=25, pady=12,
                             cursor='hand2', activebackground='#0056b3', activeforeground='white')
        login_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        cancel_btn = tk.Button(button_container, text="❌ İptal", command=password_dialog.destroy,
                              bg='#6c757d', fg='white', font=('Segoe UI', 11, 'bold'),
                              relief='flat', borderwidth=0, padx=25, pady=12,
                              cursor='hand2', activebackground='#545b62', activeforeground='white')
        cancel_btn.pack(side=tk.LEFT)
        
        # Buton hover efektleri
        def on_login_enter(e):
            login_btn.configure(bg='#0056b3')
        def on_login_leave(e):
            login_btn.configure(bg='#007bff')
        def on_cancel_enter(e):
            cancel_btn.configure(bg='#545b62')
        def on_cancel_leave(e):
            cancel_btn.configure(bg='#6c757d')
        
        login_btn.bind('<Enter>', on_login_enter)
        login_btn.bind('<Leave>', on_login_leave)
        cancel_btn.bind('<Enter>', on_cancel_enter)
        cancel_btn.bind('<Leave>', on_cancel_leave)
    
    def open_settings(self):
        """Modern ayarlar penceresini açar"""
        settings_dialog = tk.Toplevel(self.root)
        settings_dialog.title("⚙️ Uygulama Ayarları")
        settings_dialog.geometry("600x500")
        settings_dialog.transient(self.root)
        settings_dialog.grab_set()
        settings_dialog.configure(bg='#f8f9fa')
        settings_dialog.resizable(False, False)
        
        # Pencereyi merkeze yerleştir
        settings_dialog.update_idletasks()
        x = (settings_dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (settings_dialog.winfo_screenheight() // 2) - (500 // 2)
        settings_dialog.geometry(f"600x500+{x}+{y}")
        
        # Ana container - Modern card tasarımı
        main_container = tk.Frame(settings_dialog, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Header card
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 25))
        
        # Modern gölge efekti
        shadow_frame = tk.Frame(main_container, bg='#e0e0e0', height=2)
        shadow_frame.pack(fill=tk.X, pady=(0, 25))
        
        # Başlık
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=25, pady=20)
        
        tk.Label(header_inner, text="⚙️", font=('Segoe UI', 28), bg='white', fg='#495057').pack()
        tk.Label(header_inner, 
                text="Uygulama Ayarları",
                font=('Segoe UI', 18, 'bold'),
                bg='white',
                fg='#2c3e50').pack(pady=(10, 5))
        tk.Label(header_inner,
                text="Uygulamanızı kişiselleştirin",
                font=('Segoe UI', 11),
                bg='white',
                fg='#7f8c8d').pack()
        
        # Ayarlar kartı
        settings_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        settings_card.pack(fill=tk.BOTH, expand=True, pady=(0, 25))
        
        settings_inner = tk.Frame(settings_card, bg='white')
        settings_inner.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # Uygulama başlığı ayarı
        title_section = tk.Frame(settings_inner, bg='white')
        title_section.pack(fill=tk.X, pady=(0, 25))
        
        tk.Label(title_section, text="📝 Uygulama Başlığı", 
                font=('Segoe UI', 14, 'bold'), bg='white', fg='#495057').pack(anchor='w', pady=(0, 10))
        
        title_var = tk.StringVar(value=self.app_title)
        title_entry = tk.Entry(title_section, textvariable=title_var, 
                              font=('Segoe UI', 12), relief='solid', bd=1,
                              highlightthickness=1, highlightcolor='#007bff',
                              highlightbackground='#ced4da', bg='#ffffff', fg='#495057')
        title_entry.pack(fill=tk.X, ipady=8, ipadx=10, pady=(0, 5))
        
        tk.Label(title_section, text="Uygulamanın üst kısmında görünen başlık metni", 
                font=('Segoe UI', 9), bg='white', fg='#6c757d').pack(anchor='w')
        
        # Başlık konumu ayarı
        position_section = tk.Frame(settings_inner, bg='white')
        position_section.pack(fill=tk.X, pady=(0, 25))
        
        tk.Label(position_section, text="📍 Başlık Konumu", 
                font=('Segoe UI', 14, 'bold'), bg='white', fg='#495057').pack(anchor='w', pady=(0, 15))
        
        position_var = tk.StringVar(value=self.title_position)
        
        def update_position():
            self.title_position = position_var.get()
            self.update_title_position()
        
        # Modern radio button container
        radio_container = tk.Frame(position_section, bg='white')
        radio_container.pack(fill=tk.X)
        
        # Custom modern radio buttons
        radio_options = [
            ("📍 Sol", "left"),
            ("📍 Orta", "center"), 
            ("📍 Sağ", "right")
        ]
        
        for i, (text, value) in enumerate(radio_options):
            radio_frame = tk.Frame(radio_container, bg='white')
            radio_frame.pack(side=tk.LEFT, padx=(0, 20) if i < 2 else (0, 0))
            
            radio_btn = tk.Radiobutton(radio_frame, text=text, variable=position_var, value=value,
                                      command=update_position, font=('Segoe UI', 11),
                                      bg='white', fg='#495057', selectcolor='#007bff',
                                      activebackground='white', activeforeground='#007bff',
                                      relief='flat', bd=0)
            radio_btn.pack()
        
        # Butonlar kartı
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X)
        
        button_inner = tk.Frame(button_card, bg='white')
        button_inner.pack(fill=tk.X, padx=25, pady=20)
        
        button_container = tk.Frame(button_inner, bg='white')
        button_container.pack()
        
        # Kaydet fonksiyonu
        def save_settings():
            # Başlığı güncelle
            new_title = title_var.get().strip()
            if new_title:
                self.app_title = new_title
                self.title_label.config(text=self.app_title)
                self.root.title(self.app_title)
            
            # Başlık konumunu güncelle
            self.title_position = position_var.get()
            self.update_title_position()
            
            # Ayarları dosyaya kaydet
            self.save_settings_file()
            
            # Otomatik kayıt
            self.auto_save_data("Ayarlar Güncellendi")
            
            # Modern başarı mesajı
            success_msg = tk.Toplevel(settings_dialog)
            success_msg.title("✅ Başarılı")
            success_msg.geometry("300x150")
            success_msg.configure(bg='#d4edda')
            success_msg.transient(settings_dialog)
            success_msg.grab_set()
            
            # Merkeze yerleştir
            success_msg.update_idletasks()
            x = (success_msg.winfo_screenwidth() // 2) - (150 // 2)
            y = (success_msg.winfo_screenheight() // 2) - (75 // 2)
            success_msg.geometry(f"300x150+{x}+{y}")
            
            tk.Label(success_msg, text="✅", font=('Segoe UI', 24), bg='#d4edda', fg='#155724').pack(pady=(20, 5))
            tk.Label(success_msg, text="Ayarlar başarıyla kaydedildi!", 
                    font=('Segoe UI', 12, 'bold'), bg='#d4edda', fg='#155724').pack()
            
            def close_all():
                success_msg.destroy()
                settings_dialog.destroy()
            
            success_msg.after(2000, close_all)
        
        # Modern butonlar
        save_btn = tk.Button(button_container, text="💾 Kaydet", command=save_settings,
                            bg='#28a745', fg='white', font=('Segoe UI', 11, 'bold'),
                            relief='flat', borderwidth=0, padx=25, pady=12,
                            cursor='hand2', activebackground='#1e7e34', activeforeground='white')
        save_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        cancel_btn = tk.Button(button_container, text="❌ İptal", command=settings_dialog.destroy,
                              bg='#6c757d', fg='white', font=('Segoe UI', 11, 'bold'),
                              relief='flat', borderwidth=0, padx=25, pady=12,
                              cursor='hand2', activebackground='#545b62', activeforeground='white')
        cancel_btn.pack(side=tk.LEFT)
        
        # Buton hover efektleri
        def on_save_enter(e):
            save_btn.configure(bg='#1e7e34')
        def on_save_leave(e):
            save_btn.configure(bg='#28a745')
        def on_cancel_enter(e):
            cancel_btn.configure(bg='#545b62')
        def on_cancel_leave(e):
            cancel_btn.configure(bg='#6c757d')
        
        save_btn.bind('<Enter>', on_save_enter)
        save_btn.bind('<Leave>', on_save_leave)
        cancel_btn.bind('<Enter>', on_cancel_enter)
        cancel_btn.bind('<Leave>', on_cancel_leave)
        
        # save_settings_file fonksiyonunu tanımla (orijinal save_settings ile çakışmayı önlemek için)
        def save_settings_file():
            try:
                settings = {
                    'app_title': self.app_title,
                    'title_position': self.title_position
                }
                with open('settings.json', 'w', encoding='utf-8') as f:
                    import json
                    json.dump(settings, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"Ayarlar kaydedilirken hata: {e}")
        
        self.save_settings_file = save_settings_file
        
        def delete_all_data():
            """Tüm verileri silme işlemi"""
            # Şifre kontrolü
            password_dialog = tk.Toplevel(settings_dialog)
            password_dialog.title("Şifre Gerekli")
            password_dialog.geometry("300x150")
            password_dialog.transient(settings_dialog)
            password_dialog.grab_set()
            
            # Pencereyi merkeze yerleştir
            password_dialog.update_idletasks()
            x = (password_dialog.winfo_screenwidth() // 2) - (300 // 2)
            y = (password_dialog.winfo_screenheight() // 2) - (150 // 2)
            password_dialog.geometry(f"300x150+{x}+{y}")
            
            # Ana frame
            pwd_frame = ttk.Frame(password_dialog)
            pwd_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            ttk.Label(pwd_frame, text="Şifre Girin:", font=('Arial', 10, 'bold')).pack(pady=(0, 10))
            
            pwd_var = tk.StringVar()
            pwd_entry = ttk.Entry(pwd_frame, textvariable=pwd_var, show="*", width=30)
            pwd_entry.pack(pady=(0, 15))
            pwd_entry.focus()
            
            def check_delete_password():
                if pwd_var.get() == "11235":
                    password_dialog.destroy()
                    
                    # Geri alınamaz uyarısı
                    result = messagebox.askyesno(
                        "⚠️ UYARI", 
                        "TÜM VERİLER SİLİNECEK!\n\n"
                        "Bu işlem geri alınamaz!\n"
                        "Tüm müşteriler ve siparişler kalıcı olarak silinecek.\n\n"
                        "Devam etmek istiyor musunuz?",
                        icon='warning'
                    )
                    
                    if result:
                        try:
                            # Veritabanını sıfırla
                            self.db.reset_database()
                            
                            # Listeleri temizle
                            self.load_customers()
                            self.load_orders()
                            self.load_stats()
                            self.load_notifications()
                            
                            # Otomatik kayıt
                            self.auto_save_data("Tüm Veriler Silindi")
                            
                            messagebox.showinfo(
                                "Başarılı", 
                                "Tüm veriler başarıyla silindi!\n"
                                "Program yeniden başlatılacak."
                            )
                            
                            # Programı yeniden başlat
                            self.root.after(2000, self.root.quit)
                            
                        except Exception as e:
                            messagebox.showerror(
                                "Hata", 
                                f"Veriler silinirken hata oluştu:\n{str(e)}"
                            )
                else:
                    messagebox.showerror("Hata", "Yanlış şifre!")
                    pwd_var.set("")
                    pwd_entry.focus()
            
            def on_enter(event):
                check_delete_password()
            
            pwd_entry.bind('<Return>', on_enter)
            
            # Butonlar
            pwd_button_frame = ttk.Frame(pwd_frame)
            pwd_button_frame.pack(side=tk.BOTTOM, pady=(10, 0))
            
            ttk.Button(pwd_button_frame, text="Onayla", command=check_delete_password).pack(side=tk.LEFT, padx=5)
            ttk.Button(pwd_button_frame, text="İptal", command=password_dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # Butonlar
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side=tk.BOTTOM, pady=(20, 0))
        
        ttk.Button(button_frame, text="Kaydet", command=save_settings).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Verileri Sil", command=delete_all_data, 
                  style="Danger.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="İptal", command=settings_dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def update_title_position(self):
        """Başlık konumunu günceller"""
        title_frame = self.title_label.master
        
        if self.title_position == "left":
            # Başlık sola
            self.title_label.grid(row=0, column=0, sticky='w', padx=(10, 0))
            # Grid ağırlıklarını ayarla
            title_frame.grid_columnconfigure(0, weight=0)  # Başlık (sabit genişlik)
            title_frame.grid_columnconfigure(1, weight=1)  # Orta boşluk
            title_frame.grid_columnconfigure(2, weight=1)  # Sağ boşluk
            title_frame.grid_columnconfigure(3, weight=0)  # Saat (sabit genişlik)
            title_frame.grid_columnconfigure(4, weight=0)  # Ayarlar (sabit genişlik)
        elif self.title_position == "right":
            # Başlık sağa
            self.title_label.grid(row=0, column=2, sticky='e', padx=(0, 10))
            # Grid ağırlıklarını ayarla
            title_frame.grid_columnconfigure(0, weight=1)  # Sol boşluk
            title_frame.grid_columnconfigure(1, weight=1)  # Orta boşluk
            title_frame.grid_columnconfigure(2, weight=0)  # Başlık (sabit genişlik)
            title_frame.grid_columnconfigure(3, weight=0)  # Saat (sabit genişlik)
            title_frame.grid_columnconfigure(4, weight=0)  # Ayarlar (sabit genişlik)
        else:  # center
            # Başlık ortada
            self.title_label.grid(row=0, column=1, sticky='ew')
            # Grid ağırlıklarını ayarla
            title_frame.grid_columnconfigure(0, weight=1)  # Sol boşluk
            title_frame.grid_columnconfigure(1, weight=0)  # Başlık (sabit genişlik)
            title_frame.grid_columnconfigure(2, weight=1)  # Sağ boşluk
            title_frame.grid_columnconfigure(3, weight=0)  # Saat (sabit genişlik)
            title_frame.grid_columnconfigure(4, weight=0)  # Ayarlar (sabit genişlik)
    
    def start_clock(self):
        """Saat ve tarihi başlatır"""
        self.clock_running = True
        self.update_clock()
    
    def stop_clock(self):
        """Saat ve tarihi durdurur"""
        self.clock_running = False
    
    def update_clock(self):
        """Saat ve tarihi günceller"""
        if self.clock_running and self.clock_label:
            try:
                from datetime import datetime
                now = datetime.now()
                # Türkçe tarih formatı
                date_str = now.strftime("%d.%m.%Y")
                time_str = now.strftime("%H:%M:%S")
                clock_text = f"{time_str}\n{date_str}"
                self.clock_label.config(text=clock_text)
            except Exception as e:
                print(f"Saat güncelleme hatası: {e}")
            
            # 1 saniye sonra tekrar güncelle
            self.root.after(1000, self.update_clock)
    
    def load_settings(self):
        """Ayarları dosyadan yükler"""
        try:
            if os.path.exists('settings.txt'):
                with open('settings.txt', 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    for line in lines:
                        if line.startswith('title='):
                            self.app_title = line.split('=', 1)[1].strip()
                        elif line.startswith('position='):
                            self.title_position = line.split('=', 1)[1].strip()
        except Exception as e:
            print(f"Ayarlar yüklenirken hata: {e}")
    
    def save_settings(self):
        """Ayarları dosyaya kaydeder"""
        try:
            with open('settings.txt', 'w', encoding='utf-8') as f:
                f.write(f'title={self.app_title}\n')
                f.write(f'position={self.title_position}\n')
        except Exception as e:
            print(f"Ayarlar kaydedilirken hata: {e}")
    
    def setup_customers_tab(self):
        """Müşteriler sekmesini hazırlar"""
        # Üst frame - Arama ve butonlar (Oval geçişli)
        top_frame = ttk.Frame(self.customers_frame, style='Modern.TFrame')
        top_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # Arama
        ttk.Label(top_frame, text="🔍 Arama:", style='Modern.TLabel').pack(side=tk.LEFT, padx=(0, 8))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(top_frame, textvariable=self.search_var, width=35, style='Modern.TEntry')
        self.search_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.search_entry.bind('<KeyRelease>', self.search_customers)
        
        # Butonlar frame - Grid layout ile düzenli hizalama
        buttons_frame = ttk.Frame(top_frame, style='Modern.TFrame')
        buttons_frame.pack(side=tk.LEFT, padx=(0, 0))
        
        # Butonlar - Modern tasarım ve standart boyutlar
        button_style = {
            'font': ('Segoe UI', 9, 'bold'),
            'relief': 'flat',
            'borderwidth': 0,
            'padx': 15,
            'pady': 8,
            'cursor': 'hand2',
            'activebackground': '#1a252f',
            'activeforeground': 'white'
        }
        
        new_customer_btn = tk.Button(buttons_frame, text="➕ Yeni Müşteri", command=self.add_customer_dialog,
                                    bg='#2c3e50', fg='white', **button_style)
        new_customer_btn.grid(row=0, column=0, padx=3, pady=2, sticky='ew')
        
        edit_btn = tk.Button(buttons_frame, text="✏️ Düzenle", command=self.edit_customer,
                            bg='#34495e', fg='white', **button_style)
        edit_btn.grid(row=0, column=1, padx=3, pady=2, sticky='ew')
        
        delete_btn = tk.Button(buttons_frame, text="🗑️ Sil", command=self.delete_customer,
                              bg='#e74c3c', fg='white', **button_style)
        delete_btn.grid(row=0, column=2, padx=3, pady=2, sticky='ew')
        
        refresh_btn = tk.Button(buttons_frame, text="🔄 Yenile", command=self.load_customers,
                               bg='#3498db', fg='white', **button_style)
        refresh_btn.grid(row=0, column=3, padx=3, pady=2, sticky='ew')
        
        excel_export_btn = tk.Button(buttons_frame, text="📊 Excele Aktar", command=self.show_excel_export_dialog,
                                    bg='#27ae60', fg='white', **button_style)
        excel_export_btn.grid(row=0, column=4, padx=3, pady=2, sticky='ew')
        
        # Grid column weights için eşit genişlik
        for i in range(5):
            buttons_frame.grid_columnconfigure(i, weight=1)
        
        # Hover efektleri - Koyu renkler
        def on_enter(e):
            if e.widget == new_customer_btn:
                e.widget['bg'] = '#1a252f'
            elif e.widget == edit_btn:
                e.widget['bg'] = '#2c3e50'
            elif e.widget == delete_btn:
                e.widget['bg'] = '#c0392b'
            elif e.widget == refresh_btn:
                e.widget['bg'] = '#2980b9'
            elif e.widget == excel_export_btn:
                e.widget['bg'] = '#229954'
        
        def on_leave(e):
            if e.widget == new_customer_btn:
                e.widget['bg'] = '#2c3e50'
            elif e.widget == edit_btn:
                e.widget['bg'] = '#34495e'
            elif e.widget == delete_btn:
                e.widget['bg'] = '#e74c3c'
            elif e.widget == refresh_btn:
                e.widget['bg'] = '#3498db'
            elif e.widget == excel_export_btn:
                e.widget['bg'] = '#27ae60'
        
        for btn in [new_customer_btn, edit_btn, delete_btn, refresh_btn, excel_export_btn]:
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)
        
        # Müşteri listesi (Oval geçişli)
        list_frame = ttk.Frame(self.customers_frame, style='Modern.TFrame')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Treeview
        columns = ('ID', 'Ad', 'E-posta', 'Telefon', 'Şirket', 'Kayıt Tarihi')
        self.customer_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15, style='Modern.Treeview')
        
        # Treeview satır stilleri
        self.customer_tree.tag_configure('oddrow', background='#f8f9fa')
        self.customer_tree.tag_configure('evenrow', background='#ffffff')
        self.customer_tree.tag_configure('selected', background='#e3f2fd')
        
        # Sütun başlıkları - sola hizalı
        self.customer_tree.heading('ID', text='ID', anchor='w')
        self.customer_tree.heading('Ad', text='Ad Soyad', anchor='w')
        self.customer_tree.heading('E-posta', text='E-posta', anchor='w')
        self.customer_tree.heading('Telefon', text='Telefon', anchor='w')
        self.customer_tree.heading('Şirket', text='Şirket', anchor='w')
        self.customer_tree.heading('Kayıt Tarihi', text='Kayıt Tarihi', anchor='w')
        
        # Sütun genişlikleri ve hizalama
        self.customer_tree.column('ID', width=0, stretch=False, anchor='w')  # ID sütununu gizle
        self.customer_tree.column('Ad', width=150, anchor='w')
        self.customer_tree.column('E-posta', width=200, anchor='w')
        self.customer_tree.column('Telefon', width=120, anchor='w')
        self.customer_tree.column('Şirket', width=150, anchor='w')
        self.customer_tree.column('Kayıt Tarihi', width=150, anchor='w')
        
        # Modern custom scrollbar for table
        table_scrollbar_frame = tk.Frame(list_frame, bg='#f8f9fa', width=8)
        table_scrollbar_bg = tk.Canvas(table_scrollbar_frame, bg='#f8f9fa', width=8, highlightthickness=0, bd=0)
        
        # Modern scrollbar functions
        def update_table_scrollbar():
            table_scrollbar_bg.delete("thumb")
            try:
                # Get treeview scroll position
                top, bottom = self.customer_tree.yview()
                if bottom - top >= 1.0:
                    return
                    
                scrollbar_height = table_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                table_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def table_smooth_scroll(event):
            try:
                if self.customer_tree.winfo_exists():
                    self.customer_tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    update_table_scrollbar()
            except tk.TclError:
                pass
        
        # Scrollbar drag variables
        self.table_scrollbar_dragging = False
        self.table_drag_start_y = 0
        
        def on_table_scrollbar_click(event):
            try:
                top, bottom = self.customer_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = table_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.customer_tree.yview_moveto(click_position)
                update_table_scrollbar()
                
                # Start drag
                self.table_scrollbar_dragging = True
                self.table_drag_start_y = event.y
            except:
                pass
        
        def on_table_scrollbar_drag(event):
            if not self.table_scrollbar_dragging:
                return
            try:
                top, bottom = self.customer_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = table_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.customer_tree.yview_moveto(click_position)
                update_table_scrollbar()
            except:
                pass
        
        def on_table_scrollbar_release(event):
            self.table_scrollbar_dragging = False
        
        def on_table_scrollbar_enter(event):
            table_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.customer_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = table_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                table_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_table_scrollbar_leave(event):
            update_table_scrollbar()
        
        # Configure treeview scrolling
        def on_tree_scroll(*args):
            update_table_scrollbar()
        
        self.customer_tree.configure(yscrollcommand=on_tree_scroll)
        
        # Bind events
        self.customer_tree.bind("<MouseWheel>", table_smooth_scroll)
        table_scrollbar_bg.bind("<Button-1>", on_table_scrollbar_click)
        table_scrollbar_bg.bind("<B1-Motion>", on_table_scrollbar_drag)
        table_scrollbar_bg.bind("<ButtonRelease-1>", on_table_scrollbar_release)
        table_scrollbar_bg.bind("<Enter>", on_table_scrollbar_enter)
        table_scrollbar_bg.bind("<Leave>", on_table_scrollbar_leave)
        
        # Make scrollbar focusable for better interaction
        table_scrollbar_bg.configure(takefocus=True)
        
        # Pack widgets
        self.customer_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        table_scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=0)
        table_scrollbar_frame.pack_propagate(False)
        table_scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize scrollbar
        self.customer_tree.after(200, update_table_scrollbar)
        
        # Çift tıklama olayı
        self.customer_tree.bind('<Double-1>', self.on_customer_double_click)
    
    def setup_orders_tab(self):
        """Siparişler sekmesini hazırlar"""
        # Üst frame - Arama ve butonlar (Oval geçişli)
        top_frame = ttk.Frame(self.orders_frame, style='Modern.TFrame')
        top_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # Arama
        ttk.Label(top_frame, text="🔍 Arama:", style='Modern.TLabel').pack(side=tk.LEFT, padx=(0, 8))
        self.order_search_var = tk.StringVar()
        self.order_search_entry = ttk.Entry(top_frame, textvariable=self.order_search_var, width=35, style='Modern.TEntry')
        self.order_search_entry.pack(side=tk.LEFT, padx=(0, 5))
        self.order_search_entry.bind('<KeyRelease>', self.search_orders)
        
        # Butonlar frame - Grid layout ile düzenli hizalama
        buttons_frame = ttk.Frame(top_frame, style='Modern.TFrame')
        buttons_frame.pack(side=tk.LEFT, padx=(0, 0))
        
        # Butonlar - Modern tasarım ve standart boyutlar
        button_style = {
            'font': ('Segoe UI', 9, 'bold'),
            'relief': 'flat',
            'borderwidth': 0,
            'padx': 15,
            'pady': 8,
            'cursor': 'hand2',
            'activebackground': '#1a252f',
            'activeforeground': 'white'
        }
        
        new_order_btn = tk.Button(buttons_frame, text="➕ Yeni Sipariş", command=self.add_order_from_orders_tab,
                                 bg='#1a252f', fg='white', **button_style)
        new_order_btn.grid(row=0, column=0, padx=3, pady=2, sticky='ew')
        
        edit_order_btn = tk.Button(buttons_frame, text="✏️ Siparişi Düzenle", command=self.edit_order,
                                 bg='#2c3e50', fg='white', **button_style)
        edit_order_btn.grid(row=0, column=1, padx=3, pady=2, sticky='ew')
        
        details_btn = tk.Button(buttons_frame, text="👁️ Sipariş Detayları", command=self.show_order_details,
                               bg='#e74c3c', fg='white', **button_style)
        details_btn.grid(row=0, column=2, padx=3, pady=2, sticky='ew')
        
        refresh_orders_btn = tk.Button(buttons_frame, text="↻ Siparişleri Yenile", command=self.refresh_orders_manual,
                                     bg='#2980b9', fg='white', **button_style)
        refresh_orders_btn.grid(row=0, column=3, padx=3, pady=2, sticky='ew')
        
        # Grid column weights için eşit genişlik
        for i in range(4):
            buttons_frame.grid_columnconfigure(i, weight=1)
        
        # Hover efektleri - Koyu renkler
        def on_enter(e):
            if e.widget == new_order_btn:
                e.widget['bg'] = '#1a252f'
            elif e.widget == edit_order_btn:
                e.widget['bg'] = '#2c3e50'
            elif e.widget == details_btn:
                e.widget['bg'] = '#c0392b'
            elif e.widget == refresh_orders_btn:
                e.widget['bg'] = '#2980b9'
        
        def on_leave(e):
            if e.widget == new_order_btn:
                e.widget['bg'] = '#1a252f'
            elif e.widget == edit_order_btn:
                e.widget['bg'] = '#2c3e50'
            elif e.widget == details_btn:
                e.widget['bg'] = '#e74c3c'
            elif e.widget == refresh_orders_btn:
                e.widget['bg'] = '#2980b9'
        
        for btn in [new_order_btn, edit_order_btn, details_btn, refresh_orders_btn]:
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)
        # ttk.Button(top_frame, text="Gecikmiş Siparişleri Tamamla", command=self.complete_overdue_orders, 
        #           style="Accent.TButton").pack(side=tk.LEFT, padx=5)
        
        # Sipariş listesi (Oval geçişli)
        list_frame = ttk.Frame(self.orders_frame, style='Modern.TFrame')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Treeview
        columns = ('Müşteri', 'Ürün', 'Miktar', 'Fiyat', 'Toplam', 'Başlangıç', 'Bitiş', 'Durum', 'Tarih')
        self.order_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15, style='Modern.Treeview')
        
        # Treeview satır stilleri
        self.order_tree.tag_configure('oddrow', background='#f8f9fa')
        self.order_tree.tag_configure('evenrow', background='#ffffff')
        self.order_tree.tag_configure('selected', background='#e3f2fd')
        
        # Sütun başlıkları - sola hizalı
        self.order_tree.heading('Müşteri', text='Müşteri', anchor='w')
        self.order_tree.heading('Ürün', text='Ürün Adı', anchor='w')
        self.order_tree.heading('Miktar', text='Miktar', anchor='w')
        self.order_tree.heading('Fiyat', text='Birim Fiyat', anchor='w')
        self.order_tree.heading('Toplam', text='Toplam', anchor='w')
        self.order_tree.heading('Başlangıç', text='Başlangıç', anchor='w')
        self.order_tree.heading('Bitiş', text='Bitiş', anchor='w')
        self.order_tree.heading('Durum', text='Durum', anchor='w')
        self.order_tree.heading('Tarih', text='Sipariş Tarihi', anchor='w')
        
        # Sütun genişlikleri ve hizalama
        self.order_tree.column('Müşteri', width=120, anchor='w')
        self.order_tree.column('Ürün', width=150, anchor='w')
        self.order_tree.column('Miktar', width=60, anchor='w')
        self.order_tree.column('Fiyat', width=80, anchor='w')
        self.order_tree.column('Toplam', width=80, anchor='w')
        self.order_tree.column('Başlangıç', width=80, anchor='w')
        self.order_tree.column('Bitiş', width=80, anchor='w')
        self.order_tree.column('Durum', width=80, anchor='w')
        self.order_tree.column('Tarih', width=120, anchor='w')
        
        # Modern custom scrollbar for orders table
        orders_scrollbar_frame = tk.Frame(list_frame, bg='#f8f9fa', width=8)
        orders_scrollbar_bg = tk.Canvas(orders_scrollbar_frame, bg='#f8f9fa', width=8, highlightthickness=0, bd=0)
        
        # Modern scrollbar functions for orders
        def update_orders_scrollbar():
            orders_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.order_tree.yview()
                if bottom - top >= 1.0:
                    return
                    
                scrollbar_height = orders_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                orders_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        # Scrollbar drag variables for orders
        self.orders_scrollbar_dragging = False
        
        def orders_smooth_scroll(event):
            try:
                if self.order_tree.winfo_exists():
                    self.order_tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    update_orders_scrollbar()
            except tk.TclError:
                pass
        
        def on_orders_scrollbar_click(event):
            try:
                top, bottom = self.order_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = orders_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.order_tree.yview_moveto(click_position)
                update_orders_scrollbar()
                self.orders_scrollbar_dragging = True
            except:
                pass
        
        def on_orders_scrollbar_drag(event):
            if not self.orders_scrollbar_dragging:
                return
            try:
                top, bottom = self.order_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = orders_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.order_tree.yview_moveto(click_position)
                update_orders_scrollbar()
            except:
                pass
        
        def on_orders_scrollbar_release(event):
            self.orders_scrollbar_dragging = False
        
        def on_orders_scrollbar_enter(event):
            orders_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.order_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = orders_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                orders_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_orders_scrollbar_leave(event):
            update_orders_scrollbar()
        
        def on_orders_tree_scroll(*args):
            update_orders_scrollbar()
        
        self.order_tree.configure(yscrollcommand=on_orders_tree_scroll)
        
        # Bind events for orders
        self.order_tree.bind("<MouseWheel>", orders_smooth_scroll)
        orders_scrollbar_bg.bind("<Button-1>", on_orders_scrollbar_click)
        orders_scrollbar_bg.bind("<B1-Motion>", on_orders_scrollbar_drag)
        orders_scrollbar_bg.bind("<ButtonRelease-1>", on_orders_scrollbar_release)
        orders_scrollbar_bg.bind("<Enter>", on_orders_scrollbar_enter)
        orders_scrollbar_bg.bind("<Leave>", on_orders_scrollbar_leave)
        orders_scrollbar_bg.configure(takefocus=True)
        
        # Pack widgets for orders
        self.order_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        orders_scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=0)
        orders_scrollbar_frame.pack_propagate(False)
        orders_scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize orders scrollbar
        self.order_tree.after(200, update_orders_scrollbar)
        
        # Çift tıklama eventi - Sipariş düzenleme penceresi açar
        self.order_tree.bind("<Double-1>", self.on_order_double_click)
    
    def on_order_double_click(self, event):
        """Sipariş listesinde çift tıklama olayı - Sipariş düzenleme penceresini açar"""
        try:
            # Seçili satır var mı kontrol et
            selection = self.order_tree.selection()
            if selection:
                # Sipariş düzenleme fonksiyonunu çağır
                self.edit_order()
        except Exception as e:
            # Hata olursa sessizce geç (kullanıcıyı rahatsız etmeyelim)
            pass
    
    def setup_stats_tab(self):
        """Gelir ve Danışanlar sekmesini hazırlar"""
        # Ana container
        main_container = ttk.Frame(self.stats_frame, style='Modern.TFrame')
        main_container.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        
        # Grid layout için - 2x2 layout
        main_container.grid_columnconfigure(0, weight=1)
        main_container.grid_columnconfigure(1, weight=1)
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_rowconfigure(1, weight=1)
        
        # Stats kartları - 2x2 grid layout
        # Kart 1 - Toplam Danışan
        card1 = tk.Frame(main_container, bg='#3498db', relief='flat', bd=0, width=250, height=180)
        card1.grid(row=0, column=0, padx=10, pady=10, sticky='nsew')
        card1.grid_propagate(False)
        
        tk.Label(card1, text="👥", font=('Segoe UI', 30), 
                bg='#3498db', fg='white').pack(pady=(15, 5))
        tk.Label(card1, text="TOPLAM DANIŞAN", font=('Segoe UI', 10, 'bold'), 
                bg='#3498db', fg='white').pack()
        self.total_customers_label = tk.Label(card1, text="0", font=('Segoe UI', 20, 'bold'), 
                                            bg='#3498db', fg='white')
        self.total_customers_label.pack(pady=(5, 15))
        
        # Kart 2 - Güncel Danışan
        card2 = tk.Frame(main_container, bg='#f39c12', relief='flat', bd=0, width=250, height=180)
        card2.grid(row=0, column=1, padx=10, pady=10, sticky='nsew')
        card2.grid_propagate(False)
        
        tk.Label(card2, text="📋", font=('Segoe UI', 30), 
                bg='#f39c12', fg='white').pack(pady=(15, 5))
        tk.Label(card2, text="GÜNCEL DANIŞAN", font=('Segoe UI', 10, 'bold'), 
                bg='#f39c12', fg='white').pack()
        self.total_orders_label = tk.Label(card2, text="0", font=('Segoe UI', 20, 'bold'), 
                                         bg='#f39c12', fg='white')
        self.total_orders_label.pack(pady=(5, 15))
        
        # Kart 3 - Toplam Gelir
        card3 = tk.Frame(main_container, bg='#27ae60', relief='flat', bd=0, width=250, height=180)
        card3.grid(row=1, column=0, padx=10, pady=10, sticky='nsew')
        card3.grid_propagate(False)
        
        tk.Label(card3, text="💰", font=('Segoe UI', 30), 
                bg='#27ae60', fg='white').pack(pady=(15, 5))
        tk.Label(card3, text="TOPLAM GELİR", font=('Segoe UI', 10, 'bold'), 
                bg='#27ae60', fg='white').pack()
        self.total_revenue_label = tk.Label(card3, text="0 TL", font=('Segoe UI', 20, 'bold'), 
                                          bg='#27ae60', fg='white')
        self.total_revenue_label.pack(pady=(5, 15))
        
        # Kart 4 - İptal Edilen Siparişler
        card4 = tk.Frame(main_container, bg='#e74c3c', relief='flat', bd=0, width=250, height=180)
        card4.grid(row=1, column=1, padx=10, pady=10, sticky='nsew')
        card4.grid_propagate(False)
        
        tk.Label(card4, text="❌", font=('Segoe UI', 30), 
                bg='#e74c3c', fg='white').pack(pady=(15, 5))
        tk.Label(card4, text="İPTAL EDİLEN TUTAR", font=('Segoe UI', 10, 'bold'), 
                bg='#e74c3c', fg='white').pack()
        self.cancelled_revenue_label = tk.Label(card4, text="0 TL", font=('Segoe UI', 20, 'bold'), 
                                              bg='#e74c3c', fg='white')
        self.cancelled_revenue_label.pack(pady=(5, 15))
        
        # Kartların yüksekliğini eşitle
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_rowconfigure(1, weight=1)
        
    
    def setup_notifications_tab(self):
        """Bildirimler sekmesini hazırlar"""
        # Üst frame - Butonlar (beyaz arka plan)
        top_frame = ttk.Frame(self.notifications_frame, style='White.TFrame')
        top_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Modern buton stili
        button_style = {
            'font': ('Segoe UI', 9, 'bold'),
            'relief': 'flat',
            'borderwidth': 0,
            'padx': 15,
            'pady': 8,
            'cursor': 'hand2',
            'activebackground': '#1a252f',
            'activeforeground': 'black'
        }
        
        check_payments_btn = tk.Button(top_frame, text="🔍 Ödemeleri Kontrol Et", 
                                      command=self.check_notifications,
                                      bg='#2c3e50', fg='white', **button_style)
        check_payments_btn.pack(side=tk.LEFT, padx=6)
        
        refresh_payments_btn = tk.Button(top_frame, text="🔄 Ödemeleri Yenile",
                                        command=self.load_notifications,
                                        bg='#3498db', fg='white', **button_style)
        refresh_payments_btn.pack(side=tk.LEFT, padx=6)
        
        complete_overdue_btn = tk.Button(top_frame, text="✅ Gecikmiş Siparişleri Tamamla", 
                                        command=self.complete_overdue_orders,
                                        bg='#e74c3c', fg='white', **button_style)
        complete_overdue_btn.pack(side=tk.LEFT, padx=6)
        
        edit_status_btn = tk.Button(top_frame, text="✏️ Durumu Düzenle", 
                                   command=self.edit_expiring_order_status,
                                   bg='#f39c12', fg='white', **button_style)
        edit_status_btn.pack(side=tk.LEFT, padx=6)
        
        # Hover efektleri
        def on_enter(e):
            if e.widget == check_payments_btn:
                e.widget['bg'] = '#1a252f'
            elif e.widget == refresh_payments_btn:
                e.widget['bg'] = '#2980b9'
            elif e.widget == complete_overdue_btn:
                e.widget['bg'] = '#c0392b'
            elif e.widget == edit_status_btn:
                e.widget['bg'] = '#e67e22'
        
        def on_leave(e):
            if e.widget == check_payments_btn:
                e.widget['bg'] = '#2c3e50'
            elif e.widget == refresh_payments_btn:
                e.widget['bg'] = '#3498db'
            elif e.widget == complete_overdue_btn:
                e.widget['bg'] = '#e74c3c'
            elif e.widget == edit_status_btn:
                e.widget['bg'] = '#f39c12'
        
        for btn in [check_payments_btn, refresh_payments_btn, complete_overdue_btn, edit_status_btn]:
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)

        
        # Notebook for notifications - Modern stil
        self.notifications_notebook = ttk.Notebook(self.notifications_frame, style='Modern.TNotebook')
        self.notifications_notebook.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Yaklaşan bitiş tarihleri sekmesi
        self.expiring_frame = ttk.Frame(self.notifications_notebook, style='Modern.TFrame')
        self.notifications_notebook.add(self.expiring_frame, text="Yaklaşan Ödemeler")
        
        # Gecikmiş siparişler sekmesi
        self.overdue_frame = ttk.Frame(self.notifications_notebook, style='Modern.TFrame')
        self.notifications_notebook.add(self.overdue_frame, text="Gecikmiş Ödemeler")
        
        # Yaklaşan bitiş tarihleri listesi
        expiring_list_frame = ttk.Frame(self.expiring_frame, style='Modern.TFrame')
        expiring_list_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        columns = ('ID', 'Müşteri', 'Ürün', 'Bitiş Tarihi', 'Durum', 'Bildirim')
        self.expiring_tree = ttk.Treeview(expiring_list_frame, columns=columns, show='headings', height=10, style='Modern.Treeview')
        
        # Treeview satır stilleri
        self.expiring_tree.tag_configure('oddrow', background='#f8f9fa')
        self.expiring_tree.tag_configure('evenrow', background='#ffffff')
        self.expiring_tree.tag_configure('selected', background='#e3f2fd')
        
        self.expiring_tree.heading('ID', text='ID', anchor='w')
        self.expiring_tree.heading('Müşteri', text='Müşteri', anchor='w')
        self.expiring_tree.heading('Ürün', text='Ürün', anchor='w')
        self.expiring_tree.heading('Bitiş Tarihi', text='Bitiş Tarihi', anchor='w')
        self.expiring_tree.heading('Durum', text='Durum', anchor='w')
        self.expiring_tree.heading('Bildirim', text='Bildirim', anchor='w')
        
        self.expiring_tree.column('ID', width=0, stretch=False, anchor='w')  # Gizli sütun
        self.expiring_tree.column('Müşteri', width=150, anchor='w')
        self.expiring_tree.column('Ürün', width=200, anchor='w')
        self.expiring_tree.column('Bitiş Tarihi', width=120, anchor='w')
        self.expiring_tree.column('Durum', width=100, anchor='w')
        self.expiring_tree.column('Bildirim', width=100, anchor='w')
        
        # Modern custom scrollbar for expiring table
        expiring_scrollbar_frame = tk.Frame(expiring_list_frame, bg='white', width=8)
        expiring_scrollbar_bg = tk.Canvas(expiring_scrollbar_frame, bg='white', width=8, highlightthickness=0, bd=0)
        
        # Modern scrollbar functions for expiring
        def update_expiring_scrollbar():
            expiring_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.expiring_tree.yview()
                if bottom - top >= 1.0:
                    return
                    
                scrollbar_height = expiring_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                expiring_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        # Scrollbar drag variables for expiring
        self.expiring_scrollbar_dragging = False
        
        def expiring_smooth_scroll(event):
            try:
                if self.expiring_tree.winfo_exists():
                    self.expiring_tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    update_expiring_scrollbar()
            except tk.TclError:
                pass
        
        def on_expiring_scrollbar_click(event):
            try:
                top, bottom = self.expiring_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = expiring_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.expiring_tree.yview_moveto(click_position)
                update_expiring_scrollbar()
                self.expiring_scrollbar_dragging = True
            except:
                pass
        
        def on_expiring_scrollbar_drag(event):
            if not self.expiring_scrollbar_dragging:
                return
            try:
                top, bottom = self.expiring_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = expiring_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.expiring_tree.yview_moveto(click_position)
                update_expiring_scrollbar()
            except:
                pass
        
        def on_expiring_scrollbar_release(event):
            self.expiring_scrollbar_dragging = False
        
        def on_expiring_scrollbar_enter(event):
            expiring_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.expiring_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = expiring_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                expiring_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_expiring_scrollbar_leave(event):
            update_expiring_scrollbar()
        
        def on_expiring_tree_scroll(*args):
            update_expiring_scrollbar()
        
        self.expiring_tree.configure(yscrollcommand=on_expiring_tree_scroll)
        
        # Bind events for expiring
        self.expiring_tree.bind("<MouseWheel>", expiring_smooth_scroll)
        expiring_scrollbar_bg.bind("<Button-1>", on_expiring_scrollbar_click)
        expiring_scrollbar_bg.bind("<B1-Motion>", on_expiring_scrollbar_drag)
        expiring_scrollbar_bg.bind("<ButtonRelease-1>", on_expiring_scrollbar_release)
        expiring_scrollbar_bg.bind("<Enter>", on_expiring_scrollbar_enter)
        expiring_scrollbar_bg.bind("<Leave>", on_expiring_scrollbar_leave)
        expiring_scrollbar_bg.configure(takefocus=True)
        
        # Pack widgets for expiring
        self.expiring_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        expiring_scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=0)
        expiring_scrollbar_frame.pack_propagate(False)
        expiring_scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize expiring scrollbar
        self.expiring_tree.after(200, update_expiring_scrollbar)
        

        
        # Gecikmiş siparişler listesi
        overdue_list_frame = ttk.Frame(self.overdue_frame, style='Modern.TFrame')
        overdue_list_frame.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        columns = ('ID', 'Müşteri', 'Ürün', 'Bitiş Tarihi', 'Durum')
        self.overdue_tree = ttk.Treeview(overdue_list_frame, columns=columns, show='headings', height=10, style='Modern.Treeview')
        
        # Treeview satır stilleri
        self.overdue_tree.tag_configure('oddrow', background='#f8f9fa')
        self.overdue_tree.tag_configure('evenrow', background='#ffffff')
        self.overdue_tree.tag_configure('selected', background='#e3f2fd')
        
        self.overdue_tree.heading('ID', text='ID', anchor='w')
        self.overdue_tree.heading('Müşteri', text='Müşteri', anchor='w')
        self.overdue_tree.heading('Ürün', text='Ürün', anchor='w')
        self.overdue_tree.heading('Bitiş Tarihi', text='Bitiş Tarihi', anchor='w')
        self.overdue_tree.heading('Durum', text='Durum', anchor='w')
        
        self.overdue_tree.column('ID', width=0, stretch=False, anchor='w')  # Gizli sütun
        self.overdue_tree.column('Müşteri', width=150, anchor='w')
        self.overdue_tree.column('Ürün', width=200, anchor='w')
        self.overdue_tree.column('Bitiş Tarihi', width=120, anchor='w')
        self.overdue_tree.column('Durum', width=100, anchor='w')
        
        # Modern custom scrollbar for overdue table
        overdue_scrollbar_frame = tk.Frame(overdue_list_frame, bg='white', width=8)
        overdue_scrollbar_bg = tk.Canvas(overdue_scrollbar_frame, bg='white', width=8, highlightthickness=0, bd=0)
        
        # Modern scrollbar functions for overdue
        def update_overdue_scrollbar():
            overdue_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.overdue_tree.yview()
                if bottom - top >= 1.0:
                    return
                    
                scrollbar_height = overdue_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                overdue_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        # Scrollbar drag variables for overdue
        self.overdue_scrollbar_dragging = False
        
        def overdue_smooth_scroll(event):
            try:
                if self.overdue_tree.winfo_exists():
                    self.overdue_tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    update_overdue_scrollbar()
            except tk.TclError:
                pass
        
        def on_overdue_scrollbar_click(event):
            try:
                top, bottom = self.overdue_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = overdue_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.overdue_tree.yview_moveto(click_position)
                update_overdue_scrollbar()
                self.overdue_scrollbar_dragging = True
            except:
                pass
        
        def on_overdue_scrollbar_drag(event):
            if not self.overdue_scrollbar_dragging:
                return
            try:
                top, bottom = self.overdue_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = overdue_scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                self.overdue_tree.yview_moveto(click_position)
                update_overdue_scrollbar()
            except:
                pass
        
        def on_overdue_scrollbar_release(event):
            self.overdue_scrollbar_dragging = False
        
        def on_overdue_scrollbar_enter(event):
            overdue_scrollbar_bg.delete("thumb")
            try:
                top, bottom = self.overdue_tree.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = overdue_scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                overdue_scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_overdue_scrollbar_leave(event):
            update_overdue_scrollbar()
        
        def on_overdue_tree_scroll(*args):
            update_overdue_scrollbar()
        
        self.overdue_tree.configure(yscrollcommand=on_overdue_tree_scroll)
        
        # Bind events for overdue
        self.overdue_tree.bind("<MouseWheel>", overdue_smooth_scroll)
        overdue_scrollbar_bg.bind("<Button-1>", on_overdue_scrollbar_click)
        overdue_scrollbar_bg.bind("<B1-Motion>", on_overdue_scrollbar_drag)
        overdue_scrollbar_bg.bind("<ButtonRelease-1>", on_overdue_scrollbar_release)
        overdue_scrollbar_bg.bind("<Enter>", on_overdue_scrollbar_enter)
        overdue_scrollbar_bg.bind("<Leave>", on_overdue_scrollbar_leave)
        overdue_scrollbar_bg.configure(takefocus=True)
        
        # Pack widgets for overdue
        self.overdue_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        overdue_scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=0)
        overdue_scrollbar_frame.pack_propagate(False)
        overdue_scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize overdue scrollbar
        self.overdue_tree.after(200, update_overdue_scrollbar)
        

        
        # İlk yükleme
        self.load_notifications()
        
        # Uygulama kapatılırken otomatik kontrolü durdur
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def edit_expiring_order_status(self):
        """Yaklaşan ödemeler listesinden seçili siparişin durumunu düzenler"""
        # Hangi sekmenin aktif olduğunu kontrol et
        current_tab = self.notifications_notebook.select()
        tab_id = self.notifications_notebook.index(current_tab)
        
        if tab_id == 0:  # Yaklaşan ödemeler sekmesi
            selection = self.expiring_tree.selection()
            tree = self.expiring_tree
        else:  # Gecikmiş ödemeler sekmesi
            selection = self.overdue_tree.selection()
            tree = self.overdue_tree
            
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen düzenlenecek siparişi seçin!")
            return
        
        # Seçili siparişin bilgilerini al
        item = tree.item(selection[0])
        values = item['values']
        
        if not values or len(values) < 5:
            messagebox.showerror("Hata", "Sipariş bilgileri alınamadı!")
            return
        
        order_id = values[0]  # Sipariş ID (gizli sütun)
        customer_name = values[1]
        product_name = values[2]
        end_date = values[3]
        current_status = values[4]
        
        # Sipariş ID'sini kontrol et
        if not order_id:
            messagebox.showerror("Hata", "Sipariş ID bulunamadı!")
            return
        
        try:
            # Durum düzenleme dialog'u - Modern tasarım
            dialog = tk.Toplevel(self.root)
            dialog.title(f"✏️ Sipariş Durumu Düzenle - {customer_name}")
            dialog.configure(bg='#f8f9fa')
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.resizable(True, True)
            
            # Optimal boyut ayarla - Modern butonlar için büyütüldü
            dialog.geometry("1200x1000+100+30")
            dialog.minsize(1000, 800)
            
            # Ana container
            main_container = tk.Frame(dialog, bg='#f8f9fa')
            main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Header card
            header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            header_card.pack(fill=tk.X, pady=(0, 25))
            
            header_inner = tk.Frame(header_card, bg='white')
            header_inner.pack(fill=tk.X, padx=40, pady=30)
            
            tk.Label(header_inner,
                    text="✏️ Sipariş Durumu Düzenle",
                    font=('Segoe UI', 20, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack()
            
            tk.Label(header_inner,
                    text=f"{customer_name} - Sipariş durumunu güncelleyin",
                    font=('Segoe UI', 11),
                    bg='white',
                    fg='#7f8c8d').pack(pady=(8, 0))
            
            # Info card - Sipariş bilgileri
            info_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            info_card.pack(fill=tk.X, pady=(0, 25))
            
            info_inner = tk.Frame(info_card, bg='white')
            info_inner.pack(fill=tk.X, padx=40, pady=30)
            
            # Sipariş bilgileri başlığı
            tk.Label(info_inner,
                    text="Sipariş Bilgileri",
                    font=('Segoe UI', 14, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(0, 20))
            
            # Bilgi grid'i
            info_grid = tk.Frame(info_inner, bg='white')
            info_grid.pack(fill=tk.X)
            
            # İlk satır - Müşteri ve Ürün
            row1 = tk.Frame(info_grid, bg='white')
            row1.pack(fill=tk.X, pady=(0, 15))
            
            # Müşteri (Sol)
            customer_col = tk.Frame(row1, bg='white')
            customer_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 20))
            
            tk.Label(customer_col, text="👤 Müşteri", 
                    font=('Segoe UI', 10, 'bold'), 
                    bg='white', fg='#7f8c8d').pack(anchor='w')
            tk.Label(customer_col, text=customer_name, 
                    font=('Segoe UI', 12, 'bold'), 
                    bg='white', fg='#2c3e50').pack(anchor='w', pady=(5, 0))
            
            # Ürün (Sağ)
            product_col = tk.Frame(row1, bg='white')
            product_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(20, 0))
            
            tk.Label(product_col, text="📦 Ürün/Hizmet", 
                    font=('Segoe UI', 10, 'bold'), 
                    bg='white', fg='#7f8c8d').pack(anchor='w')
            tk.Label(product_col, text=product_name, 
                    font=('Segoe UI', 12, 'bold'), 
                    bg='white', fg='#2c3e50').pack(anchor='w', pady=(5, 0))
            
            # İkinci satır - Bitiş Tarihi ve Mevcut Durum
            row2 = tk.Frame(info_grid, bg='white')
            row2.pack(fill=tk.X, pady=(0, 15))
            
            # Bitiş Tarihi (Sol)
            date_col = tk.Frame(row2, bg='white')
            date_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 20))
            
            tk.Label(date_col, text="📅 Bitiş Tarihi", 
                    font=('Segoe UI', 10, 'bold'), 
                    bg='white', fg='#7f8c8d').pack(anchor='w')
            tk.Label(date_col, text=end_date, 
                    font=('Segoe UI', 12, 'bold'), 
                    bg='white', fg='#e74c3c').pack(anchor='w', pady=(5, 0))
            
            # Mevcut Durum (Sağ)
            status_col = tk.Frame(row2, bg='white')
            status_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(20, 0))
            
            tk.Label(status_col, text="⚡ Mevcut Durum", 
                    font=('Segoe UI', 10, 'bold'), 
                    bg='white', fg='#7f8c8d').pack(anchor='w')
            tk.Label(status_col, text=current_status, 
                    font=('Segoe UI', 12, 'bold'), 
                    bg='white', fg='#f39c12').pack(anchor='w', pady=(5, 0))
            
            # Status card - Durum seçimi
            status_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            status_card.pack(fill=tk.X, pady=(0, 25))
            
            status_inner = tk.Frame(status_card, bg='white')
            status_inner.pack(fill=tk.X, padx=40, pady=30)
            
            # Durum seçimi başlığı
            tk.Label(status_inner,
                    text="Yeni Durum Seçin",
                    font=('Segoe UI', 14, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(0, 20))
            
            # Durum seçimi container
            status_select = tk.Frame(status_inner, bg='white')
            status_select.pack(fill=tk.X)
            
            tk.Label(status_select, text="🔄 Sipariş Durumu", 
                    font=('Segoe UI', 11, 'bold'), 
                    bg='white', fg='#7f8c8d').pack(anchor='w', pady=(0, 15))
            
            status_var = tk.StringVar(value=current_status)
            
            # Modern durum butonları
            status_options = [
                ("⏳ Beklemede", "Beklemede", "#f39c12"),
                ("🔧 Devam Ediyor", "Devam Ediyor", "#3498db"),
                ("✅ Tamamlandı", "Tamamlandı", "#27ae60"),
                ("❌ İptal", "İptal", "#e74c3c")
            ]
            
            status_buttons_frame = tk.Frame(status_select, bg='white')
            status_buttons_frame.pack(fill=tk.X, pady=(5, 0))
            
            status_buttons = []
            
            def select_status(status_value):
                status_var.set(status_value)
                # Buton renklerini güncelle
                for btn, (_, btn_status, btn_color) in zip(status_buttons, status_options):
                    if btn_status == status_value:
                        btn.configure(bg=btn_color, fg='white', relief='solid', bd=2)
                    else:
                        btn.configure(bg='#ecf0f1', fg='#34495e', relief='flat', bd=0)
            
            # Durum butonlarını oluştur
            for i, (display_text, status_value, color) in enumerate(status_options):
                btn = tk.Button(status_buttons_frame,
                              text=display_text,
                              command=lambda s=status_value: select_status(s),
                              font=('Segoe UI', 12, 'bold'),
                              bg='#ecf0f1' if status_value != current_status else color,
                              fg='#34495e' if status_value != current_status else 'white',
                              relief='flat' if status_value != current_status else 'solid',
                              bd=0 if status_value != current_status else 2,
                              padx=20,
                              pady=12,
                              cursor='hand2',
                              borderwidth=0)
                
                # 2x2 grid düzeni
                row = i // 2
                col = i % 2
                btn.grid(row=row, column=col, padx=8, pady=8, sticky='ew')
                status_buttons.append(btn)
                
                # Hover efektleri
                def on_enter(e, original_color=color):
                    if e.widget['bg'] != original_color:  # Seçili değilse
                        e.widget.configure(bg=original_color, fg='white')
                
                def on_leave(e):
                    current_sel_status = status_var.get()
                    btn_index = status_buttons.index(e.widget)
                    btn_status = status_options[btn_index][1]
                    if btn_status != current_sel_status:
                        e.widget.configure(bg='#ecf0f1', fg='#34495e')
                    else:
                        # Seçili buton kendi renginde kalır
                        btn_color = status_options[btn_index][2]
                        e.widget.configure(bg=btn_color, fg='white')
                
                btn.bind('<Enter>', on_enter)
                btn.bind('<Leave>', on_leave)
            
            # Grid ağırlıklarını ayarla
            status_buttons_frame.columnconfigure(0, weight=1)
            status_buttons_frame.columnconfigure(1, weight=1)
            
            # Durum açıklamaları
            status_info = tk.Label(status_select, 
                                  text="💡 İpucu: Durumu değiştirmek için yukarıdaki butonlara tıklayın", 
                                  font=('Segoe UI', 9, 'italic'), 
                                  bg='white', fg='#95a5a6', justify=tk.LEFT)
            status_info.pack(anchor='w', pady=(15, 0))
            
            def save_status():
                new_status = status_var.get()
                if new_status == current_status:
                    messagebox.showinfo("Bilgi", "Durum değişmedi!")
                    dialog.destroy()
                    return
                
                try:
                    # Durumu güncelle
                    success = self.db.update_order_status(order_id, new_status)
                    if success:
                        messagebox.showinfo("Başarılı", "Sipariş durumu güncellendi!")
                        # Otomatik kayıt
                        self.auto_save_data("Sipariş Durumu Güncellendi")
                        # Cache'i temizle ve listeleri yenile
                        self.clear_cache()
                        self.load_notifications()
                        self.load_orders()  # Sipariş sekmesini de yenile
                        dialog.destroy()
                    else:
                        messagebox.showerror("Hata", "Durum güncellenirken hata oluştu!")
                except Exception as e:
                    messagebox.showerror("Hata", f"Durum güncellenirken hata oluştu: {str(e)}")
            
            # Button card - Alt kısım
            button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            button_card.pack(fill=tk.X, pady=(0, 0))
            
            button_inner = tk.Frame(button_card, bg='white')
            button_inner.pack(fill=tk.X, padx=40, pady=(20, 30))
            
            # Button container - Sağa hizalı
            btn_container = tk.Frame(button_inner, bg='white')
            btn_container.pack(anchor='e')
            
            # İptal butonu - Kırmızı
            cancel_btn = tk.Button(btn_container, 
                                 text="✖ İptal",
                                 command=dialog.destroy,
                                 font=('Segoe UI', 11, 'bold'),
                                 bg='#e74c3c', 
                                 fg='white',
                                 relief='flat',
                                 borderwidth=0,
                                 padx=30,
                                 pady=12,
                                 cursor='hand2')
            cancel_btn.pack(side=tk.LEFT, padx=(0, 15))
            
            # Kaydet butonu - Yeşil
            save_btn = tk.Button(btn_container, 
                               text="✓ Durumu Güncelle",
                               command=save_status,
                               font=('Segoe UI', 11, 'bold'),
                               bg='#27ae60', 
                               fg='white',
                               relief='flat',
                               borderwidth=0,
                               padx=30,
                               pady=12,
                               cursor='hand2')
            save_btn.pack(side=tk.LEFT)
            
            # Hover efektleri
            def on_save_hover(e):
                save_btn.config(bg='#219a52')
                
            def on_save_leave(e):
                save_btn.config(bg='#27ae60')
                
            def on_cancel_hover(e):
                cancel_btn.config(bg='#c0392b')
                
            def on_cancel_leave(e):
                cancel_btn.config(bg='#e74c3c')
                
            save_btn.bind('<Enter>', on_save_hover)
            save_btn.bind('<Leave>', on_save_leave)
            cancel_btn.bind('<Enter>', on_cancel_hover)
            cancel_btn.bind('<Leave>', on_cancel_leave)
            
        except Exception as e:
            messagebox.showerror("Hata", f"Sipariş düzenleme penceresi açılırken hata oluştu: {str(e)}")
    
    def edit_expiring_order(self):
        """Yaklaşan ödemeler listesinden seçili siparişi düzenler"""
        # Hangi sekmenin aktif olduğunu kontrol et
        current_tab = self.notifications_notebook.select()
        tab_id = self.notifications_notebook.index(current_tab)
        
        if tab_id == 0:  # Yaklaşan ödemeler sekmesi
            selection = self.expiring_tree.selection()
            tree = self.expiring_tree
        else:  # Gecikmiş ödemeler sekmesi
            selection = self.overdue_tree.selection()
            tree = self.overdue_tree
            
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen düzenlenecek siparişi seçin!")
            return
        
        # Seçili siparişin bilgilerini al
        item = tree.item(selection[0])
        values = item['values']
        
        if not values or len(values) < 5:
            messagebox.showerror("Hata", "Sipariş bilgileri alınamadı!")
            return
        
        order_id = values[0]  # Sipariş ID (gizli sütun)
        
        # Sipariş ID'sini kontrol et
        if not order_id:
            messagebox.showerror("Hata", "Sipariş ID bulunamadı!")
            return
        
        # Mevcut edit_order fonksiyonunu çağır
        self.edit_order_by_id(order_id)
    
    def edit_order_by_id(self, order_id):
        """Sipariş ID'sine göre düzenleme penceresini açar"""
        try:
            # Sipariş bilgilerini getir
            order = self.db.get_order_by_id(order_id)
            
            if not order:
                messagebox.showerror("Hata", "Sipariş bulunamadı!")
                return
            
            # Mevcut edit_order fonksiyonunu çağır
            self.edit_order()
            
        except Exception as e:
            messagebox.showerror("Hata", f"Sipariş düzenleme penceresi açılırken hata oluştu: {str(e)}")
    
    def format_date_for_display(self, date_str):
        """Veritabanı tarihini görüntüleme formatına çevirir (DD.MM.YYYY)"""
        try:
            if not date_str:
                return ""
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            return date_obj.strftime("%d.%m.%Y")
        except:
            return date_str
    
    def format_date_for_database(self, date_str):
        """Görüntüleme tarihini veritabanı formatına çevirir (YYYY-MM-DD)"""
        try:
            if not date_str:
                return ""
            # DD.MM.YYYY formatını parse et
            date_obj = datetime.strptime(date_str, "%d.%m.%Y")
            return date_obj.strftime("%Y-%m-%d")
        except:
            # Eğer YYYY-MM-DD formatındaysa olduğu gibi bırak
            try:
                datetime.strptime(date_str, "%Y-%m-%d")
                return date_str
            except:
                return ""
    
    def validate_phone_number(self, phone):
        """Telefon numarasının sadece rakam içerip içermediğini kontrol eder"""
        # Sadece rakamları al (boşluk, tire, parantez gibi karakterleri kaldır)
        digits_only = ''.join(filter(str.isdigit, phone))
        return digits_only == phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
    
    def clean_phone_number(self, phone):
        """Telefon numarasından sadece rakamları alır"""
        return ''.join(filter(str.isdigit, phone))
    
    def validate_phone_input(self, P):
        """Telefon alanına sadece rakam girilmesini sağlar"""
        # Boş string'e izin ver (silme işlemi için)
        if P == "":
            return True
        # Sadece rakam kontrolü
        return P.isdigit()
    
    def auto_save_data(self, operation_type=""):
        """Verileri otomatik olarak kaydeder"""
        try:
            # Veritabanı bağlantısını yenile (commit işlemi için)
            import sqlite3
            conn = sqlite3.connect(self.db.db_name)
            conn.commit()
            conn.close()
            
            # Kayıt bilgisi
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_message = f"✅ Otomatik kayıt alındı: {timestamp}"
            if operation_type:
                save_message += f" ({operation_type})"
            
            # Konsola kayıt bilgisi yazdır
            print(save_message)
            
            # Ses sistemi kaldırıldı
            
        except Exception as e:
            print(f"❌ Otomatik kayıt hatası: {e}")
    
    def show_auto_save_notification(self):
        """Otomatik kayıt bildirimi gösterir"""
        try:
            # Küçük bir bildirim penceresi
            notification = tk.Toplevel(self.root)
            notification.title("Otomatik Kayıt")
            notification.geometry("300x100")
            notification.transient(self.root)
            notification.attributes('-topmost', True)
            
            # Pencereyi merkeze yerleştir
            notification.update_idletasks()
            x = (notification.winfo_screenwidth() // 2) - (300 // 2)
            y = (notification.winfo_screenheight() // 2) - (100 // 2)
            notification.geometry(f"300x100+{x}+{y}")
            
            # Mesaj
            timestamp = datetime.now().strftime("%H:%M:%S")
            ttk.Label(notification, text=f"💾 Otomatik Kayıt Alındı", 
                     font=('Arial', 12, 'bold')).pack(pady=10)
            ttk.Label(notification, text=f"Saat: {timestamp}", 
                     font=('Arial', 10)).pack(pady=5)
            
            # 2 saniye sonra otomatik kapat
            notification.after(2000, notification.destroy)
            
        except Exception as e:
            print(f"Bildirim hatası: {e}")
    
    # Excele Aktar Fonksiyonları
    def show_excel_export_dialog(self):
        """Modern Excel export dialog'unu gösterir"""
        dialog = tk.Toplevel(self.root)
        dialog.title("📊 Excele Aktar")
        dialog.geometry("800x600")
        dialog.configure(bg='#f8f9fa')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Pencereyi ortalama
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (600 // 2)
        dialog.geometry(f"800x600+{x}+{y}")
        
        # Ana container - Modern card design
        main_container = tk.Frame(dialog, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # Modern card frame
        card_frame = tk.Frame(main_container, bg='#ffffff', relief='flat', bd=1)
        card_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header - Modern başlık
        header_frame = tk.Frame(card_frame, bg='#3498db', height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Başlık metni - ortalanmış
        header_label = tk.Label(header_frame, text="📊 EXCELE AKTAR", 
                               font=('Segoe UI', 20, 'bold'), 
                               bg='#3498db', fg='white')
        header_label.pack(expand=True)
        
        # Content area
        content_frame = tk.Frame(card_frame, bg='#ffffff')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        # Export Türü Seçimi Card
        export_type_card = tk.Frame(content_frame, bg='#f8f9fa', relief='flat', bd=1, pady=20)
        export_type_card.pack(fill=tk.X, pady=(0, 25))
        
        # Export türü başlığı
        export_type_header = tk.Label(export_type_card, text="📋 Ne Export Etmek İstiyorsunuz?", 
                                     font=('Segoe UI', 16, 'bold'), 
                                     bg='#f8f9fa', fg='#2c3e50')
        export_type_header.pack(pady=(15, 20))
        
        export_type = tk.StringVar(value="customers")
        
        # Export türü butonları - Modern button style
        export_buttons_frame = tk.Frame(export_type_card, bg='#f8f9fa')
        export_buttons_frame.pack(pady=(0, 20))
        
        # Export türü buton seçim fonksiyonu
        def select_export_type(type_value, selected_btn, other_btn, selected_color, other_color):
            export_type.set(type_value)
            selected_btn.config(bg=selected_color)  # Seçili orijinal renk
            other_btn.config(bg='#95a5a6')          # Seçili olmayan gri renk
        
        # Danışan Listesi Butonu
        customers_btn = tk.Button(export_buttons_frame, text="👥 Danışan Listesi", 
                                 bg='#3498db', fg='white',  # Mavi renk
                                 font=('Segoe UI', 14, 'bold'),
                                 relief='flat', borderwidth=0,
                                 padx=40, pady=15,
                                 cursor='hand2')
        customers_btn.pack(side=tk.LEFT, padx=15)
        
        # Sipariş Listesi Butonu
        orders_btn = tk.Button(export_buttons_frame, text="📦 Sipariş Listesi", 
                              bg='#95a5a6', fg='white',  # Varsayılan seçili değil
                              font=('Segoe UI', 14, 'bold'),
                              relief='flat', borderwidth=0,
                              padx=40, pady=15,
                              cursor='hand2')
        orders_btn.pack(side=tk.LEFT, padx=15)
        
        # Hover efektleri - Export türü butonları
        def on_enter_customers(e):
            if export_type.get() == "customers":
                customers_btn.config(bg='#2980b9')  # Koyu mavi hover
            else:
                customers_btn.config(bg='#7f8c8d')  # Gri hover
        def on_leave_customers(e):
            if export_type.get() == "customers":
                customers_btn.config(bg='#3498db')  # Orijinal mavi
            else:
                customers_btn.config(bg='#95a5a6')  # Orijinal gri
                
        def on_enter_orders(e):
            if export_type.get() == "orders":
                orders_btn.config(bg='#229954')     # Koyu yeşil hover
            else:
                orders_btn.config(bg='#7f8c8d')     # Gri hover
        def on_leave_orders(e):
            if export_type.get() == "orders":
                orders_btn.config(bg='#27ae60')     # Orijinal yeşil
            else:
                orders_btn.config(bg='#95a5a6')     # Orijinal gri
        
        customers_btn.bind('<Enter>', on_enter_customers)
        customers_btn.bind('<Leave>', on_leave_customers)
        orders_btn.bind('<Enter>', on_enter_orders)
        orders_btn.bind('<Leave>', on_leave_orders)
        
        # Buton komutlarını ayarla
        customers_btn.config(command=lambda: select_export_type("customers", customers_btn, orders_btn, '#3498db', '#27ae60'))
        orders_btn.config(command=lambda: select_export_type("orders", orders_btn, customers_btn, '#27ae60', '#3498db'))
        
        
        
        def start_export():
            try:
                export_type_val = export_type.get()
                
                if export_type_val == "customers":
                    self.export_customers_to_excel()
                else:  # orders
                    self.export_orders_to_excel()
                
                messagebox.showinfo("Başarılı", "Export işlemi tamamlandı!")
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("Hata", f"Export sırasında hata oluştu:\n{str(e)}")
        
        # İşlemler Card - Modern button area
        actions_card = tk.Frame(card_frame, bg='#f8f9fa', height=150)
        actions_card.pack(fill=tk.X, side=tk.BOTTOM, pady=30)
        actions_card.pack_propagate(False)
        
        # Buton container - ortalanmış
        button_container = tk.Frame(actions_card, bg='#f8f9fa')
        button_container.pack(expand=True, pady=20)
        
        # Excele Aktar Butonu - Büyük ve modern
        export_btn = tk.Button(button_container, text="🚀 Excele Aktar", 
                              command=start_export,
                              bg='#27ae60', fg='white',
                              font=('Segoe UI', 16, 'bold'),
                              relief='flat', borderwidth=0,
                              padx=45, pady=16,
                              cursor='hand2')
        export_btn.pack(side=tk.LEFT, padx=20)
        
        # Hover effects
        def on_enter_export(e):
            export_btn.config(bg='#229954')
        def on_leave_export(e):
            export_btn.config(bg='#27ae60')
            
        export_btn.bind('<Enter>', on_enter_export)
        export_btn.bind('<Leave>', on_leave_export)
        
        # İptal Butonu - Büyük ve kırmızı
        cancel_btn = tk.Button(button_container, text="❌ İptal", 
                              command=dialog.destroy,
                              bg='#e74c3c', fg='white',
                              font=('Segoe UI', 16, 'bold'),
                              relief='flat', borderwidth=0,
                              padx=45, pady=16,
                              cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=20)
        
        # Hover effects
        def on_enter_cancel(e):
            cancel_btn.config(bg='#c0392b')
        def on_leave_cancel(e):
            cancel_btn.config(bg='#e74c3c')
            
        cancel_btn.bind('<Enter>', on_enter_cancel)
        cancel_btn.bind('<Leave>', on_leave_cancel)
    
    def export_customers_to_excel(self):
        """Tüm müşterileri Excel'e export eder"""
        try:
            # Dosya adı seç
            filename = filedialog.asksaveasfilename(
                title="Müşteri Listesini Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel dosyaları", "*.xlsx"), ("Tüm dosyalar", "*.*")],
                initialfile=f"musteri_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Müşteri verilerini al
            customers = self.db.get_all_customers()
            
            if not customers:
                messagebox.showwarning("Uyarı", "Export edilecek müşteri bulunamadı!")
                return
            
            # DataFrame oluştur
            data = []
            for customer in customers:
                data.append({
                    'ID': customer[0],
                    'Ad Soyad': customer[1],
                    'E-posta': customer[2] or '',
                    'Telefon': customer[3],
                    'Adres': customer[4] or '',
                    'Şirket': customer[5] or '',
                    'Kayıt Tarihi': self.format_date_for_display(customer[6]),
                    'Notlar': customer[7] or ''
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Müşteriler', index=False)
                
                # Stil ayarları
                workbook = writer.book
                worksheet = writer.sheets['Müşteriler']
                
                # Başlık stili
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Sütun genişliklerini ayarla
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Alternatif satır renkleri
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            messagebox.showinfo("Başarılı", f"Müşteri listesi başarıyla export edildi!\n\nDosya: {filename}\n\nToplam: {len(customers)} müşteri")
            
        except Exception as e:
            messagebox.showerror("Export Hatası", f"Export sırasında hata oluştu:\n{str(e)}")
    
    def export_filtered_customers_to_excel(self):
        """Filtrelenmiş müşterileri Excel'e export eder"""
        try:
            # Arama terimini al
            search_term = self.customer_search_var.get().strip()
            if not search_term:
                messagebox.showwarning("Uyarı", "Lütfen önce arama yapın!")
                return
            
            # Dosya adı seç
            filename = filedialog.asksaveasfilename(
                title="Filtrelenmiş Müşteri Listesini Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel dosyaları", "*.xlsx"), ("Tüm dosyalar", "*.*")],
                initialfile=f"filtrelenmis_musteriler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Filtrelenmiş müşteri verilerini al
            customers = self.db.search_customers(search_term)
            
            if not customers:
                messagebox.showwarning("Uyarı", "Arama sonucu bulunamadı!")
                return
            
            # DataFrame oluştur
            data = []
            for customer in customers:
                data.append({
                    'ID': customer[0],
                    'Ad Soyad': customer[1],
                    'E-posta': customer[2] or '',
                    'Telefon': customer[3],
                    'Adres': customer[4] or '',
                    'Şirket': customer[5] or '',
                    'Kayıt Tarihi': self.format_date_for_display(customer[6]),
                    'Notlar': customer[7] or ''
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz (aynı stil ayarları)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Filtrelenmiş Müşteriler', index=False)
                
                # Stil ayarları
                workbook = writer.book
                worksheet = writer.sheets['Filtrelenmiş Müşteriler']
                
                # Başlık stili
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Sütun genişliklerini ayarla
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Alternatif satır renkleri
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            messagebox.showinfo("Başarılı", f"Filtrelenmiş müşteri listesi export edildi!\n\nDosya: {filename}\n\nArama: '{search_term}'\nToplam: {len(customers)} müşteri")
            
        except Exception as e:
            messagebox.showerror("Export Hatası", f"Export sırasında hata oluştu:\n{str(e)}")
    
    def export_orders_to_excel(self):
        """Tüm siparişleri Excel'e export eder"""
        try:
            # Dosya adı seç
            filename = filedialog.asksaveasfilename(
                title="Sipariş Listesini Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel dosyaları", "*.xlsx"), ("Tüm dosyalar", "*.*")],
                initialfile=f"siparis_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Sipariş verilerini al
            orders = self.db.get_all_orders()
            
            if not orders:
                messagebox.showwarning("Uyarı", "Export edilecek sipariş bulunamadı!")
                return
            
            # DataFrame oluştur
            data = []
            for order in orders:
                data.append({
                    'ID': order[0],
                    'Müşteri ID': order[1],
                    'Sipariş Tarihi': self.format_date_for_display(order[2]),
                    'Başlangıç Tarihi': self.format_date_for_display(order[3]),
                    'Bitiş Tarihi': self.format_date_for_display(order[4]),
                    'Ürün Adı': order[5],
                    'Miktar': order[6],
                    'Birim Fiyat': f"{order[7]:.2f} TL" if order[7] is not None else "0.00 TL",
                    'Toplam Fiyat': f"{order[8]:.2f} TL" if order[8] is not None else "0.00 TL",
                    'Durum': order[9],
                    'Bildirim Gönderildi': 'Evet' if order[10] == 1 else 'Hayır',
                    'Müşteri Adı': order[11] if len(order) > 11 else 'Bilinmiyor'
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Siparişler', index=False)
                
                # Stil ayarları
                workbook = writer.book
                worksheet = writer.sheets['Siparişler']
                
                # Başlık stili
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Sütun genişliklerini ayarla
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Alternatif satır renkleri
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            messagebox.showinfo("Başarılı", f"Sipariş listesi başarıyla export edildi!\n\nDosya: {filename}\n\nToplam: {len(orders)} sipariş")
            
        except Exception as e:
            messagebox.showerror("Export Hatası", f"Export sırasında hata oluştu:\n{str(e)}")
    
    def export_filtered_orders_to_excel(self):
        """Filtrelenmiş siparişleri Excel'e export eder"""
        try:
            # Arama terimini al
            search_term = self.order_search_var.get().strip()
            if not search_term:
                messagebox.showwarning("Uyarı", "Lütfen önce arama yapın!")
                return
            
            # Dosya adı seç
            filename = filedialog.asksaveasfilename(
                title="Filtrelenmiş Sipariş Listesini Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel dosyaları", "*.xlsx"), ("Tüm dosyalar", "*.*")],
                initialfile=f"filtrelenmis_siparisler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Filtrelenmiş sipariş verilerini al (basit arama)
            orders = self.db.get_all_orders()  # Tüm siparişleri al ve filtrele
            filtered_orders = []
            
            for order in orders:
                # order[11] müşteri adı (JOIN'den gelen), order[5] ürün adı, order[9] durum
                customer_name = order[11] if len(order) > 11 else ''
                product_name = order[5] if order[5] else ''
                status = order[9] if order[9] else ''
                
                if (search_term.lower() in customer_name.lower() or  # Müşteri adı
                    search_term.lower() in product_name.lower() or  # Ürün adı
                    search_term.lower() in status.lower()):         # Durum
                    filtered_orders.append(order)
            
            if not filtered_orders:
                messagebox.showwarning("Uyarı", "Arama sonucu bulunamadı!")
                return
            
            # DataFrame oluştur
            data = []
            for order in filtered_orders:
                data.append({
                    'ID': order[0],
                    'Müşteri Adı': order[1],
                    'Ürün Adı': order[2],
                    'Miktar': order[3],
                    'Birim Fiyat': f"{order[4]:.2f} TL" if order[4] is not None else "0.00 TL",
                    'Toplam Fiyat': f"{order[5]:.2f} TL" if order[5] is not None else "0.00 TL",
                    'Başlangıç Tarihi': self.format_date_for_display(order[6]),
                    'Bitiş Tarihi': self.format_date_for_display(order[7]),
                    'Durum': order[8],
                    'Oluşturma Tarihi': self.format_date_for_display(order[9])
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz (aynı stil ayarları)
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Filtrelenmiş Siparişler', index=False)
                
                # Stil ayarları
                workbook = writer.book
                worksheet = writer.sheets['Filtrelenmiş Siparişler']
                
                # Başlık stili
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Sütun genişliklerini ayarla
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Alternatif satır renkleri
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            messagebox.showinfo("Başarılı", f"Filtrelenmiş sipariş listesi export edildi!\n\nDosya: {filename}\n\nArama: '{search_term}'\nToplam: {len(filtered_orders)} sipariş")
            
        except Exception as e:
            messagebox.showerror("Export Hatası", f"Export sırasında hata oluştu:\n{str(e)}")
    
    def export_orders_by_status_to_excel(self, status):
        """Belirli durumdaki siparişleri Excel'e export eder"""
        try:
            # Dosya adı seç
            filename = filedialog.asksaveasfilename(
                title=f"{status} Durumundaki Siparişleri Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel dosyaları", "*.xlsx"), ("Tüm dosyalar", "*.*")],
                initialfile=f"{status.lower().replace(' ', '_')}_siparisler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not filename:
                return
            
            # Duruma göre sipariş verilerini al
            orders = self.db.get_orders_by_status(status)
            
            if not orders:
                messagebox.showwarning("Uyarı", f"'{status}' durumunda sipariş bulunamadı!")
                return
            
            # DataFrame oluştur
            data = []
            for order in orders:
                data.append({
                    'ID': order[0],
                    'Müşteri Adı': order[1],
                    'Ürün Adı': order[2],
                    'Miktar': order[3],
                    'Birim Fiyat': f"{order[4]:.2f} TL" if order[4] is not None else "0.00 TL",
                    'Toplam Fiyat': f"{order[5]:.2f} TL" if order[5] is not None else "0.00 TL",
                    'Başlangıç Tarihi': self.format_date_for_display(order[6]),
                    'Bitiş Tarihi': self.format_date_for_display(order[7]),
                    'Durum': order[8],
                    'Oluşturma Tarihi': self.format_date_for_display(order[9])
                })
            
            df = pd.DataFrame(data)
            
            # Excel'e yaz
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=f'{status} Siparişler', index=False)
                
                # Stil ayarları
                workbook = writer.book
                worksheet = writer.sheets[f'{status} Siparişler']
                
                # Başlık stili
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF", size=12)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Sütun genişliklerini ayarla
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column].width = adjusted_width
                
                # Alternatif satır renkleri
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            messagebox.showinfo("Başarılı", f"{status} durumundaki siparişler export edildi!\n\nDosya: {filename}\n\nDurum: {status}\nToplam: {len(orders)} sipariş")
            
        except Exception as e:
            messagebox.showerror("Export Hatası", f"Export sırasında hata oluştu:\n{str(e)}")
    
    # Optimize edilmiş fonksiyonlar
    
    def on_closing(self):
        """Uygulama kapatılırken çağrılır"""
        try:
            # Saati durdur
            self.stop_clock()
            print("Saat ve tarih durduruldu.")
            
            # Son otomatik kayıt
            self.auto_save_data("Program Kapatılıyor")
            self.show_auto_save_notification()
            
            # Kullanıcıya bilgi ver
            messagebox.showinfo("Otomatik Kayıt", "💾 Veriler otomatik olarak kaydedildi!\n\nProgram kapatılıyor...")
            
            # Otomatik bildirim kontrolü kaldırıldı
            
        except Exception as e:
            print(f"Uygulama kapatma hatası: {e}")
        
        # Uygulamayı kapat
        self.root.destroy()
    
    def load_customers(self):
        """Müşteri listesini yükler (cache ile optimize edilmiş)"""
        # Cache kontrolü (5 dakika)
        cache_age = (datetime.now() - self.cache_timestamp).total_seconds()
        if cache_age < 300 and self.customer_cache:  # 5 dakika cache
            customers = self.customer_cache
        else:
            customers = self.db.get_all_customers()
            self.customer_cache = customers
            self.cache_timestamp = datetime.now()
        
        # Mevcut listeyi temizle
        for item in self.customer_tree.get_children():
            self.customer_tree.delete(item)
        
        # Müşterileri listele
        for i, customer in enumerate(customers):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.customer_tree.insert('', 'end', values=(
                customer[0],  # ID (gizli)
                customer[1],  # Ad
                customer[2],  # E-posta
                customer[3],  # Telefon
                customer[5],  # Şirket
                customer[6]   # Kayıt tarihi
            ), tags=(tag,))
    
    def search_customers(self, event=None):
        """Müşteri arama"""
        search_term = self.search_var.get().strip()
        
        # Mevcut listeyi temizle
        for item in self.customer_tree.get_children():
            self.customer_tree.delete(item)
        
        if search_term:
            customers = self.db.search_customers(search_term)
        else:
            customers = self.db.get_all_customers()
        
        for i, customer in enumerate(customers):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.customer_tree.insert('', 'end', values=(
                customer[0],  # ID (gizli)
                customer[1], customer[2], 
                customer[3], customer[5], customer[6]
            ), tags=(tag,))
    
    def search_orders(self, event=None):
        """Sipariş arama"""
        search_term = self.order_search_var.get().strip()
        
        # Mevcut listeyi temizle
        for item in self.order_tree.get_children():
            self.order_tree.delete(item)
        
        if search_term:
            orders = self.db.search_orders(search_term)
        else:
            orders = self.db.get_all_orders()
        
        for i, order in enumerate(orders):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.order_tree.insert('', 'end', values=(
                order[11] if len(order) > 11 else "Bilinmiyor",  # Müşteri adı (12. sütun, indeks 11)
                order[5],  # Ürün adı
                order[6],  # Miktar
                f"{order[7]:.2f}",  # Birim fiyat
                f"{order[8]:.2f}",  # Toplam fiyat
                self.format_date_for_display(order[3]),  # Başlangıç tarihi
                self.format_date_for_display(order[4]),  # Bitiş tarihi
                order[9],  # Durum
                order[2]   # Sipariş tarihi
            ), tags=(tag,))
    
    def add_customer_dialog(self):
        """Yeni müşteri ekleme dialog'u"""
        dialog = tk.Toplevel(self.root)
        dialog.title("👤 Yeni Danışan Ekle")
        dialog.configure(bg='#f8f9fa')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        
        # Optimal boyut ayarla - çok daha dar
        dialog.geometry("800x1000+100+30")
        dialog.minsize(700, 800)
        
        # Modern scrollable container - minimal padding
        container_frame = tk.Frame(dialog, bg='#f8f9fa')
        container_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Canvas for scrolling
        canvas = tk.Canvas(container_frame, bg='#f8f9fa', highlightthickness=0, bd=0)
        
        # Modern custom scrollbar frame - minimal width
        scrollbar_frame = tk.Frame(container_frame, bg='#f8f9fa', width=8)
        scrollbar_bg = tk.Canvas(scrollbar_frame, bg='#f8f9fa', width=8, highlightthickness=0, bd=0)
        
        # Scrollable content frame
        scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
        
        # Configure canvas scrolling
        def configure_scroll_region(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            
        scrollable_frame.bind("<Configure>", configure_scroll_region)
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Update canvas width when container resizes
        def configure_canvas_width(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind('<Configure>', configure_canvas_width)
        
        # Modern smooth mouse wheel scrolling
        def smooth_scroll(event):
            # Smooth scroll with smaller increments
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    update_modern_scrollbar()
            except tk.TclError:
                pass
        
        # Modern scrollbar visual
        scrollbar_thumb = None
        
        def update_modern_scrollbar():
            nonlocal scrollbar_thumb
            # Clear previous thumb
            scrollbar_bg.delete("thumb")
            
            # Get scroll position
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:  # No scroll needed
                    return
                    
                # Calculate thumb position and size
                scrollbar_height = scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                # Draw modern thumb - adjusted for smaller width
                scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        # Scrollbar drag variables
        self.dialog_scrollbar_dragging = False
        
        # Scrollbar drag functionality
        def on_scrollbar_click(event):
            if canvas.yview() == (0.0, 1.0):  # No scroll needed
                return
            
            scrollbar_height = scrollbar_bg.winfo_height()
            click_position = max(0, min(1, event.y / scrollbar_height))
            canvas.yview_moveto(click_position)
            update_modern_scrollbar()
            
            # Start drag
            self.dialog_scrollbar_dragging = True
        
        def on_scrollbar_drag(event):
            if not self.dialog_scrollbar_dragging:
                return
            if canvas.yview() == (0.0, 1.0):
                return
            
            scrollbar_height = scrollbar_bg.winfo_height()
            click_position = max(0, min(1, event.y / scrollbar_height))
            canvas.yview_moveto(click_position)
            update_modern_scrollbar()
        
        def on_scrollbar_release(event):
            self.dialog_scrollbar_dragging = False
        
        # Hover effects for modern scrollbar
        def on_scrollbar_enter(event):
            scrollbar_bg.delete("thumb")
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_scrollbar_leave(event):
            update_modern_scrollbar()
        
        # Bind events
        def safe_close_dialog():
            try:
                dialog.unbind('<MouseWheel>')
                dialog.unbind('<Return>')
                dialog.unbind('<Escape>')
            except:
                pass
            dialog.destroy()
        
        try:
            dialog.bind("<MouseWheel>", smooth_scroll)
        except:
            pass
        
        dialog.protocol("WM_DELETE_WINDOW", safe_close_dialog)
        scrollbar_bg.bind("<Button-1>", on_scrollbar_click)
        scrollbar_bg.bind("<B1-Motion>", on_scrollbar_drag)
        scrollbar_bg.bind("<ButtonRelease-1>", on_scrollbar_release)
        scrollbar_bg.bind("<Enter>", on_scrollbar_enter)
        scrollbar_bg.bind("<Leave>", on_scrollbar_leave)
        
        # Make scrollbar focusable
        scrollbar_bg.configure(takefocus=True)
        
        # Update scrollbar when canvas scrolls
        def on_canvas_configure(event=None):
            canvas.after_idle(update_modern_scrollbar)
        
        canvas.bind("<Configure>", on_canvas_configure)
        
        # Pack widgets - no extra padding on scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_frame.pack(side="right", fill="y", padx=0)
        scrollbar_frame.pack_propagate(False)  # Keep fixed width
        scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize scrollbar
        canvas.after(100, update_modern_scrollbar)
        
        # Dialog kapatma event'ini yakala
        def on_close():
            try:
                canvas.unbind_all("<MouseWheel>")
            except:
                pass
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        # Ana container - Uygun padding (artık scrollable_frame içinde)
        main_container = tk.Frame(scrollable_frame, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Header card
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 25))
        
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(header_inner,
                text="👤 Yeni Danışan Ekle",
                font=('Segoe UI', 22, 'bold'),
                bg='white',
                fg='#2c3e50').pack()
        
        tk.Label(header_inner,
                text="Danışan bilgilerini eksiksiz doldurun",
                font=('Segoe UI', 12),
                bg='white',
                fg='#7f8c8d').pack(pady=(10, 0))
        
        # Form card
        form_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        form_card.pack(fill=tk.BOTH, expand=True, pady=(0, 25))
        
        form_inner = tk.Frame(form_card, bg='white')
        form_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Form başlığı
        tk.Label(form_inner,
                text="Danışan Bilgileri",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(0, 25))
        
        # Form container - Grid layout
        form_container = tk.Frame(form_inner, bg='white')
        form_container.pack(fill=tk.BOTH, expand=True)
        
        # İlk satır - Ad Soyad ve Telefon
        row1 = tk.Frame(form_container, bg='white')
        row1.pack(fill=tk.X, pady=(0, 30))
        
        # Ad Soyad * (Sol)
        name_col = tk.Frame(row1, bg='white')
        name_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 20))
        
        tk.Label(name_col, text="👤 Ad Soyad *", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#e74c3c').pack(anchor='w', pady=(0, 8))
        name_entry = tk.Entry(name_col, 
                            font=('Segoe UI', 12),
                            relief='flat', bd=0,
                            highlightthickness=2,
                            highlightcolor='#27ae60',
                            bg='#f8f9fa',
                            insertbackground='#27ae60')
        name_entry.pack(fill=tk.X, ipady=12)
        
        # Telefon * (Sağ)
        phone_col = tk.Frame(row1, bg='white')
        phone_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(20, 0))
        
        tk.Label(phone_col, text="📱 Telefon *", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#e74c3c').pack(anchor='w', pady=(0, 8))
        vcmd = (dialog.register(self.validate_phone_input), '%P')
        phone_entry = tk.Entry(phone_col, 
                             font=('Segoe UI', 12),
                             relief='flat', bd=0,
                             validate='key', validatecommand=vcmd,
                             highlightthickness=2,
                             highlightcolor='#27ae60',
                             bg='#f8f9fa',
                             insertbackground='#27ae60')
        phone_entry.pack(fill=tk.X, ipady=12)
        
        # İkinci satır - E-posta
        row2 = tk.Frame(form_container, bg='white')
        row2.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row2, text="📧 E-posta", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        email_entry = tk.Entry(row2, 
                             font=('Segoe UI', 12),
                             relief='flat', bd=0,
                             highlightthickness=2,
                             highlightcolor='#27ae60',
                             bg='#f8f9fa',
                             insertbackground='#27ae60')
        email_entry.pack(fill=tk.X, ipady=12)
        
        # Üçüncü satır - Adres
        row3 = tk.Frame(form_container, bg='white')
        row3.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row3, text="🏠 Adres", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        address_entry = tk.Entry(row3, 
                               font=('Segoe UI', 12),
                               relief='flat', bd=0,
                               highlightthickness=2,
                               highlightcolor='#27ae60',
                               bg='#f8f9fa',
                               insertbackground='#27ae60')
        address_entry.pack(fill=tk.X, ipady=12)
        
        # Dördüncü satır - Şirket
        row4 = tk.Frame(form_container, bg='white')
        row4.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row4, text="🏢 Şirket", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        company_entry = tk.Entry(row4, 
                               font=('Segoe UI', 12),
                               relief='flat', bd=0,
                               highlightthickness=2,
                               highlightcolor='#27ae60',
                               bg='#f8f9fa',
                               insertbackground='#27ae60')
        company_entry.pack(fill=tk.X, ipady=12)
        
        # Beşinci satır - Notlar
        row5 = tk.Frame(form_container, bg='white')
        row5.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row5, text="📝 Notlar", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        notes_text = tk.Text(row5, 
                           height=4,
                           font=('Segoe UI', 11),
                           relief='flat', bd=0,
                           highlightthickness=2,
                           highlightcolor='#27ae60',
                           bg='#f8f9fa',
                           insertbackground='#27ae60',
                           wrap=tk.WORD)
        notes_text.pack(fill=tk.X)
        
        def save_customer():
            name = name_entry.get().strip()
            if not name:
                messagebox.showerror("Hata", "Ad Soyad alanı zorunludur!")
                name_entry.focus()
                return
            
            phone = phone_entry.get().strip()
            if not phone:
                messagebox.showerror("Hata", "Telefon alanı zorunludur!")
                phone_entry.focus()
                return
            
            # Telefon numarası uzunluk kontrolü
            if len(phone) < 10:
                messagebox.showerror("Hata", "Telefon numarası en az 10 haneli olmalıdır!")
                phone_entry.focus()
                return
            
            email = email_entry.get().strip()
            address = address_entry.get().strip()
            company = company_entry.get().strip()
            notes = notes_text.get("1.0", tk.END).strip()
            
            try:
                self.db.add_customer(name, email, phone, address, company, notes)
                self.clear_cache()  # Cache'i temizle
                messagebox.showinfo("Başarılı", "Müşteri başarıyla eklendi!")
                
                # Otomatik kayıt
                self.auto_save_data("Müşteri Eklendi")
                
                dialog.destroy()
                self.load_customers()
            except ValueError as e:
                # Aynı isimde müşteri varsa özel hata mesajı
                messagebox.showerror("Müşteri Hatası", f"Böyle bir müşteriniz var sipariş açın!\n\n{str(e)}")
            except Exception as e:
                messagebox.showerror("Hata", f"Müşteri eklenirken hata oluştu: {str(e)}")
        
        # Button card - Alt kısım
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X, pady=(0, 0))
        
        button_inner = tk.Frame(button_card, bg='white')
        button_inner.pack(fill=tk.X, padx=50, pady=(30, 40))
        
        # Zorunlu alan uyarısı
        tk.Label(button_inner,
                text="* Zorunlu alanlar",
                font=('Segoe UI', 10, 'italic'),
                bg='white',
                fg='#e74c3c').pack(anchor='w', pady=(0, 20))
        
        # Button container - Sağa hizalı
        btn_container = tk.Frame(button_inner, bg='white')
        btn_container.pack(anchor='e')
        
        # İptal butonu - Kırmızı
        cancel_btn = tk.Button(btn_container, 
                             text="✖ İptal",
                             command=dialog.destroy,
                             font=('Segoe UI', 11, 'bold'),
                             bg='#e74c3c', 
                             fg='white',
                             relief='flat',
                             borderwidth=0,
                             padx=30,
                             pady=12,
                             cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Kaydet butonu
        save_btn = tk.Button(btn_container, 
                           text="✓ Danışan Ekle",
                           command=save_customer,
                           font=('Segoe UI', 11, 'bold'),
                           bg='#27ae60', 
                           fg='white',
                           relief='flat',
                           borderwidth=0,
                           padx=30,
                           pady=12,
                           cursor='hand2')
        save_btn.pack(side=tk.LEFT)
        
        # Hover efektleri
        def on_save_hover(e):
            save_btn.config(bg='#219a52')
            
        def on_save_leave(e):
            save_btn.config(bg='#27ae60')
            
        def on_cancel_hover(e):
            cancel_btn.config(bg='#c0392b')
            
        def on_cancel_leave(e):
            cancel_btn.config(bg='#e74c3c')
            
        save_btn.bind('<Enter>', on_save_hover)
        save_btn.bind('<Leave>', on_save_leave)
        cancel_btn.bind('<Enter>', on_cancel_hover)
        cancel_btn.bind('<Leave>', on_cancel_leave)
        
        # Focus ve kısayollar
        name_entry.focus()
        dialog.bind('<Return>', lambda e: save_customer())
        def safe_close(e=None):
            try:
                dialog.unbind('<MouseWheel>')
                dialog.unbind('<Return>')
                dialog.unbind('<Escape>')
            except:
                pass
            dialog.destroy()
        dialog.bind('<Escape>', safe_close)
    
    def edit_customer(self):
        """Müşteri düzenleme"""
        selection = self.customer_tree.selection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen düzenlenecek müşteriyi seçin!")
            return
        
        customer_id = self.customer_tree.item(selection[0])['values'][0]
        customers = self.db.get_all_customers()
        customer = None
        
        for c in customers:
            if c[0] == customer_id:
                customer = c
                break
        
        if not customer:
            messagebox.showerror("Hata", "Müşteri bulunamadı!")
            return
        
        # Düzenleme dialog'u - Modern tasarım
        dialog = tk.Toplevel(self.root)
        dialog.title("✏️ Danışan Düzenle")
        dialog.configure(bg='#f8f9fa')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        
        # Optimal boyut ayarla - "Yeni Danışan Ekle" ile aynı
        dialog.geometry("800x1000")
        dialog.minsize(700, 800)
        
        # Pencereyi merkeze yerleştir
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (1000 // 2)
        dialog.geometry(f"800x1000+{x}+{y}")
        
        # Modern scrollable container - minimal padding
        container_frame = tk.Frame(dialog, bg='#f8f9fa')
        container_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Canvas for scrolling
        canvas = tk.Canvas(container_frame, bg='#f8f9fa', highlightthickness=0, bd=0)
        
        # Modern custom scrollbar frame - minimal width
        scrollbar_frame = tk.Frame(container_frame, bg='#f8f9fa', width=8)
        scrollbar_bg = tk.Canvas(scrollbar_frame, bg='#f8f9fa', width=8, highlightthickness=0, bd=0)
        
        # Scrollable content frame
        scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
        
        # Mouse wheel scrolling with smooth effect
        def smooth_scroll(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                    update_modern_scrollbar()
            except tk.TclError:
                pass
        
        # Modern scrollbar visual
        scrollbar_thumb = None
        
        def update_modern_scrollbar():
            nonlocal scrollbar_thumb
            # Clear previous thumb
            scrollbar_bg.delete("thumb")
            
            # Get scroll position
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return  # No scrolling needed
                
                # Calculate thumb position and size
                scrollbar_height = scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                # Draw modern thumb - minimal width
                scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        # Scrollbar interaction variables
        scrollbar_dragging = False
        
        def on_scrollbar_click(event):
            nonlocal scrollbar_dragging
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                
                scrollbar_height = scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                canvas.yview_moveto(click_position)
                update_modern_scrollbar()
                scrollbar_dragging = True
            except:
                pass
        
        def on_scrollbar_drag(event):
            if not scrollbar_dragging:
                return
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                
                scrollbar_height = scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                canvas.yview_moveto(click_position)
                update_modern_scrollbar()
            except:
                pass
        
        def on_scrollbar_release(event):
            nonlocal scrollbar_dragging
            scrollbar_dragging = False
        
        def on_scrollbar_enter(event):
            scrollbar_bg.delete("thumb")
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_scrollbar_leave(event):
            update_modern_scrollbar()
        
        # Bind events
        def safe_close_dialog():
            try:
                dialog.unbind('<MouseWheel>')
                dialog.unbind('<Return>')
                dialog.unbind('<Escape>')
            except:
                pass
            dialog.destroy()
        
        try:
            dialog.bind("<MouseWheel>", smooth_scroll)
        except:
            pass
        
        dialog.protocol("WM_DELETE_WINDOW", safe_close_dialog)
        scrollbar_bg.bind("<Button-1>", on_scrollbar_click)
        scrollbar_bg.bind("<B1-Motion>", on_scrollbar_drag)
        scrollbar_bg.bind("<ButtonRelease-1>", on_scrollbar_release)
        scrollbar_bg.bind("<Enter>", on_scrollbar_enter)
        scrollbar_bg.bind("<Leave>", on_scrollbar_leave)
        
        # Make scrollbar focusable
        scrollbar_bg.configure(takefocus=True)
        
        # Update scrollbar when canvas scrolls
        def on_canvas_configure(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.after_idle(update_modern_scrollbar)
        
        canvas.bind("<Configure>", on_canvas_configure)
        scrollable_frame.bind("<Configure>", on_canvas_configure)
        
        # Create canvas window
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Update canvas window width when canvas width changes
        def on_canvas_configure_width(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        canvas.bind("<Configure>", on_canvas_configure_width)
        
        # Pack widgets - no extra padding on scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_frame.pack(side="right", fill="y", padx=0)
        scrollbar_frame.pack_propagate(False)  # Keep fixed width
        scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize scrollbar
        canvas.after(100, update_modern_scrollbar)
        
        # Dialog kapatma event'ini yakala
        def on_close():
            try:
                canvas.unbind_all("<MouseWheel>")
            except:
                pass
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        # Ana container - Uygun padding (artık scrollable_frame içinde)
        main_container = tk.Frame(scrollable_frame, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Header card
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 25))
        
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(header_inner,
                text="✏️ Danışan Düzenle",
                font=('Segoe UI', 22, 'bold'),
                bg='white',
                fg='#2c3e50').pack()
        
        tk.Label(header_inner,
                text="Danışan bilgilerini güncelleyin",
                font=('Segoe UI', 12),
                bg='white',
                fg='#7f8c8d').pack(pady=(10, 0))
        
        # Form card
        form_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        form_card.pack(fill=tk.BOTH, expand=True, pady=(0, 25))
        
        form_inner = tk.Frame(form_card, bg='white')
        form_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Form başlığı
        tk.Label(form_inner,
                text="Danışan Bilgileri",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(0, 25))
        
        # Form container - Grid layout
        form_container = tk.Frame(form_inner, bg='white')
        form_container.pack(fill=tk.BOTH, expand=True)
        
        # İlk satır - Ad Soyad ve Telefon
        row1 = tk.Frame(form_container, bg='white')
        row1.pack(fill=tk.X, pady=(0, 30))
        
        # Ad Soyad * (Sol)
        name_col = tk.Frame(row1, bg='white')
        name_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 20))
        
        tk.Label(name_col, text="👤 Ad Soyad *", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#e74c3c').pack(anchor='w', pady=(0, 8))
        name_entry = tk.Entry(name_col, 
                            font=('Segoe UI', 12),
                            relief='flat', bd=0,
                            highlightthickness=2,
                            highlightcolor='#f39c12',
                            bg='#f8f9fa',
                            insertbackground='#f39c12')
        name_entry.insert(0, customer[1])
        name_entry.pack(fill=tk.X, ipady=12)
        
        # Telefon * (Sağ)
        phone_col = tk.Frame(row1, bg='white')
        phone_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(20, 0))
        
        tk.Label(phone_col, text="📱 Telefon *", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#e74c3c').pack(anchor='w', pady=(0, 8))
        vcmd = (dialog.register(self.validate_phone_input), '%P')
        phone_entry = tk.Entry(phone_col, 
                             font=('Segoe UI', 12),
                             relief='flat', bd=0,
                             validate='key', validatecommand=vcmd,
                             highlightthickness=2,
                             highlightcolor='#f39c12',
                             bg='#f8f9fa',
                             insertbackground='#f39c12')
        phone_entry.insert(0, customer[3] or "")
        phone_entry.pack(fill=tk.X, ipady=12)
        
        # İkinci satır - E-posta
        row2 = tk.Frame(form_container, bg='white')
        row2.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row2, text="📧 E-posta", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        email_entry = tk.Entry(row2, 
                             font=('Segoe UI', 12),
                             relief='flat', bd=0,
                             highlightthickness=2,
                             highlightcolor='#f39c12',
                             bg='#f8f9fa',
                             insertbackground='#f39c12')
        email_entry.insert(0, customer[2] or "")
        email_entry.pack(fill=tk.X, ipady=12)
        
        # Üçüncü satır - Adres
        row3 = tk.Frame(form_container, bg='white')
        row3.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row3, text="🏠 Adres", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        address_entry = tk.Entry(row3, 
                               font=('Segoe UI', 12),
                               relief='flat', bd=0,
                               highlightthickness=2,
                               highlightcolor='#f39c12',
                               bg='#f8f9fa',
                               insertbackground='#f39c12')
        address_entry.insert(0, customer[4] or "")
        address_entry.pack(fill=tk.X, ipady=12)
        
        # Dördüncü satır - Şirket
        row4 = tk.Frame(form_container, bg='white')
        row4.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row4, text="🏢 Şirket", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        company_entry = tk.Entry(row4, 
                               font=('Segoe UI', 12),
                               relief='flat', bd=0,
                               highlightthickness=2,
                               highlightcolor='#f39c12',
                               bg='#f8f9fa',
                               insertbackground='#f39c12')
        company_entry.insert(0, customer[5] or "")
        company_entry.pack(fill=tk.X, ipady=12)
        
        # Beşinci satır - Notlar
        row5 = tk.Frame(form_container, bg='white')
        row5.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(row5, text="📝 Notlar", 
                font=('Segoe UI', 11, 'bold'), 
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 8))
        notes_text = tk.Text(row5, 
                           height=4,
                           font=('Segoe UI', 11),
                           relief='flat', bd=0,
                           highlightthickness=2,
                           highlightcolor='#f39c12',
                           bg='#f8f9fa',
                           insertbackground='#f39c12',
                           wrap=tk.WORD)
        notes_text.insert("1.0", customer[7] or "")
        notes_text.pack(fill=tk.X)
        
        def update_customer():
            name = name_entry.get().strip()
            if not name:
                messagebox.showerror("Hata", "Ad Soyad alanı zorunludur!")
                name_entry.focus()
                return
            
            phone = phone_entry.get().strip()
            if not phone:
                messagebox.showerror("Hata", "Telefon alanı zorunludur!")
                phone_entry.focus()
                return
            
            # Telefon numarası uzunluk kontrolü
            if len(phone) < 10:
                messagebox.showerror("Hata", "Telefon numarası en az 10 haneli olmalıdır!")
                phone_entry.focus()
                return
            
            email = email_entry.get().strip()
            address = address_entry.get().strip()
            company = company_entry.get().strip()
            notes = notes_text.get("1.0", tk.END).strip()
            
            try:
                self.db.update_customer(customer_id, name, email, phone, address, company, notes)
                self.clear_cache()  # Cache'i temizle
                messagebox.showinfo("Başarılı", "Müşteri başarıyla güncellendi!")
                
                # Otomatik kayıt
                self.auto_save_data("Müşteri Güncellendi")
                
                dialog.destroy()
                self.load_customers()
            except Exception as e:
                messagebox.showerror("Hata", f"Müşteri güncellenirken hata oluştu: {str(e)}")
        
        # Button card - Alt kısım
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X, pady=(0, 0))
        
        button_inner = tk.Frame(button_card, bg='white')
        button_inner.pack(fill=tk.X, padx=50, pady=(30, 40))
        
        # Zorunlu alan uyarısı
        tk.Label(button_inner,
                text="* Zorunlu alanlar",
                font=('Segoe UI', 10, 'italic'),
                bg='white',
                fg='#e74c3c').pack(anchor='w', pady=(0, 20))
        
        # Button container - Sağa hizalı
        btn_container = tk.Frame(button_inner, bg='white')
        btn_container.pack(anchor='e')
        
        # İptal butonu - Kırmızı
        cancel_btn = tk.Button(btn_container, 
                             text="✖ İptal",
                             command=dialog.destroy,
                             font=('Segoe UI', 11, 'bold'),
                             bg='#e74c3c', 
                             fg='white',
                             relief='flat',
                             borderwidth=0,
                             padx=30,
                             pady=12,
                             cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Güncelle butonu - Turuncu tema
        update_btn = tk.Button(btn_container, 
                             text="✓ Güncelle",
                             command=update_customer,
                             font=('Segoe UI', 11, 'bold'),
                             bg='#f39c12', 
                             fg='white',
                             relief='flat',
                             borderwidth=0,
                             padx=30,
                             pady=12,
                             cursor='hand2')
        update_btn.pack(side=tk.LEFT)
        
        # Hover efektleri
        def on_update_hover(e):
            update_btn.config(bg='#e67e22')
            
        def on_update_leave(e):
            update_btn.config(bg='#f39c12')
            
        def on_cancel_hover(e):
            cancel_btn.config(bg='#c0392b')
            
        def on_cancel_leave(e):
            cancel_btn.config(bg='#e74c3c')
            
        update_btn.bind('<Enter>', on_update_hover)
        update_btn.bind('<Leave>', on_update_leave)
        cancel_btn.bind('<Enter>', on_cancel_hover)
        cancel_btn.bind('<Leave>', on_cancel_leave)
        
        # Focus ve kısayollar
        name_entry.focus()
        dialog.bind('<Return>', lambda e: update_customer())
        def safe_close(e=None):
            try:
                dialog.unbind('<MouseWheel>')
                dialog.unbind('<Return>')
                dialog.unbind('<Escape>')
            except:
                pass
            dialog.destroy()
        dialog.bind('<Escape>', safe_close)
    
    def delete_customer(self):
        """Müşteri silme"""
        selection = self.customer_tree.selection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen silinecek müşteriyi seçin!")
            return
        
        customer_id = self.customer_tree.item(selection[0])['values'][0]
        customer_name = self.customer_tree.item(selection[0])['values'][1]
        
        # Müşterinin siparişlerini kontrol et
        try:
            customer_orders = self.db.get_customer_orders(customer_id)
            if customer_orders:
                # Müşterinin siparişi varsa silmeyi engelle
                order_count = len(customer_orders)
                messagebox.showwarning(
                    "Silme Engellendi", 
                    f"'{customer_name}' müşterisi silinemez!\n\n"
                    f"Bu müşteriye ait {order_count} adet sipariş bulunmaktadır.\n"
                    f"Müşteriyi silmek için önce tüm siparişlerini silmeniz gerekir.\n\n"
                    f"Siparişler sekmesinden müşterinin siparişlerini görüntüleyebilir ve silebilirsiniz."
                )
                return
        except Exception as e:
            messagebox.showerror("Hata", f"Müşteri siparişleri kontrol edilirken hata oluştu: {str(e)}")
            return
        
        result = messagebox.askyesno("Onay", f"'{customer_name}' müşterisini silmek istediğinizden emin misiniz?\n\nBu işlem geri alınamaz!")
        
        if result:
            try:
                self.db.delete_customer(customer_id)
                self.clear_cache()  # Cache'i temizle
                messagebox.showinfo("Başarılı", "Müşteri başarıyla silindi!")
                
                # Otomatik kayıt
                self.auto_save_data("Müşteri Silindi")
                
                self.load_customers()
                self.load_stats()
            except Exception as e:
                messagebox.showerror("Hata", f"Müşteri silinirken hata oluştu: {str(e)}")
    
    def on_customer_double_click(self, event):
        """Müşteriye çift tıklandığında sipariş ekleme"""
        selection = self.customer_tree.selection()
        if selection:
            self.add_order_dialog()
    
    def add_order_dialog(self):
        """Yeni sipariş ekleme dialog'u"""
        # Müşteri seçimi kontrolü
        selection = self.customer_tree.selection()
        if not selection:
            # Müşteri seçili değilse, müşteri seçme dialog'u göster
            customer = self.select_customer_dialog()
            if not customer:
                return
            customer_id, customer_name = customer
        else:
            customer_id = self.customer_tree.item(selection[0])['values'][0]
            customer_name = self.customer_tree.item(selection[0])['values'][1]
        
        # Modern Yeni Sipariş penceresi - Daha büyük boyut
        dialog = tk.Toplevel(self.root)
        dialog.title(f"📝 Yeni Sipariş - {customer_name}")
        dialog.geometry("1050x1150")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        dialog.resizable(True, True)
        
        # Pencereyi merkeze yerleştir
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (1050 // 2)
        y = (dialog.winfo_screenheight() // 2) - (1150 // 2)
        dialog.geometry(f"1050x1150+{x}+{y}")
        
        # Ana container - Daha geniş padding
        main_container = tk.Frame(dialog, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        # Header card - Müşteri bilgisi
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 20))
        
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=40, pady=30)
        
        tk.Label(header_inner,
                text="📝 Yeni Sipariş Oluştur",
                font=('Segoe UI', 22, 'bold'),
                bg='white',
                fg='#2c3e50').pack()
        
        tk.Label(header_inner,
                text=f"👤 Müşteri: {customer_name}",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#27ae60').pack(pady=(8, 0))
        
        # Form kartı
        form_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        form_card.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        form_inner = tk.Frame(form_card, bg='white')
        form_inner.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        # Ürün bilgileri bölümü
        tk.Label(form_inner,
                text="🛍️ Ürün Bilgileri",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        # Ürün adı
        product_frame = tk.Frame(form_inner, bg='white')
        product_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(product_frame,
                text="Ürün/Hizmet Adı *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        product_entry = tk.Entry(product_frame,
                               font=('Segoe UI', 12),
                               relief='flat',
                               bd=0,
                               bg='#f8f9fa',
                               fg='#2c3e50',
                               insertbackground='#3498db')
        product_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)
        
        # Miktar ve Fiyat - Yan yana
        quantity_price_frame = tk.Frame(form_inner, bg='white')
        quantity_price_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Sol: Miktar
        left_frame = tk.Frame(quantity_price_frame, bg='white')
        left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(left_frame,
                text="Miktar *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        quantity_entry = tk.Entry(left_frame,
                                font=('Segoe UI', 12),
                                relief='flat',
                                bd=0,
                                bg='#f8f9fa',
                                fg='#2c3e50',
                                insertbackground='#3498db')
        quantity_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)
        
        # Sağ: Birim Fiyat
        right_frame = tk.Frame(quantity_price_frame, bg='white')
        right_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        
        tk.Label(right_frame,
                text="Birim Fiyat (₺) *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        price_entry = tk.Entry(right_frame,
                             font=('Segoe UI', 12),
                             relief='flat',
                             bd=0,
                             bg='#f8f9fa',
                             fg='#2c3e50',
                             insertbackground='#3498db')
        price_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)
        
        # Tarih alanları bölümü
        tk.Label(form_inner,
                text="📅 Tarih Bilgileri",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(25, 15))
        
        # Başlangıç ve Bitiş tarihi - Yan yana
        dates_frame = tk.Frame(form_inner, bg='white')
        dates_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Sol: Başlangıç tarihi
        start_date_container = tk.Frame(dates_frame, bg='white')
        start_date_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(start_date_container,
                text="Başlangıç Tarihi *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        start_date_input_frame = tk.Frame(start_date_container, bg='white')
        start_date_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        start_date_entry = tk.Entry(start_date_input_frame,
                                  font=('Segoe UI', 12),
                                  relief='flat',
                                  bd=0,
                                  bg='#f8f9fa',
                                  fg='#2c3e50',
                                  insertbackground='#3498db')
        start_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(0, 8))
        
        # Modern takvim butonu
        def open_start_date_picker():
            self.show_modern_date_picker(start_date_entry, "Başlangıç Tarihi Seç")
        
        start_date_btn = tk.Button(start_date_input_frame,
                                 text="📅",
                                 command=open_start_date_picker,
                                 font=('Segoe UI', 12),
                                 bg='#3498db',
                                 fg='white',
                                 relief='flat',
                                 borderwidth=0,
                                 padx=8,
                                 pady=8,
                                 cursor='hand2',
                                 activebackground='#2980b9',
                                 activeforeground='white')
        start_date_btn.pack(side=tk.RIGHT)
        
        # Sağ: Bitiş tarihi
        end_date_container = tk.Frame(dates_frame, bg='white')
        end_date_container.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        
        tk.Label(end_date_container,
                text="Bitiş Tarihi *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        end_date_input_frame = tk.Frame(end_date_container, bg='white')
        end_date_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        end_date_entry = tk.Entry(end_date_input_frame,
                                font=('Segoe UI', 12),
                                relief='flat',
                                bd=0,
                                bg='#f8f9fa',
                                fg='#2c3e50',
                                insertbackground='#3498db')
        end_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(0, 8))
        
        # Modern takvim butonu
        def open_end_date_picker():
            self.show_modern_date_picker(end_date_entry, "Bitiş Tarihi Seç")
        
        end_date_btn = tk.Button(end_date_input_frame,
                               text="📅",
                               command=open_end_date_picker,
                               font=('Segoe UI', 12),
                               bg='#e74c3c',
                               fg='white',
                               relief='flat',
                               borderwidth=0,
                               padx=8,
                               pady=8,
                               cursor='hand2',
                               activebackground='#c0392b',
                               activeforeground='white')
        end_date_btn.pack(side=tk.RIGHT)
        
        # Varsayılan tarihleri ayarla
        from datetime import datetime, timedelta
        today = datetime.now().date()
        start_date_entry.insert(0, today.strftime("%d.%m.%Y"))
        end_date_entry.insert(0, (today + timedelta(days=7)).strftime("%d.%m.%Y"))
        
        # Durum bölümü
        tk.Label(form_inner,
                text="📊 Sipariş Durumu",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(25, 15))
        
        status_frame = tk.Frame(form_inner, bg='white')
        status_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(status_frame,
                text="Başlangıç Durumu",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w', pady=(0, 10))
        
        # Modern durum seçim butonları
        status_var = tk.StringVar(value="Beklemede")
        status_options = [
            ("⏳ Beklemede", "Beklemede", "#f39c12"),
            ("🔧 Hazırlanıyor", "Hazırlanıyor", "#3498db"),
            ("✅ Tamamlandı", "Tamamlandı", "#27ae60"),
            ("❌ İptal", "İptal", "#e74c3c")
        ]
        
        status_buttons_frame = tk.Frame(status_frame, bg='white')
        status_buttons_frame.pack(fill=tk.X, pady=(5, 0))
        
        status_buttons = []
        
        def select_status(status_value):
            status_var.set(status_value)
            # Buton renklerini güncelle
            for btn, (_, btn_status, btn_color) in zip(status_buttons, status_options):
                if btn_status == status_value:
                    btn.configure(bg=btn_color, fg='white', relief='solid', bd=2)
                else:
                    btn.configure(bg='#ecf0f1', fg='#34495e', relief='flat', bd=0)
        
        # Durum butonlarını oluştur
        for i, (display_text, status_value, color) in enumerate(status_options):
            btn = tk.Button(status_buttons_frame,
                          text=display_text,
                          command=lambda s=status_value: select_status(s),
                          font=('Segoe UI', 11, 'bold'),
                          bg='#ecf0f1' if status_value != "Beklemede" else color,
                          fg='#34495e' if status_value != "Beklemede" else 'white',
                          relief='flat' if status_value != "Beklemede" else 'solid',
                          bd=0 if status_value != "Beklemede" else 2,
                          padx=15,
                          pady=8,
                          cursor='hand2',
                          activebackground='#bdc3c7',
                          activeforeground='#2c3e50')
            
            # 2x2 grid düzeni
            row = i // 2
            col = i % 2
            btn.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
            status_buttons.append(btn)
            
            # Hover efektleri
            def on_enter(e, original_color=color):
                if e.widget['bg'] != '#2c3e50':  # Seçili değilse
                    e.widget.configure(bg=original_color, fg='white')
            
            def on_leave(e):
                current_status = status_var.get()
                btn_index = status_buttons.index(e.widget)
                btn_status = status_options[btn_index][1]
                if btn_status != current_status:
                    e.widget.configure(bg='#ecf0f1', fg='#34495e')
                else:
                    # Seçili buton kendi renginde kalır
                    btn_color = status_options[btn_index][2]
                    e.widget.configure(bg=btn_color, fg='white')
            
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)
        
        # Grid ağırlıklarını ayarla
        status_buttons_frame.columnconfigure(0, weight=1)
        status_buttons_frame.columnconfigure(1, weight=1)
        
        def save_order():
            product = product_entry.get().strip()
            if not product:
                messagebox.showerror("Hata", "Ürün adı zorunludur!")
                return
            
            try:
                quantity = int(quantity_entry.get())
                price = float(price_entry.get())
            except ValueError:
                messagebox.showerror("Hata", "Miktar ve fiyat sayısal değer olmalıdır!")
                return
            
            # Tarih doğrulama
            start_date = start_date_entry.get().strip()
            end_date = end_date_entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showerror("Hata", "Başlangıç ve bitiş tarihi zorunludur!")
                return
            
            # Tarihleri veritabanı formatına çevir
            start_date_db = self.format_date_for_database(start_date)
            end_date_db = self.format_date_for_database(end_date)
            
            if not start_date_db or not end_date_db:
                messagebox.showerror("Hata", "Tarih formatı GG.AA.YYYY olmalıdır! (Örnek: 27.07.2025)")
                return
            
            # Bitiş tarihi başlangıç tarihinden sonra olmalı
            if start_date_db >= end_date_db:
                messagebox.showerror("Hata", "Bitiş tarihi başlangıç tarihinden sonra olmalıdır!")
                return
            
            status = status_var.get()
            
            try:
                self.db.add_order(customer_id, product, quantity, price, start_date_db, end_date_db, status)
                messagebox.showinfo("Başarılı", "Sipariş başarıyla eklendi!")
                
                # Otomatik kayıt
                self.auto_save_data("Sipariş Eklendi")
                
                dialog.destroy()
                self.clear_cache()  # Cache'i temizle
                self.load_orders()
                self.load_stats()
            except Exception as e:
                messagebox.showerror("Hata", f"Sipariş eklenirken hata oluştu: {str(e)}")
        
        # Modern buton kartı
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X, pady=(0, 0))
        
        button_container = tk.Frame(button_card, bg='white')
        button_container.pack(pady=40)
        
        # Modern buton stilleri - büyütülmüş
        modern_button_style = {
            'font': ('Segoe UI', 16, 'bold'),
            'relief': 'flat',
            'borderwidth': 0,
            'padx': 45,
            'pady': 16,
            'cursor': 'hand2'
        }
        
        # Kaydet butonu (Yeşil)
        save_btn = tk.Button(button_container,
                           text="💾 Siparişi Kaydet",
                           command=save_order,
                           bg='#27ae60',
                           fg='white',
                           activebackground='#229954',
                           activeforeground='white',
                           **modern_button_style)
        save_btn.pack(side=tk.LEFT, padx=20)
        
        # İptal butonu (Kırmızı)
        cancel_btn = tk.Button(button_container,
                             text="❌ İptal",
                             command=dialog.destroy,
                             bg='#e74c3c',
                             fg='white',
                             activebackground='#c0392b',
                             activeforeground='white',
                             **modern_button_style)
        cancel_btn.pack(side=tk.LEFT, padx=20)
        
        # Hover efektleri
        def on_enter(e):
            if e.widget == save_btn:
                e.widget['bg'] = '#229954'
            elif e.widget == cancel_btn:
                e.widget['bg'] = '#c0392b'
        
        def on_leave(e):
            if e.widget == save_btn:
                e.widget['bg'] = '#27ae60'
            elif e.widget == cancel_btn:
                e.widget['bg'] = '#e74c3c'
        
        save_btn.bind('<Enter>', on_enter)
        save_btn.bind('<Leave>', on_leave)
        cancel_btn.bind('<Enter>', on_enter)
        cancel_btn.bind('<Leave>', on_leave)
    
    def select_customer_dialog(self):
        """Modern müşteri seçme dialog'u"""
        dialog = tk.Toplevel(self.root)
        dialog.title("👤 Danışan Seç")
        dialog.geometry("800x1000")
        dialog.configure(bg='#f8f9fa')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(True, True)
        dialog.minsize(700, 800)
        
        # Pencereyi merkeze yerleştir
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
        y = (dialog.winfo_screenheight() // 2) - (1000 // 2)
        dialog.geometry(f"800x1000+{x}+{y}")
        
        # Modern scrollable container - minimal padding
        container_frame = tk.Frame(dialog, bg='#f8f9fa')
        container_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Canvas for scrolling
        canvas = tk.Canvas(container_frame, bg='#f8f9fa', highlightthickness=0, bd=0)
        
        # Modern custom scrollbar frame - minimal width
        scrollbar_frame = tk.Frame(container_frame, bg='#f8f9fa', width=8)
        scrollbar_bg = tk.Canvas(scrollbar_frame, bg='#f8f9fa', width=8, highlightthickness=0, bd=0)
        
        # Scrollable content frame
        scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
        
        # Mouse wheel scrolling with smooth effect
        def smooth_scroll(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                    update_modern_scrollbar()
            except tk.TclError:
                pass
        
        # Modern scrollbar visual
        scrollbar_thumb = None
        
        def update_modern_scrollbar():
            nonlocal scrollbar_thumb
            # Clear previous thumb
            scrollbar_bg.delete("thumb")
            
            # Get scroll position
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return  # No scrolling needed
                
                # Calculate thumb position and size
                scrollbar_height = scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                
                # Draw modern thumb - minimal width
                scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#c0c0c0', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        # Scrollbar interaction variables
        scrollbar_dragging = False
        
        def on_scrollbar_click(event):
            nonlocal scrollbar_dragging
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                
                scrollbar_height = scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                canvas.yview_moveto(click_position)
                update_modern_scrollbar()
                scrollbar_dragging = True
            except:
                pass
        
        def on_scrollbar_drag(event):
            if not scrollbar_dragging:
                return
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                
                scrollbar_height = scrollbar_bg.winfo_height()
                click_position = max(0, min(1, event.y / scrollbar_height))
                canvas.yview_moveto(click_position)
                update_modern_scrollbar()
            except:
                pass
        
        def on_scrollbar_release(event):
            nonlocal scrollbar_dragging
            scrollbar_dragging = False
        
        def on_scrollbar_enter(event):
            scrollbar_bg.delete("thumb")
            try:
                top, bottom = canvas.yview()
                if bottom - top >= 1.0:
                    return
                scrollbar_height = scrollbar_bg.winfo_height()
                thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                thumb_y = int(scrollbar_height * top)
                scrollbar_bg.create_rectangle(
                    1, thumb_y, 7, thumb_y + thumb_height,
                    fill='#999999', outline='', tags="thumb",
                    width=0
                )
            except:
                pass
        
        def on_scrollbar_leave(event):
            update_modern_scrollbar()
        
        # Bind events
        def safe_close_dialog():
            try:
                dialog.unbind('<MouseWheel>')
                dialog.unbind('<Return>')
                dialog.unbind('<Escape>')
            except:
                pass
            dialog.destroy()
        
        try:
            dialog.bind("<MouseWheel>", smooth_scroll)
        except:
            pass
        
        dialog.protocol("WM_DELETE_WINDOW", safe_close_dialog)
        scrollbar_bg.bind("<Button-1>", on_scrollbar_click)
        scrollbar_bg.bind("<B1-Motion>", on_scrollbar_drag)
        scrollbar_bg.bind("<ButtonRelease-1>", on_scrollbar_release)
        scrollbar_bg.bind("<Enter>", on_scrollbar_enter)
        scrollbar_bg.bind("<Leave>", on_scrollbar_leave)
        
        # Make scrollbar focusable
        scrollbar_bg.configure(takefocus=True)
        
        # Update scrollbar when canvas scrolls
        def on_canvas_configure(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.after_idle(update_modern_scrollbar)
        
        canvas.bind("<Configure>", on_canvas_configure)
        scrollable_frame.bind("<Configure>", on_canvas_configure)
        
        # Create canvas window
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Update canvas window width when canvas width changes
        def on_canvas_configure_width(event):
            canvas.itemconfig(canvas_window, width=event.width)
        
        canvas.bind("<Configure>", on_canvas_configure_width)
        
        # Pack widgets - no extra padding on scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_frame.pack(side="right", fill="y", padx=0)
        scrollbar_frame.pack_propagate(False)  # Keep fixed width
        scrollbar_bg.pack(fill="both", expand=True)
        
        # Initialize scrollbar
        canvas.after(100, update_modern_scrollbar)
        
        # Dialog kapatma event'ini yakala
        def on_close():
            try:
                canvas.unbind_all("<MouseWheel>")
            except:
                pass
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        # Ana container - Modern padding (artık scrollable_frame içinde)
        main_container = tk.Frame(scrollable_frame, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)
        
        # Header card
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 20))
        
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=40, pady=30)
        
        tk.Label(header_inner,
                text="👤 Danışan Seçin",
                font=('Segoe UI', 24, 'bold'),
                bg='white',
                fg='#2c3e50').pack()
        
        tk.Label(header_inner,
                text="Sipariş oluşturmak için bir danışan seçin",
                font=('Segoe UI', 12),
                bg='white',
                fg='#7f8c8d').pack(pady=(5, 0))
        
        # Search card
        search_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        search_card.pack(fill=tk.X, pady=(0, 20))
        
        search_inner = tk.Frame(search_card, bg='white')
        search_inner.pack(fill=tk.X, padx=40, pady=25)
        
        tk.Label(search_inner,
                text="🔍 Danışan Ara",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(0, 10))
        
        # Modern arama kutusu container
        search_input_frame = tk.Frame(search_inner, bg='white')
        search_input_frame.pack(fill=tk.X, pady=(0, 0))
        
        # Modern arama kutusu - gradient efektli
        search_outer = tk.Frame(search_input_frame, bg='#ecf0f1', relief='flat', bd=0)
        search_outer.pack(fill=tk.X, pady=8)
        
        search_box = tk.Frame(search_outer, bg='#ffffff', relief='flat', bd=0)
        search_box.pack(fill=tk.X, padx=2, pady=2)
        
        # Modern çerçeve efekti
        search_border = tk.Frame(search_box, bg='#bdc3c7', height=1)
        search_border.pack(fill=tk.X, side=tk.BOTTOM)
        
        # İçerik container
        search_content = tk.Frame(search_box, bg='#ffffff')
        search_content.pack(fill=tk.X, padx=16, pady=16)
        
        # Arama ikonu - daha büyük ve parlak
        search_icon = tk.Label(search_content, text="🔍", font=('Segoe UI', 16), 
                              bg='#ffffff', fg='#3498db')
        search_icon.pack(side=tk.LEFT, padx=(0, 12))
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_content, textvariable=search_var,
                              font=('Segoe UI', 13),
                              relief='flat',
                              bd=0,
                              bg='#ffffff',
                              fg='#2c3e50',
                              insertbackground='#3498db')
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6)
        
        # Gelişmiş focus ve hover efektleri
        def on_focus_in(event):
            search_border.configure(bg='#3498db', height=2)
            search_icon.configure(fg='#2980b9')
            search_outer.configure(bg='#e8f4f8')
        
        def on_focus_out(event):
            search_border.configure(bg='#bdc3c7', height=1)
            search_icon.configure(fg='#3498db')
            search_outer.configure(bg='#ecf0f1')
        
        def on_enter(event):
            if search_entry != event.widget.focus_get():
                search_outer.configure(bg='#e8f4f8')
        
        def on_leave(event):
            if search_entry != event.widget.focus_get():
                search_outer.configure(bg='#ecf0f1')
        
        search_entry.bind('<FocusIn>', on_focus_in)
        search_entry.bind('<FocusOut>', on_focus_out)
        search_outer.bind('<Enter>', on_enter)
        search_outer.bind('<Leave>', on_leave)
        
        # List card - Sabit yükseklik
        list_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        list_card.pack(fill=tk.X, pady=(0, 20))
        list_card.configure(height=400)
        
        list_inner = tk.Frame(list_card, bg='white')
        list_inner.pack(fill=tk.X, padx=40, pady=25)
        
        tk.Label(list_inner,
                text="📋 Danışan Listesi",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        # Treeview container - Sabit yükseklik
        tree_frame = tk.Frame(list_inner, bg='white')
        tree_frame.pack(fill=tk.X)
        tree_frame.configure(height=300)
        
        # Modern Treeview
        columns = ('ID', 'Ad', 'E-posta', 'Telefon', 'Şirket')
        customer_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', 
                                   height=12, style='Modern.Treeview')
        
        # Sütun başlıkları ve genişlikleri - sola hizalı
        customer_tree.heading('ID', text='ID', anchor='w')
        customer_tree.heading('Ad', text='Ad Soyad', anchor='w')
        customer_tree.heading('E-posta', text='E-posta', anchor='w')
        customer_tree.heading('Telefon', text='Telefon', anchor='w')
        customer_tree.heading('Şirket', text='Şirket', anchor='w')
        
        customer_tree.column('ID', width=100, anchor='w')
        customer_tree.column('Ad', width=250, anchor='w')
        customer_tree.column('E-posta', width=280, anchor='w')
        customer_tree.column('Telefon', width=160, anchor='w')
        customer_tree.column('Şirket', width=180, anchor='w')
        
        # Modern Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, 
                                command=customer_tree.yview, style='Modern.Vertical.TScrollbar')
        customer_tree.configure(yscrollcommand=scrollbar.set)
        
        customer_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Müşterileri yükle
        def load_customers():
            for item in customer_tree.get_children():
                customer_tree.delete(item)
            
            search_term = search_var.get().strip()
            if search_term:
                customers = self.db.search_customers(search_term)
            else:
                customers = self.db.get_all_customers()
            
            for i, customer in enumerate(customers):
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                customer_tree.insert('', 'end', values=(
                    customer[0], customer[1], customer[2], 
                    customer[3], customer[5] if customer[5] else "—"
                ), tags=(tag,))
        
        # İlk yükleme
        load_customers()
        
        # Arama fonksiyonu
        def on_search(event=None):
            load_customers()
        
        search_entry.bind('<KeyRelease>', on_search)
        search_entry.focus()
        
        # Seçim değişkeni
        selected_customer = None
        
        # Çift tıklama ile seçim
        def on_double_click(event):
            selection = customer_tree.selection()
            if selection:
                nonlocal selected_customer
                customer_data = customer_tree.item(selection[0])['values']
                selected_customer = (customer_data[0], customer_data[1])  # ID, Name
                dialog.destroy()
        
        customer_tree.bind('<Double-1>', on_double_click)
        
        # Button card
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X)
        
        button_inner = tk.Frame(button_card, bg='white')
        button_inner.pack(fill=tk.X, padx=40, pady=30)
        
        # Butonları ortala
        button_frame = tk.Frame(button_inner, bg='white')
        button_frame.pack(anchor='center')
        
        def select_customer():
            selection = customer_tree.selection()
            if not selection:
                messagebox.showwarning("Uyarı", "Lütfen bir danışan seçin!")
                return
            
            nonlocal selected_customer
            customer_data = customer_tree.item(selection[0])['values']
            selected_customer = (customer_data[0], customer_data[1])  # ID, Name
            dialog.destroy()
        
        # Modern butonlar - Daha büyük
        select_btn = tk.Button(button_frame, text="✅ Seç ve Devam Et", 
                             command=select_customer,
                             bg='#27ae60', fg='white',
                             font=('Segoe UI', 13, 'bold'),
                             relief='flat',
                             borderwidth=0,
                             padx=40,
                             pady=15,
                             cursor='hand2')
        select_btn.pack(side=tk.LEFT, padx=15)
        
        cancel_btn = tk.Button(button_frame, text="❌ İptal", 
                             command=dialog.destroy,
                             bg='#e74c3c', fg='white',
                             font=('Segoe UI', 13, 'bold'),
                             relief='flat',
                             borderwidth=0,
                             padx=40,
                             pady=15,
                             cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=15)
        
        # Buton hover efektleri
        def on_enter_select(e):
            select_btn.configure(bg='#229954')
        def on_leave_select(e):
            select_btn.configure(bg='#27ae60')
        def on_enter_cancel(e):
            cancel_btn.configure(bg='#c0392b')
        def on_leave_cancel(e):
            cancel_btn.configure(bg='#e74c3c')
        
        select_btn.bind('<Enter>', on_enter_select)
        select_btn.bind('<Leave>', on_leave_select)
        cancel_btn.bind('<Enter>', on_enter_cancel)
        cancel_btn.bind('<Leave>', on_leave_cancel)
        
        # Dialog'u bekle
        dialog.wait_window()
        return selected_customer
    
    def add_order_from_orders_tab(self):
        """Siparişler sekmesinden yeni sipariş ekleme"""
        # Doğrudan müşteri seçme dialog'unu aç
        customer = self.select_customer_dialog()
        if customer:
            customer_id, customer_name = customer
            self.create_order_dialog(customer_id, customer_name)
    
    def create_order_dialog(self, customer_id, customer_name):
        """Modern sipariş oluşturma dialog'u"""
        # Modern Yeni Sipariş penceresi - Daha büyük boyut
        dialog = tk.Toplevel(self.root)
        dialog.title(f"📝 Yeni Sipariş - {customer_name}")
        dialog.geometry("1050x1150")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        dialog.resizable(True, True)
        
        # Pencereyi merkeze yerleştir
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (1050 // 2)
        y = (dialog.winfo_screenheight() // 2) - (1150 // 2)
        dialog.geometry(f"1050x1150+{x}+{y}")
        
        # Ana container - Daha geniş padding
        main_container = tk.Frame(dialog, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        # Header card - Müşteri bilgisi
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 20))
        
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=40, pady=30)
        
        tk.Label(header_inner,
                text="📝 Yeni Sipariş Oluştur",
                font=('Segoe UI', 22, 'bold'),
                bg='white',
                fg='#2c3e50').pack()
        
        tk.Label(header_inner,
                text=f"👤 Müşteri: {customer_name}",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#27ae60').pack(pady=(8, 0))
        
        # Form kartı
        form_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        form_card.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        form_inner = tk.Frame(form_card, bg='white')
        form_inner.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        # Ürün bilgileri bölümü
        tk.Label(form_inner,
                text="🛍️ Ürün Bilgileri",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        # Ürün adı
        product_frame = tk.Frame(form_inner, bg='white')
        product_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(product_frame,
                text="Ürün/Hizmet Adı *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        product_entry = tk.Entry(product_frame,
                               font=('Segoe UI', 12),
                               relief='flat',
                               bd=0,
                               bg='#f8f9fa',
                               fg='#2c3e50',
                               insertbackground='#3498db')
        product_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)
        
        # Miktar ve Fiyat - Yan yana
        quantity_price_frame = tk.Frame(form_inner, bg='white')
        quantity_price_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Sol: Miktar
        left_frame = tk.Frame(quantity_price_frame, bg='white')
        left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(left_frame,
                text="Miktar *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        quantity_entry = tk.Entry(left_frame,
                                font=('Segoe UI', 12),
                                relief='flat',
                                bd=0,
                                bg='#f8f9fa',
                                fg='#2c3e50',
                                insertbackground='#3498db')
        quantity_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)
        
        # Sağ: Birim Fiyat
        right_frame = tk.Frame(quantity_price_frame, bg='white')
        right_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        
        tk.Label(right_frame,
                text="Birim Fiyat (₺) *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        price_entry = tk.Entry(right_frame,
                             font=('Segoe UI', 12),
                             relief='flat',
                             bd=0,
                             bg='#f8f9fa',
                             fg='#2c3e50',
                             insertbackground='#3498db')
        price_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)
        
        # Tarih alanları bölümü
        tk.Label(form_inner,
                text="📅 Tarih Bilgileri",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(25, 15))
        
        # Başlangıç ve Bitiş tarihi - Yan yana
        dates_frame = tk.Frame(form_inner, bg='white')
        dates_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Sol: Başlangıç tarihi
        start_date_container = tk.Frame(dates_frame, bg='white')
        start_date_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(start_date_container,
                text="Başlangıç Tarihi *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        start_date_input_frame = tk.Frame(start_date_container, bg='white')
        start_date_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        start_date_entry = tk.Entry(start_date_input_frame,
                                  font=('Segoe UI', 12),
                                  relief='flat',
                                  bd=0,
                                  bg='#f8f9fa',
                                  fg='#2c3e50',
                                  insertbackground='#3498db')
        start_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(0, 8))
        
        # Modern takvim butonu
        def open_start_date_picker():
            self.show_modern_date_picker(start_date_entry, "Başlangıç Tarihi Seç")
        
        start_date_btn = tk.Button(start_date_input_frame,
                                 text="📅",
                                 command=open_start_date_picker,
                                 font=('Segoe UI', 12),
                                 bg='#3498db',
                                 fg='white',
                                 relief='flat',
                                 borderwidth=0,
                                 padx=8,
                                 pady=8,
                                 cursor='hand2',
                                 activebackground='#2980b9',
                                 activeforeground='white')
        start_date_btn.pack(side=tk.RIGHT)
        
        # Sağ: Bitiş tarihi
        end_date_container = tk.Frame(dates_frame, bg='white')
        end_date_container.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        
        tk.Label(end_date_container,
                text="Bitiş Tarihi *",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w')
        
        end_date_input_frame = tk.Frame(end_date_container, bg='white')
        end_date_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        end_date_entry = tk.Entry(end_date_input_frame,
                                font=('Segoe UI', 12),
                                relief='flat',
                                bd=0,
                                bg='#f8f9fa',
                                fg='#2c3e50',
                                insertbackground='#3498db')
        end_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(0, 8))
        
        # Modern takvim butonu
        def open_end_date_picker():
            self.show_modern_date_picker(end_date_entry, "Bitiş Tarihi Seç")
        
        end_date_btn = tk.Button(end_date_input_frame,
                               text="📅",
                               command=open_end_date_picker,
                               font=('Segoe UI', 12),
                               bg='#e74c3c',
                               fg='white',
                               relief='flat',
                               borderwidth=0,
                               padx=8,
                               pady=8,
                               cursor='hand2',
                               activebackground='#c0392b',
                               activeforeground='white')
        end_date_btn.pack(side=tk.RIGHT)
        
        # Varsayılan tarihleri ayarla
        from datetime import datetime, timedelta
        today = datetime.now().date()
        start_date_entry.insert(0, today.strftime("%d.%m.%Y"))
        end_date_entry.insert(0, (today + timedelta(days=7)).strftime("%d.%m.%Y"))
        
        # Durum bölümü
        tk.Label(form_inner,
                text="📊 Sipariş Durumu",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg='#2c3e50').pack(anchor='w', pady=(25, 15))
        
        status_frame = tk.Frame(form_inner, bg='white')
        status_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(status_frame,
                text="Başlangıç Durumu",
                font=('Segoe UI', 11, 'bold'),
                bg='white',
                fg='#34495e').pack(anchor='w', pady=(0, 10))
        
        # Modern durum seçim butonları
        status_var = tk.StringVar(value="Beklemede")
        status_options = [
            ("⏳ Beklemede", "Beklemede", "#f39c12"),
            ("🔧 Hazırlanıyor", "Hazırlanıyor", "#3498db"),
            ("✅ Tamamlandı", "Tamamlandı", "#27ae60"),
            ("❌ İptal", "İptal", "#e74c3c")
        ]
        
        status_buttons_frame = tk.Frame(status_frame, bg='white')
        status_buttons_frame.pack(fill=tk.X, pady=(5, 0))
        
        status_buttons = []
        
        def select_status(status_value):
            status_var.set(status_value)
            # Buton renklerini güncelle
            for btn, (_, btn_status, btn_color) in zip(status_buttons, status_options):
                if btn_status == status_value:
                    btn.configure(bg=btn_color, fg='white', relief='solid', bd=2)
                else:
                    btn.configure(bg='#ecf0f1', fg='#34495e', relief='flat', bd=0)
        
        # Durum butonlarını oluştur
        for i, (display_text, status_value, color) in enumerate(status_options):
            btn = tk.Button(status_buttons_frame,
                          text=display_text,
                          command=lambda s=status_value: select_status(s),
                          font=('Segoe UI', 11, 'bold'),
                          bg='#ecf0f1' if status_value != "Beklemede" else color,
                          fg='#34495e' if status_value != "Beklemede" else 'white',
                          relief='flat' if status_value != "Beklemede" else 'solid',
                          bd=0 if status_value != "Beklemede" else 2,
                          padx=15,
                          pady=8,
                          cursor='hand2',
                          activebackground='#bdc3c7',
                          activeforeground='#2c3e50')
            
            # 2x2 grid düzeni
            row = i // 2
            col = i % 2
            btn.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
            status_buttons.append(btn)
            
            # Hover efektleri
            def on_enter(e, original_color=color):
                if e.widget['bg'] != '#2c3e50':  # Seçili değilse
                    e.widget.configure(bg=original_color, fg='white')
            
            def on_leave(e):
                current_status = status_var.get()
                btn_index = status_buttons.index(e.widget)
                btn_status = status_options[btn_index][1]
                if btn_status != current_status:
                    e.widget.configure(bg='#ecf0f1', fg='#34495e')
                else:
                    # Seçili buton kendi renginde kalır
                    btn_color = status_options[btn_index][2]
                    e.widget.configure(bg=btn_color, fg='white')
            
            btn.bind('<Enter>', on_enter)
            btn.bind('<Leave>', on_leave)
        
        # Grid ağırlıklarını ayarla
        status_buttons_frame.columnconfigure(0, weight=1)
        status_buttons_frame.columnconfigure(1, weight=1)
        
        def save_order():
            product = product_entry.get().strip()
            if not product:
                messagebox.showerror("Hata", "Ürün adı zorunludur!")
                return
            
            try:
                quantity = int(quantity_entry.get())
                price = float(price_entry.get())
            except ValueError:
                messagebox.showerror("Hata", "Miktar ve fiyat sayısal değer olmalıdır!")
                return
            
            # Tarih doğrulama
            start_date = start_date_entry.get().strip()
            end_date = end_date_entry.get().strip()
            
            if not start_date or not end_date:
                messagebox.showerror("Hata", "Başlangıç ve bitiş tarihi zorunludur!")
                return
            
            # Tarihleri veritabanı formatına çevir
            start_date_db = self.format_date_for_database(start_date)
            end_date_db = self.format_date_for_database(end_date)
            
            if not start_date_db or not end_date_db:
                messagebox.showerror("Hata", "Tarih formatı GG.AA.YYYY olmalıdır! (Örnek: 27.07.2025)")
                return
            
            # Bitiş tarihi başlangıç tarihinden sonra olmalı
            if start_date_db >= end_date_db:
                messagebox.showerror("Hata", "Bitiş tarihi başlangıç tarihinden sonra olmalıdır!")
                return
            
            status = status_var.get()
            
            try:
                self.db.add_order(customer_id, product, quantity, price, start_date_db, end_date_db, status)
                messagebox.showinfo("Başarılı", "Sipariş başarıyla eklendi!")
                
                # Otomatik kayıt
                self.auto_save_data("Sipariş Eklendi")
                
                dialog.destroy()
                self.clear_cache()  # Cache'i temizle
                self.load_orders()
                self.load_stats()
            except Exception as e:
                messagebox.showerror("Hata", f"Sipariş eklenirken hata oluştu: {str(e)}")
        
        # Modern buton kartı
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X, pady=(0, 0))
        
        button_container = tk.Frame(button_card, bg='white')
        button_container.pack(pady=40)
        
        # Modern buton stilleri - büyütülmüş
        modern_button_style = {
            'font': ('Segoe UI', 16, 'bold'),
            'relief': 'flat',
            'borderwidth': 0,
            'padx': 45,
            'pady': 16,
            'cursor': 'hand2'
        }
        
        # Kaydet butonu (Yeşil)
        save_btn = tk.Button(button_container,
                           text="💾 Siparişi Kaydet",
                           command=save_order,
                           bg='#27ae60',
                           fg='white',
                           activebackground='#229954',
                           activeforeground='white',
                           **modern_button_style)
        save_btn.pack(side=tk.LEFT, padx=20)
        
        # İptal butonu (Kırmızı)
        cancel_btn = tk.Button(button_container,
                             text="❌ İptal",
                             command=dialog.destroy,
                             bg='#e74c3c',
                             fg='white',
                             activebackground='#c0392b',
                             activeforeground='white',
                             **modern_button_style)
        cancel_btn.pack(side=tk.LEFT, padx=20)
        
        # Hover efektleri
        def on_enter(e):
            if e.widget == save_btn:
                e.widget['bg'] = '#229954'
            elif e.widget == cancel_btn:
                e.widget['bg'] = '#c0392b'
        
        def on_leave(e):
            if e.widget == save_btn:
                e.widget['bg'] = '#27ae60'
            elif e.widget == cancel_btn:
                e.widget['bg'] = '#e74c3c'
        
        save_btn.bind('<Enter>', on_enter)
        save_btn.bind('<Leave>', on_leave)
        cancel_btn.bind('<Enter>', on_enter)
        cancel_btn.bind('<Leave>', on_leave)
    
    def load_orders(self):
        """Sipariş listesini yükler (cache ile optimize edilmiş)"""
        # Cache kontrolü (5 dakika)
        cache_age = (datetime.now() - self.cache_timestamp).total_seconds()
        if cache_age < 300 and self.order_cache:  # 5 dakika cache
            orders = self.order_cache
        else:
            try:
                orders = self.db.get_all_orders()
                self.order_cache = orders
                self.cache_timestamp = datetime.now()
            except Exception as e:
                messagebox.showerror("Hata", f"Siparişler yüklenirken hata oluştu: {str(e)}")
                return
        
        # Mevcut listeyi temizle
        for item in self.order_tree.get_children():
            self.order_tree.delete(item)
        
        # Siparişleri listele
        for i, order in enumerate(orders):
            # None değerleri kontrol et
            price = order[7] if order[7] is not None else 0.0
            total_price = order[8] if order[8] is not None else 0.0
            
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            self.order_tree.insert('', 'end', values=(
                order[11] if len(order) > 11 else "Bilinmiyor",  # Müşteri adı (12. sütun, indeks 11)
                order[5],  # Ürün adı
                order[6],  # Miktar
                f"{price:.2f}",  # Birim fiyat
                f"{total_price:.2f}",  # Toplam fiyat
                self.format_date_for_display(order[3]),  # Başlangıç tarihi
                self.format_date_for_display(order[4]),  # Bitiş tarihi
                order[9],  # Durum
                order[2]   # Sipariş tarihi
            ), tags=(tag,))
    
    def show_order_details(self):
        """Sipariş detaylarını gösterir"""
        selection = self.order_tree.selection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen detayları görüntülenecek siparişi seçin!")
            return
        
        # ID sütunu artık görünmediği için, sipariş ID'sini veritabanından almak gerekiyor
        # Seçili satırın müşteri adı ve ürün adını kullanarak ID'yi bulacağız
        selected_values = self.order_tree.item(selection[0])['values']
        customer_name = selected_values[0]  # Müşteri adı
        product_name = selected_values[1]   # Ürün adı
        
        # Veritabanından bu siparişin ID'sini bul
        conn = sqlite3.connect(self.db.db_name)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT o.id FROM orders o
            JOIN customers c ON o.customer_id = c.id
            WHERE c.name = ? AND o.product_name = ?
        ''', (customer_name, product_name))
        result = cursor.fetchone()
        conn.close()
        
        if not result:
            messagebox.showerror("Hata", "Sipariş bulunamadı!")
            return
            
        order_id = result[0]
        
        try:
            # Sipariş detaylarını getir
            conn = sqlite3.connect(self.db.db_name)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT o.*, c.name as customer_name
                FROM orders o
                JOIN customers c ON o.customer_id = c.id
                WHERE o.id = ?
            ''', (order_id,))
            
            order = cursor.fetchone()
            conn.close()
            
            if not order:
                messagebox.showerror("Hata", "Sipariş bulunamadı!")
                return
            
            # Modern Sipariş Detayları penceresi
            dialog = tk.Toplevel(self.root)
            dialog.title(f"👁️ Sipariş Detayları - #{order_id}")
            dialog.geometry("800x1000")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.configure(bg='#f8f9fa')
            dialog.resizable(True, True)
            
            # Pencereyi merkeze yerleştir
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
            y = (dialog.winfo_screenheight() // 2) - (1000 // 2)
            dialog.geometry(f"800x1000+{x}+{y}")
            
            # Canvas ve scrollbar için ana frame
            main_frame = tk.Frame(dialog, bg='#f8f9fa')
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Canvas oluştur
            canvas = tk.Canvas(main_frame, bg='#f8f9fa', highlightthickness=0)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Custom scrollbar
            scrollbar_frame = tk.Frame(main_frame, bg='#e5e5e5', width=8)
            scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y)
            scrollbar_frame.pack_propagate(False)
            
            scrollbar = tk.Frame(scrollbar_frame, bg='#c0c0c0', width=8)
            scrollbar.place(relx=0, rely=0, relwidth=1, relheight=0.3)
            
            # Scrollable frame
            scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
            canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            
            # Ana container
            main_container = tk.Frame(scrollable_frame, bg='#f8f9fa')
            main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Başlık kartı
            header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            header_card.pack(fill=tk.X, pady=(0, 20))
            
            header_inner = tk.Frame(header_card, bg='white')
            header_inner.pack(fill=tk.X, padx=30, pady=25)
            
            tk.Label(header_inner,
                    text=f"👁️ Sipariş Detayları - #{order_id}",
                    font=('Segoe UI', 20, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack()
            
            # JOIN sonucu customer_name en son element
            customer_name = order[-1] if order else 'Bilinmiyor'
            tk.Label(header_inner,
                    text=f"👤 Müşteri: {customer_name}",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#3498db').pack(pady=(10, 0))
            
            # Detay kartı
            detail_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            detail_card.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            
            # Detay içeriği
            detail_inner = tk.Frame(detail_card, bg='white')
            detail_inner.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Ürün Bilgileri Bölümü
            tk.Label(detail_inner,
                    text="📦 Ürün Bilgileri",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(0, 15))
            
            product_frame = tk.Frame(detail_inner, bg='white')
            product_frame.pack(fill=tk.X, pady=(0, 20))
            
            # Ürün bilgileri
            product_info = [
                ("Ürün Adı:", order[5] if len(order) > 5 else "Bilinmiyor"),
                ("Miktar:", str(order[6]) if len(order) > 6 else "0"),
                ("Birim Fiyat:", f"{order[7]:.2f} ₺" if len(order) > 7 and order[7] is not None else "0.00 ₺"),
                ("Toplam Fiyat:", f"{order[8]:.2f} ₺" if len(order) > 8 and order[8] is not None else "0.00 ₺")
            ]
            
            for i, (label, value) in enumerate(product_info):
                info_frame = tk.Frame(product_frame, bg='white')
                info_frame.pack(fill=tk.X, pady=8)
                
                tk.Label(info_frame, text=label, font=('Segoe UI', 12, 'bold'),
                        bg='white', fg='#34495e').pack(side=tk.LEFT)
                tk.Label(info_frame, text=value, font=('Segoe UI', 12),
                        bg='white', fg='#2c3e50').pack(side=tk.LEFT, padx=(10, 0))
            
            # Tarih Bilgileri Bölümü
            tk.Label(detail_inner,
                    text="📅 Tarih Bilgileri",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(20, 15))
            
            date_frame = tk.Frame(detail_inner, bg='white')
            date_frame.pack(fill=tk.X, pady=(0, 20))
            
            # Tarih bilgileri
            date_info = [
                ("Sipariş Tarihi:", order[2] if len(order) > 2 else "Bilinmiyor"),
                ("Başlangıç Tarihi:", order[3] if len(order) > 3 else "Bilinmiyor"),
                ("Bitiş Tarihi:", order[4] if len(order) > 4 else "Bilinmiyor")
            ]
            
            for i, (label, value) in enumerate(date_info):
                info_frame = tk.Frame(date_frame, bg='white')
                info_frame.pack(fill=tk.X, pady=8)
                
                tk.Label(info_frame, text=label, font=('Segoe UI', 12, 'bold'),
                        bg='white', fg='#34495e').pack(side=tk.LEFT)
                tk.Label(info_frame, text=value, font=('Segoe UI', 12),
                        bg='white', fg='#2c3e50').pack(side=tk.LEFT, padx=(10, 0))
            
            # Durum Bilgisi
            tk.Label(detail_inner,
                    text="📊 Sipariş Durumu",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(20, 15))
            
            status_frame = tk.Frame(detail_inner, bg='white')
            status_frame.pack(fill=tk.X, pady=(0, 20))
            
            current_status = order[9] if len(order) > 9 else "Bilinmiyor"
            
            # Durum rengini belirle
            status_colors = {
                "Beklemede": "#f39c12",
                "Hazırlanıyor": "#3498db", 
                "Devam Ediyor": "#3498db",
                "Tamamlandı": "#27ae60",
                "İptal": "#e74c3c"
            }
            status_color = status_colors.get(current_status, "#95a5a6")
            
            status_label = tk.Label(status_frame,
                                  text=f"📍 {current_status}",
                                  font=('Segoe UI', 14, 'bold'),
                                  bg=status_color,
                                  fg='white',
                                  padx=20,
                                  pady=10,
                                  relief='flat')
            status_label.pack(anchor='w')
            
            # Kapat butonu
            button_frame = tk.Frame(detail_inner, bg='white')
            button_frame.pack(pady=(20, 20))
            
            def on_dialog_close():
                try:
                    dialog.unbind('<MouseWheel>')
                except:
                    pass
                dialog.destroy()
            
            close_btn = tk.Button(button_frame,
                                text="❌ Kapat",
                                command=on_dialog_close,
                                font=('Segoe UI', 14, 'bold'),
                                bg='#e74c3c',
                                fg='white',
                                relief='flat',
                                borderwidth=0,
                                padx=30,
                                pady=12,
                                cursor='hand2',
                                activebackground='#c0392b',
                                activeforeground='white')
            close_btn.pack()
            
            # Hover efektleri
            def on_close_hover_enter(e):
                close_btn.configure(bg='#c0392b')
            
            def on_close_hover_leave(e):
                close_btn.configure(bg='#e74c3c')
            
            close_btn.bind('<Enter>', on_close_hover_enter)
            close_btn.bind('<Leave>', on_close_hover_leave)
            
            # Scroll sistem fonksiyonları
            def update_scrollbar():
                canvas.update_idletasks()
                scroll_top = canvas.canvasy(0)
                scroll_bottom = canvas.canvasy(canvas.winfo_height())
                total_height = scrollable_frame.winfo_reqheight()
                
                if total_height <= canvas.winfo_height():
                    scrollbar.place_forget()
                    return
                
                scrollbar_height = max(0.1, canvas.winfo_height() / total_height)
                scrollbar_top = scroll_top / total_height
                
                scrollbar.place(relx=0, rely=scrollbar_top, relwidth=1, relheight=scrollbar_height)
            
            def on_canvas_configure(event):
                canvas.configure(scrollregion=canvas.bbox("all"))
                update_scrollbar()
                
                # Canvas genişliğini pencereye uyarla
                canvas_width = event.width
                canvas.itemconfig(canvas_window, width=canvas_width)
            
            def on_mousewheel(event):
                try:
                    if canvas.winfo_exists():
                        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                        update_scrollbar()
                except tk.TclError:
                    pass
            
            def on_frame_configure(event):
                canvas.configure(scrollregion=canvas.bbox("all"))
                update_scrollbar()
            
            # Event bindings
            scrollable_frame.bind('<Configure>', on_frame_configure)
            canvas.bind('<Configure>', on_canvas_configure)
            
            try:
                dialog.bind('<MouseWheel>', on_mousewheel)
            except:
                pass
            
            # Dialog kapatma protokolü
            dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
            
            # İlk güncelleme
            dialog.after(100, update_scrollbar)
            
        except Exception as e:
            messagebox.showerror("Hata", f"Sipariş detayları yüklenirken hata oluştu: {str(e)}")
    
    def edit_order(self):
        """Modern siparişi düzenleme penceresi"""
        selection = self.order_tree.selection()
        if not selection:
            messagebox.showwarning("Uyarı", "Lütfen düzenlenecek siparişi seçin!")
            return
        
        # ID sütunu artık görünmediği için, sipariş ID'sini veritabanından almak gerekiyor
        # Seçili satırın müşteri adı ve ürün adını kullanarak ID'yi bulacağız
        selected_values = self.order_tree.item(selection[0])['values']
        customer_name = selected_values[0]  # Müşteri adı
        product_name = selected_values[1]   # Ürün adı
        
        # Veritabanından bu siparişin ID'sini bul
        conn = sqlite3.connect(self.db.db_name)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT o.id FROM orders o
            JOIN customers c ON o.customer_id = c.id
            WHERE c.name = ? AND o.product_name = ?
        ''', (customer_name, product_name))
        result = cursor.fetchone()
        conn.close()
        
        if not result:
            messagebox.showerror("Hata", "Sipariş bulunamadı!")
            return
            
        order_id = result[0]
        
        try:
            # Sipariş bilgilerini getir
            order = self.db.get_order_by_id(order_id)
            
            if not order:
                messagebox.showerror("Hata", "Sipariş bulunamadı!")
                return
            
            # Modern Sipariş Düzenleme penceresi
            dialog = tk.Toplevel(self.root)
            dialog.title(f"✏️ Sipariş Düzenle - #{order_id}")
            dialog.geometry("800x1000")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.configure(bg='#f8f9fa')
            dialog.resizable(True, True)
            
            # Pencereyi merkeze yerleştir
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (800 // 2)
            y = (dialog.winfo_screenheight() // 2) - (1000 // 2)
            dialog.geometry(f"800x1000+{x}+{y}")
            
            # Canvas ve scrollbar için ana frame
            main_frame = tk.Frame(dialog, bg='#f8f9fa')
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Canvas oluştur
            canvas = tk.Canvas(main_frame, bg='#f8f9fa', highlightthickness=0)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Custom scrollbar
            scrollbar_frame = tk.Frame(main_frame, bg='#e5e5e5', width=8)
            scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y)
            scrollbar_frame.pack_propagate(False)
            
            scrollbar = tk.Frame(scrollbar_frame, bg='#c0c0c0', width=8)
            scrollbar.place(relx=0, rely=0, relwidth=1, relheight=0.3)
            
            # Scrollable frame
            scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
            canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            
            # Ana container
            main_container = tk.Frame(scrollable_frame, bg='#f8f9fa')
            main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Başlık kartı
            header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            header_card.pack(fill=tk.X, pady=(0, 20))
            
            header_inner = tk.Frame(header_card, bg='white')
            header_inner.pack(fill=tk.X, padx=30, pady=25)
            
            tk.Label(header_inner,
                    text=f"✏️ Sipariş Düzenle - #{order_id}",
                    font=('Segoe UI', 20, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack()
            
            # JOIN sonucu customer_name en son element
            customer_name = order[-1] if order else 'Bilinmiyor'
            tk.Label(header_inner,
                    text=f"👤 Müşteri: {customer_name}",
                    font=('Segoe UI', 14, 'bold'),
                    bg='white',
                    fg='#e74c3c').pack(pady=(10, 0))
            
            # Form kartı
            form_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            form_card.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            
            # Form içeriği
            form_inner = tk.Frame(form_card, bg='white')
            form_inner.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Ürün bilgileri bölümü
            tk.Label(form_inner,
                    text="📦 Ürün Bilgileri",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(0, 15))
            
            # Ürün adı
            tk.Label(form_inner,
                    text="Ürün Adı *",
                    font=('Segoe UI', 11, 'bold'),
                    bg='white',
                    fg='#34495e').pack(anchor='w')
            
            product_entry = tk.Entry(form_inner,
                                   font=('Segoe UI', 14),
                                   relief='flat',
                                   bd=0,
                                   bg='#f8f9fa',
                                   fg='#2c3e50',
                                   insertbackground='#3498db')
            product_entry.pack(fill=tk.X, pady=(5, 0), ipady=12)
            product_entry.insert(0, order[5] if len(order) > 5 else "")
            
            # Miktar ve Fiyat - Yan yana
            quantity_price_frame = tk.Frame(form_inner, bg='white')
            quantity_price_frame.pack(fill=tk.X, pady=(0, 15))
            
            # Sol: Miktar
            left_frame = tk.Frame(quantity_price_frame, bg='white')
            left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            
            tk.Label(left_frame,
                    text="Miktar *",
                    font=('Segoe UI', 11, 'bold'),
                    bg='white',
                    fg='#34495e').pack(anchor='w')
            
            quantity_entry = tk.Entry(left_frame,
                                    font=('Segoe UI', 14),
                                    relief='flat',
                                    bd=0,
                                    bg='#f8f9fa',
                                    fg='#2c3e50',
                                    insertbackground='#3498db')
            quantity_entry.pack(fill=tk.X, pady=(5, 0), ipady=12)
            quantity_entry.insert(0, str(order[6]) if len(order) > 6 else "1")
            
            # Sağ: Birim Fiyat
            right_frame = tk.Frame(quantity_price_frame, bg='white')
            right_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
            
            tk.Label(right_frame,
                    text="Birim Fiyat (₺) *",
                    font=('Segoe UI', 11, 'bold'),
                    bg='white',
                    fg='#34495e').pack(anchor='w')
            
            price_entry = tk.Entry(right_frame,
                                 font=('Segoe UI', 14),
                                 relief='flat',
                                 bd=0,
                                 bg='#f8f9fa',
                                 fg='#2c3e50',
                                 insertbackground='#3498db')
            price_entry.pack(fill=tk.X, pady=(5, 0), ipady=12)
            price_entry.insert(0, str(order[7]) if len(order) > 7 and order[7] else "0")
            
            # Tarih alanları bölümü
            tk.Label(form_inner,
                    text="📅 Tarih Bilgileri",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(25, 15))
            
            # Başlangıç ve Bitiş tarihi - Yan yana
            dates_frame = tk.Frame(form_inner, bg='white')
            dates_frame.pack(fill=tk.X, pady=(0, 15))
            
            # Sol: Başlangıç tarihi
            start_date_container = tk.Frame(dates_frame, bg='white')
            start_date_container.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
            
            tk.Label(start_date_container,
                    text="Başlangıç Tarihi *",
                    font=('Segoe UI', 11, 'bold'),
                    bg='white',
                    fg='#34495e').pack(anchor='w')
            
            start_date_input_frame = tk.Frame(start_date_container, bg='white')
            start_date_input_frame.pack(fill=tk.X, pady=(5, 0))
            
            start_date_entry = tk.Entry(start_date_input_frame,
                                      font=('Segoe UI', 14),
                                      relief='flat',
                                      bd=0,
                                      bg='#f8f9fa',
                                      fg='#2c3e50',
                                      insertbackground='#3498db')
            start_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=12, padx=(0, 8))
            
            # Modern takvim butonu
            def open_start_date_picker():
                self.show_modern_date_picker(start_date_entry, "Başlangıç Tarihi Seç")
            
            start_date_btn = tk.Button(start_date_input_frame,
                                     text="📅",
                                     command=open_start_date_picker,
                                     font=('Segoe UI', 14),
                                     bg='#3498db',
                                     fg='white',
                                     relief='flat',
                                     borderwidth=0,
                                     padx=12,
                                     pady=12,
                                     cursor='hand2',
                                     activebackground='#2980b9',
                                     activeforeground='white')
            start_date_btn.pack(side=tk.RIGHT)
            
            # Sağ: Bitiş tarihi
            end_date_container = tk.Frame(dates_frame, bg='white')
            end_date_container.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
            
            tk.Label(end_date_container,
                    text="Bitiş Tarihi *",
                    font=('Segoe UI', 11, 'bold'),
                    bg='white',
                    fg='#34495e').pack(anchor='w')
            
            end_date_input_frame = tk.Frame(end_date_container, bg='white')
            end_date_input_frame.pack(fill=tk.X, pady=(5, 0))
            
            end_date_entry = tk.Entry(end_date_input_frame,
                                    font=('Segoe UI', 14),
                                    relief='flat',
                                    bd=0,
                                    bg='#f8f9fa',
                                    fg='#2c3e50',
                                    insertbackground='#3498db')
            end_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=12, padx=(0, 8))
            
            # Modern takvim butonu
            def open_end_date_picker():
                self.show_modern_date_picker(end_date_entry, "Bitiş Tarihi Seç")
            
            end_date_btn = tk.Button(end_date_input_frame,
                                   text="📅",
                                   command=open_end_date_picker,
                                   font=('Segoe UI', 14),
                                   bg='#e74c3c',
                                   fg='white',
                                   relief='flat',
                                   borderwidth=0,
                                   padx=12,
                                   pady=12,
                                   cursor='hand2',
                                   activebackground='#c0392b',
                                   activeforeground='white')
            end_date_btn.pack(side=tk.RIGHT)
            
            # Tarih verilerini ayarla
            if len(order) > 3 and order[3]:
                try:
                    start_date_obj = datetime.strptime(order[3], "%Y-%m-%d")
                    start_date_entry.insert(0, start_date_obj.strftime("%d.%m.%Y"))
                except:
                    start_date_entry.insert(0, self.format_date_for_display(order[3]))
            else:
                start_date_entry.insert(0, datetime.now().strftime("%d.%m.%Y"))
            
            if len(order) > 4 and order[4]:
                try:
                    end_date_obj = datetime.strptime(order[4], "%Y-%m-%d")
                    end_date_entry.insert(0, end_date_obj.strftime("%d.%m.%Y"))
                except:
                    end_date_entry.insert(0, self.format_date_for_display(order[4]))
            else:
                end_date_entry.insert(0, (datetime.now() + timedelta(days=7)).strftime("%d.%m.%Y"))
            
            # Durum bölümü
            tk.Label(form_inner,
                    text="📊 Sipariş Durumu",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(25, 15))
            
            status_frame = tk.Frame(form_inner, bg='white')
            status_frame.pack(fill=tk.X, pady=(0, 15))
            
            tk.Label(status_frame,
                    text="Mevcut Durum",
                    font=('Segoe UI', 11, 'bold'),
                    bg='white',
                    fg='#34495e').pack(anchor='w', pady=(0, 10))
            
            # Modern durum seçim butonları
            status_var = tk.StringVar(value=order[9] if len(order) > 9 else "Beklemede")
            status_options = [
                ("⏳ Beklemede", "Beklemede", "#f39c12"),
                ("🔧 Hazırlanıyor", "Hazırlanıyor", "#3498db"),
                ("✅ Tamamlandı", "Tamamlandı", "#27ae60"),
                ("❌ İptal", "İptal", "#e74c3c")
            ]
            
            status_buttons_frame = tk.Frame(status_frame, bg='white')
            status_buttons_frame.pack(fill=tk.X, pady=(5, 0))
            
            status_buttons = []
            
            def select_status(status_value):
                status_var.set(status_value)
                # Buton renklerini güncelle
                for btn, (_, btn_status, btn_color) in zip(status_buttons, status_options):
                    if btn_status == status_value:
                        btn.configure(bg=btn_color, fg='white', relief='solid', bd=2)
                    else:
                        btn.configure(bg='#ecf0f1', fg='#34495e', relief='flat', bd=0)
            
            # Durum butonlarını oluştur
            for i, (display_text, status_value, color) in enumerate(status_options):
                current_status = order[9] if len(order) > 9 else "Beklemede"
                btn = tk.Button(status_buttons_frame,
                              text=display_text,
                              command=lambda s=status_value: select_status(s),
                              font=('Segoe UI', 13, 'bold'),
                              bg='#ecf0f1' if status_value != current_status else color,
                              fg='#34495e' if status_value != current_status else 'white',
                              relief='flat' if status_value != current_status else 'solid',
                              bd=0 if status_value != current_status else 2,
                              padx=20,
                              pady=12,
                              cursor='hand2',
                              activebackground='#bdc3c7',
                              activeforeground='#2c3e50')
                
                # 2x2 grid düzeni
                row = i // 2
                col = i % 2
                btn.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                status_buttons.append(btn)
                
                # Hover efektleri
                def on_enter(e, original_color=color):
                    if e.widget['bg'] != original_color:  # Seçili değilse
                        e.widget.configure(bg=original_color, fg='white')
                
                def on_leave(e):
                    current_status = status_var.get()
                    btn_index = status_buttons.index(e.widget)
                    btn_status = status_options[btn_index][1]
                    if btn_status != current_status:
                        e.widget.configure(bg='#ecf0f1', fg='#34495e')
                    else:
                        # Seçili buton kendi renginde kalır
                        btn_color = status_options[btn_index][2]
                        e.widget.configure(bg=btn_color, fg='white')
                
                btn.bind('<Enter>', on_enter)
                btn.bind('<Leave>', on_leave)
            
            # Grid ağırlıklarını ayarla
            status_buttons_frame.columnconfigure(0, weight=1)
            status_buttons_frame.columnconfigure(1, weight=1)
            
            # Butonlar bölümü
            tk.Label(form_inner,
                    text="⚡ İşlemler",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w', pady=(20, 10))
            
            button_container = tk.Frame(form_inner, bg='white')
            button_container.pack(pady=(0, 20))
            
            def save_changes():
                try:
                    # Verileri al
                    product_name = product_entry.get().strip()
                    quantity = int(quantity_entry.get().strip())
                    price = float(price_entry.get().strip())
                    start_date_db = self.format_date_for_database(start_date_entry.get().strip())
                    end_date_db = self.format_date_for_database(end_date_entry.get().strip())
                    status = status_var.get()
                    
                    # Validasyon
                    if not product_name:
                        messagebox.showerror("Hata", "Ürün adı boş olamaz!")
                        return
                    
                    if quantity <= 0:
                        messagebox.showerror("Hata", "Miktar 0'dan büyük olmalıdır!")
                        return
                    
                    if price < 0:
                        messagebox.showerror("Hata", "Fiyat negatif olamaz!")
                        return
                    
                    if not start_date_db or not end_date_db:
                        messagebox.showerror("Hata", "Tarihler geçerli olmalıdır!")
                        return
                    
                    # Siparişi güncelle
                    success = self.db.update_order(order_id, product_name, quantity, price, start_date_db, end_date_db, status)
                    
                    if success:
                        messagebox.showinfo("Başarılı", "Sipariş başarıyla güncellendi!")
                        
                        # Otomatik kayıt
                        self.auto_save_data("Sipariş Güncellendi")
                        
                        # Cache'i temizle ve listeyi yenile
                        self.clear_cache()
                        dialog.destroy()
                        self.load_orders()  # Listeyi yenile
                    else:
                        messagebox.showerror("Hata", "Sipariş güncellenirken hata oluştu!")
                        
                except ValueError:
                    messagebox.showerror("Hata", "Lütfen sayısal değerleri doğru girin!")
                except Exception as e:
                    messagebox.showerror("Hata", f"Sipariş güncellenirken hata oluştu: {str(e)}")
            
            # Kaydet butonu
            save_btn = tk.Button(button_container,
                               text="💾 Değişiklikleri Kaydet",
                               command=save_changes,
                               font=('Segoe UI', 16, 'bold'),
                               bg='#27ae60',
                               fg='white',
                               relief='flat',
                               borderwidth=0,
                               padx=35,
                               pady=15,
                               cursor='hand2',
                               activebackground='#229954',
                               activeforeground='white')
            save_btn.pack(side=tk.LEFT, padx=(0, 20))
            
            # İptal butonu
            cancel_btn = tk.Button(button_container,
                                 text="❌ İptal",
                                 command=dialog.destroy,
                                 font=('Segoe UI', 16, 'bold'),
                                 bg='#e74c3c',
                                 fg='white',
                                 relief='flat',
                                 borderwidth=0,
                                 padx=35,
                                 pady=15,
                                 cursor='hand2',
                                 activebackground='#c0392b',
                                 activeforeground='white')
            cancel_btn.pack(side=tk.LEFT)
            
            # Hover efektleri
            def on_save_hover_enter(e):
                save_btn.configure(bg='#229954')
            
            def on_save_hover_leave(e):
                save_btn.configure(bg='#27ae60')
            
            def on_cancel_hover_enter(e):
                cancel_btn.configure(bg='#c0392b')
            
            def on_cancel_hover_leave(e):
                cancel_btn.configure(bg='#e74c3c')
            
            save_btn.bind('<Enter>', on_save_hover_enter)
            save_btn.bind('<Leave>', on_save_hover_leave)
            cancel_btn.bind('<Enter>', on_cancel_hover_enter)
            cancel_btn.bind('<Leave>', on_cancel_hover_leave)
            
            # Scroll sistem fonksiyonları
            def update_scrollbar():
                canvas.update_idletasks()
                scroll_top = canvas.canvasy(0)
                scroll_bottom = canvas.canvasy(canvas.winfo_height())
                total_height = scrollable_frame.winfo_reqheight()
                
                if total_height <= canvas.winfo_height():
                    scrollbar.place_forget()
                    return
                
                scrollbar_height = max(0.1, canvas.winfo_height() / total_height)
                scrollbar_top = scroll_top / total_height
                
                scrollbar.place(relx=0, rely=scrollbar_top, relwidth=1, relheight=scrollbar_height)
            
            def on_canvas_configure(event):
                canvas.configure(scrollregion=canvas.bbox("all"))
                update_scrollbar()
                
                # Canvas genişliğini pencereye uyarla
                canvas_width = event.width
                canvas.itemconfig(canvas_window, width=canvas_width)
            
            def on_mousewheel(event):
                try:
                    if canvas.winfo_exists():
                        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                        update_scrollbar()
                except tk.TclError:
                    pass
            
            def on_frame_configure(event):
                canvas.configure(scrollregion=canvas.bbox("all"))
                update_scrollbar()
            
            # Event bindings
            scrollable_frame.bind('<Configure>', on_frame_configure)
            canvas.bind('<Configure>', on_canvas_configure)
            
            try:
                dialog.bind('<MouseWheel>', on_mousewheel)
            except:
                pass
            
            # Dialog kapatma eventi
            def on_dialog_close():
                try:
                    dialog.unbind('<MouseWheel>')
                except:
                    pass
                dialog.destroy()
            
            dialog.protocol("WM_DELETE_WINDOW", on_dialog_close)
            
            # İlk güncelleme
            dialog.after(100, update_scrollbar)
            
        except Exception as e:
            messagebox.showerror("Hata", f"Sipariş düzenleme penceresi açılırken hata oluştu: {str(e)}")
    
    def load_stats(self):
        """Gelir ve Danışanları yükler"""
        try:
            stats = self.db.get_order_statistics()
            
            self.total_customers_label.config(text=f"{stats['total_customers']}")
            self.total_orders_label.config(text=f"{stats['total_orders']}")
            formatted_revenue = f"{int(stats['total_revenue']):,}".replace(',', '.')
            self.total_revenue_label.config(text=f"{formatted_revenue} TL")
            formatted_cancelled = f"{int(stats['cancelled_revenue']):,}".replace(',', '.')
            self.cancelled_revenue_label.config(text=f"{formatted_cancelled} TL")
        except Exception as e:
            messagebox.showerror("Hata", f"Gelir ve Danışanlar yüklenirken hata oluştu: {str(e)}")
    

    
    # Otomatik bildirim fonksiyonları kaldırıldı
    
    def load_notifications(self):
        """Bildirimleri yükler"""
        # Yaklaşan bitiş tarihleri
        for item in self.expiring_tree.get_children():
            self.expiring_tree.delete(item)
        
        try:
            expiring_orders = self.db.get_expiring_orders()
            for i, order in enumerate(expiring_orders):
                notification_status = "Gönderildi" if order[9] else "Bekliyor"  # notification_sent
                formatted_date = self.format_date_for_display(order[7])  # end_date
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                self.expiring_tree.insert('', 'end', values=(
                    order[0],  # Sipariş ID
                    order[1],  # Müşteri adı
                    order[2],  # Ürün adı
                    formatted_date,  # Bitiş tarihi (formatlanmış)
                    order[8],  # Durum
                    notification_status
                ), tags=(tag,))
        except Exception as e:
            messagebox.showerror("Hata", f"Yaklaşan siparişler yüklenirken hata oluştu: {str(e)}")
        
        # Gecikmiş siparişler
        for item in self.overdue_tree.get_children():
            self.overdue_tree.delete(item)
        
        try:
            overdue_orders = self.db.get_overdue_orders()
            for i, order in enumerate(overdue_orders):
                formatted_date = self.format_date_for_display(order[7])  # end_date
                tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                self.overdue_tree.insert('', 'end', values=(
                    order[0],  # Sipariş ID
                    order[1],  # Müşteri adı
                    order[2],  # Ürün adı
                    formatted_date,  # Bitiş tarihi (formatlanmış)
                    order[8]   # Durum
                ), tags=(tag,))
        except Exception as e:
            messagebox.showerror("Hata", f"Gecikmiş siparişler yüklenirken hata oluştu: {str(e)}")
    
    def insert_with_urgent_words(self, text_widget, text, base_style, urgent_words):
        """Metni parçalara ayırarak özel kelimeleri kırmızı yazar"""
        import re
        
        # Gün sayısı pattern'i (1 gün, 2 gün, 3 gün, vb.)
        day_pattern = r'\b(\d+)\s+(gün)\b'
        
        current_pos = 0
        
        # Önce gün sayılarını bul ve işle
        for match in re.finditer(day_pattern, text, re.IGNORECASE):
            # Match'ten önceki kısmı ekle
            before_text = text[current_pos:match.start()]
            if before_text:
                self.insert_urgent_words_simple(text_widget, before_text, base_style, urgent_words)
            
            # Gün sayısını kırmızı ekle
            text_widget.insert(tk.END, match.group(), "urgent_text")
            current_pos = match.end()
        
        # Kalan metni işle
        remaining_text = text[current_pos:]
        if remaining_text:
            self.insert_urgent_words_simple(text_widget, remaining_text, base_style, urgent_words)
    
    def insert_urgent_words_simple(self, text_widget, text, base_style, urgent_words):
        """Basit kelime bazlı kırmızı vurgulama"""
        # Boşlukları koruyarak işle
        import re
        words_with_spaces = re.split(r'(\s+)', text)
        
        for part in words_with_spaces:
            if part.isspace():
                # Boşluk karakterleri olduğu gibi ekle
                text_widget.insert(tk.END, part, base_style)
            elif part.strip():
                # Kelimeyi temizle (noktalama işaretleri hariç)
                clean_word = part.strip(".,!?:;")
                
                if clean_word.upper() in [w.upper() for w in urgent_words]:
                    text_widget.insert(tk.END, part, "urgent_text")
                else:
                    text_widget.insert(tk.END, part, base_style)

    def check_notifications(self):
        """Günlük ödeme kontrolü ve bildirimleri gösterir"""
        try:
            payment_data = self.db.get_daily_payment_check()
            notification_messages = []
            notification_styles = []  # Her mesaj için stil ("header" veya "item")

            # Bugün bitenler
            if payment_data['today']:
                notification_messages.append("🔔 BUGÜN BİTEN SİPARİŞLER:")
                notification_styles.append("header_today")
                for order in payment_data['today']:
                    formatted_date = self.format_date_for_display(order[7])
                    notification_messages.append(
                        f"   • {order[1]} müşterisinin '{order[2]}' siparişi BUGÜN bitiyor! (Bitiş: {formatted_date})"
                    )
                    notification_styles.append("item")
                notification_messages.append("")
                notification_styles.append("empty")

            # Yarın bitenler
            if payment_data['tomorrow']:
                notification_messages.append("⏰ YARIN BİTEN SİPARİŞLER:")
                notification_styles.append("header_tomorrow")
                for order in payment_data['tomorrow']:
                    formatted_date = self.format_date_for_display(order[7])
                    notification_messages.append(
                        f"   • {order[1]} müşterisinin '{order[2]}' siparişi YARIN bitiyor! (Bitiş: {formatted_date})"
                    )
                    notification_styles.append("item")
                notification_messages.append("")
                notification_styles.append("empty")

            # Bu hafta bitenler
            if payment_data['week']:
                notification_messages.append("📅 BU HAFTA BİTEN SİPARİŞLER:")
                notification_styles.append("header_week")
                for order in payment_data['week']:
                    days_left = int(order[10])
                    formatted_date = self.format_date_for_display(order[7])
                    notification_messages.append(
                        f"   • {order[1]} müşterisinin '{order[2]}' siparişi {days_left} gün sonra bitiyor! (Bitiş: {formatted_date})"
                    )
                    notification_styles.append("item")
                notification_messages.append("")
                notification_styles.append("empty")

            # Gecikmiş siparişler
            if payment_data['overdue']:
                notification_messages.append("🚨 GECİKMİŞ SİPARİŞLER:")
                notification_styles.append("header_overdue")
                for order in payment_data['overdue']:
                    days_overdue = int(order[10])
                    formatted_date = self.format_date_for_display(order[7])
                    notification_messages.append(
                        f"   • {order[1]} müşterisinin '{order[2]}' siparişi {days_overdue} gün gecikmiş! (Bitiş: {formatted_date})"
                    )
                    notification_styles.append("item")
                notification_messages.append("")
                notification_styles.append("empty")

            if notification_messages:
                # Modern bildirim penceresi oluştur
                notification_window = tk.Toplevel(self.root)
                notification_window.title("🔔 Ödeme Kontrol Bildirimleri")
                notification_window.geometry("1100x850")
                notification_window.configure(bg='#f8f9fa')
                notification_window.transient(self.root)
                notification_window.grab_set()
                notification_window.resizable(False, False)
                
                # Pencereyi ortalama
                notification_window.update_idletasks()
                x = (notification_window.winfo_screenwidth() // 2) - (1100 // 2)
                y = (notification_window.winfo_screenheight() // 2) - (850 // 2)
                notification_window.geometry(f"1100x850+{x}+{y}")
                
                # Ana container - Modern card design
                main_container = tk.Frame(notification_window, bg='#f8f9fa')
                main_container.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
                
                # Modern card frame
                card_frame = tk.Frame(main_container, bg='#ffffff', relief='flat', bd=1)
                card_frame.pack(fill=tk.BOTH, expand=True)
                
                # Header - Modern başlık
                header_frame = tk.Frame(card_frame, bg='#e74c3c', height=80)
                header_frame.pack(fill=tk.X)
                header_frame.pack_propagate(False)
                
                # Başlık metni - ortalanmış
                header_label = tk.Label(header_frame, text="🔔 ÖDEME KONTROL BİLDİRİMLERİ", 
                                       font=('Segoe UI', 20, 'bold'), 
                                       bg='#e74c3c', fg='white')
                header_label.pack(expand=True)
                
                # Content area - Scrollable
                content_frame = tk.Frame(card_frame, bg='#ffffff')
                content_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
                
                # Scrollbar ile Text widget
                text_frame = tk.Frame(content_frame, bg='#ffffff')
                text_frame.pack(fill=tk.BOTH, expand=True)
                
                # Scrollbar
                scrollbar = tk.Scrollbar(text_frame)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # Modern text widget
                text_widget = tk.Text(text_frame, wrap=tk.WORD, 
                                     font=("Segoe UI", 12),
                                     bg='#f8f9fa', fg='#2c3e50',
                                     relief='flat', borderwidth=0,
                                     padx=20, pady=15,
                                     yscrollcommand=scrollbar.set)
                text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.config(command=text_widget.yview)

                # Modern stil tanımları
                text_widget.tag_configure("header_today", 
                                        font=("Segoe UI", 16, "bold"), 
                                        spacing1=15, spacing3=10, 
                                        foreground="#e74c3c",
                                        background="#fff5f5")
                                        
                text_widget.tag_configure("header_tomorrow", 
                                        font=("Segoe UI", 16, "bold"), 
                                        spacing1=15, spacing3=10, 
                                        foreground="#f39c12",
                                        background="#fef9e7")
                                        
                text_widget.tag_configure("header_week", 
                                        font=("Segoe UI", 16, "bold"), 
                                        spacing1=15, spacing3=10, 
                                        foreground="#3498db",
                                        background="#ebf3fd")
                                        
                text_widget.tag_configure("header_overdue", 
                                        font=("Segoe UI", 16, "bold"), 
                                        spacing1=15, spacing3=10, 
                                        foreground="#c0392b",
                                        background="#f8d7da")
                                        
                text_widget.tag_configure("item", 
                                        font=("Segoe UI", 13), 
                                        lmargin1=30, lmargin2=30, 
                                        spacing1=5, spacing3=8,
                                        foreground="#2c3e50")
                                        
                text_widget.tag_configure("product_name", 
                                        font=("Segoe UI", 13, "bold"), 
                                        foreground="#2c3e50")
                                        
                text_widget.tag_configure("urgent_text", 
                                        font=("Segoe UI", 13, "bold"), 
                                        foreground="#e74c3c")
                                        
                text_widget.tag_configure("empty", font=("Segoe UI", 8))

                # İçeriği ekle - Ürün adlarını kalın, özel kelimeleri kırmızı yap
                def insert_formatted_message(text, base_style):
                    """Mesajı özel formatlarla ekler"""
                    # Kırmızı yapılacak kelimeler
                    urgent_words = ["YARIN", "BUGÜN", "GECİKMİŞ", "geçikmiş"]
                    # Gün sayıları için pattern (1 gün, 2 gün, 3 gün, vb.)
                    import re
                    
                    # Önce ürün adını işle
                    if "'" in text:
                        parts = text.split("'")
                        if len(parts) >= 3:
                            # İlk kısım
                            self.insert_with_urgent_words(text_widget, parts[0], base_style, urgent_words)
                            # Ürün adı
                            text_widget.insert(tk.END, f"'{parts[1]}'", "product_name")
                            # Kalan kısım
                            remaining = "'".join(parts[2:])
                            self.insert_with_urgent_words(text_widget, remaining, base_style, urgent_words)
                        else:
                            self.insert_with_urgent_words(text_widget, text, base_style, urgent_words)
                    else:
                        self.insert_with_urgent_words(text_widget, text, base_style, urgent_words)
                    
                    text_widget.insert(tk.END, "\n")
                
                for msg, style in zip(notification_messages, notification_styles):
                    if style == "item":
                        insert_formatted_message(msg, style)
                    else:
                        text_widget.insert(tk.END, msg + "\n", style)

                text_widget.config(state=tk.DISABLED)
                
                # İşlemler Card - Modern button area
                actions_card = tk.Frame(card_frame, bg='#f8f9fa', height=100)
                actions_card.pack(fill=tk.X, side=tk.BOTTOM, pady=20)
                actions_card.pack_propagate(False)
                
                # Buton container - ortalanmış
                button_container = tk.Frame(actions_card, bg='#f8f9fa')
                button_container.pack(expand=True, pady=15)
                
                # Tamam Butonu - Büyük ve ortalanmış
                ok_btn = tk.Button(button_container, text="✅ Tamam", 
                                  command=notification_window.destroy,
                                  bg='#27ae60', fg='white',
                                  font=('Segoe UI', 16, 'bold'),
                                  relief='flat', borderwidth=0,
                                  padx=50, pady=15,
                                  cursor='hand2')
                ok_btn.pack()
                
                # Hover effects
                def on_enter_ok(e):
                    ok_btn.config(bg='#229954')
                def on_leave_ok(e):
                    ok_btn.config(bg='#27ae60')
                    
                ok_btn.bind('<Enter>', on_enter_ok)
                ok_btn.bind('<Leave>', on_leave_ok)

                # Bildirimleri yenile
                self.load_notifications()
            else:
                messagebox.showinfo("Ödeme Kontrol", "Ödemesi gelen bulunmuyor.")
        except Exception as e:
            messagebox.showerror("Hata", f"Günlük ödeme kontrolü yapılırken hata oluştu: {str(e)}")
    
    def complete_overdue_orders(self):
        """Gecikmiş siparişlerin durumunu 'Tamamlandı' yapar"""
        try:
            # Önce gecikmiş siparişleri kontrol et
            overdue_orders = self.db.get_overdue_orders()
            
            if not overdue_orders:
                messagebox.showinfo(
                    "Bilgi",
                    "Tamamlanacak gecikmiş sipariş bulunamadı."
                )
                return
            
            # Modern onay penceresi oluştur
            confirm_dialog = tk.Toplevel(self.root)
            confirm_dialog.title("🔄 Gecikmiş Siparişleri Tamamla")
            confirm_dialog.geometry("1000x700")
            confirm_dialog.transient(self.root)
            confirm_dialog.grab_set()
            confirm_dialog.resizable(True, True)
            confirm_dialog.configure(bg='#f8f9fa')
            
            # Pencereyi merkeze yerleştir
            confirm_dialog.update_idletasks()
            x = (confirm_dialog.winfo_screenwidth() // 2) - (1000 // 2)
            y = (confirm_dialog.winfo_screenheight() // 2) - (700 // 2)
            confirm_dialog.geometry(f"1000x700+{x}+{y}")
            
            # Ana container - Modern card tasarımı
            main_container = tk.Frame(confirm_dialog, bg='#f8f9fa')
            main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Header card
            header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            header_card.pack(fill=tk.X, pady=(0, 20))
            
            # Modern gölge efekti için border frame
            shadow_frame = tk.Frame(main_container, bg='#e0e0e0', height=2)
            shadow_frame.pack(fill=tk.X, pady=(0, 20))
            
            # Başlık - Modern tipografi
            title_frame = tk.Frame(header_card, bg='white')
            title_frame.pack(fill=tk.X, padx=30, pady=20)
            
            tk.Label(title_frame, 
                    text="🔄 Gecikmiş Siparişleri Tamamlama",
                    font=('Segoe UI', 24, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w')
            
            tk.Label(title_frame,
                    text="Seçili siparişlerin durumunu toplu olarak güncelleyin",
                    font=('Segoe UI', 11),
                    bg='white',
                    fg='#7f8c8d').pack(anchor='w', pady=(5, 0))
            
            # Uyarı kartı - Modern alert tasarımı
            warning_card = tk.Frame(main_container, bg='#fff3cd', relief='flat', bd=0)
            warning_card.pack(fill=tk.X, pady=(0, 20))
            
            warning_inner = tk.Frame(warning_card, bg='#fff3cd')
            warning_inner.pack(fill=tk.X, padx=20, pady=15)
            
            tk.Label(warning_inner,
                    text="⚠️ Dikkat",
                    font=('Segoe UI', 14, 'bold'),
                    bg='#fff3cd',
                    fg='#856404').pack(anchor='w')
            
            tk.Label(warning_inner,
                    text="Seçilen gecikmiş siparişlerin durumu 'Tamamlandı' olarak işaretlenecek.\nBu işlem geri alınamaz, lütfen dikkatli olun.",
                    font=('Segoe UI', 12, 'bold'),
                    bg='#fff3cd',
                    fg='#e74c3c',
                    justify='left').pack(anchor='w', pady=(5, 0))
            
            # Modern sipariş listesi kartı
            list_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            list_card.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            
            # Liste başlığı
            list_header = tk.Frame(list_card, bg='white')
            list_header.pack(fill=tk.X, padx=30, pady=(20, 0))
            
            tk.Label(list_header,
                    text=f"📋 Tamamlanacak Siparişler ({len(overdue_orders)} adet)",
                    font=('Segoe UI', 16, 'bold'),
                    bg='white',
                    fg='#2c3e50').pack(anchor='w')
            
            tk.Label(list_header,
                    text="Tamamlamak istediğiniz siparişleri seçin (Ctrl+Click ile çoklu seçim)",
                    font=('Segoe UI', 10),
                    bg='white',
                    fg='#7f8c8d').pack(anchor='w', pady=(2, 0))
            
            # Liste container
            list_container = tk.Frame(list_card, bg='white')
            list_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
            
            # Modern Treeview
            columns = ('Status', 'Müşteri', 'Ürün', 'Bitiş Tarihi', 'Gecikme', 'Durum')
            tree = ttk.Treeview(list_container, columns=columns, show='headings', height=10, selectmode='extended', style='Modern.Treeview')
            
            # Modern sütun başlıkları - sola hizalı
            tree.heading('Status', text='✓', anchor='w')
            tree.heading('Müşteri', text='Müşteri Adı', anchor='w')
            tree.heading('Ürün', text='Ürün/Hizmet', anchor='w')
            tree.heading('Bitiş Tarihi', text='Bitiş Tarihi', anchor='w')
            tree.heading('Gecikme', text='Gecikme', anchor='w')
            tree.heading('Durum', text='Mevcut Durum', anchor='w')
            
            # Modern sütun genişlikleri ve hizalama
            tree.column('Status', width=40, minwidth=40, anchor='w')
            tree.column('Müşteri', width=200, minwidth=150, anchor='w')
            tree.column('Ürün', width=250, minwidth=200, anchor='w')
            tree.column('Bitiş Tarihi', width=120, minwidth=100, anchor='w')
            tree.column('Gecikme', width=100, minwidth=80, anchor='w')
            tree.column('Durum', width=120, minwidth=100, anchor='w')
            
            # Modern renk şeması
            tree.tag_configure('overdue_critical', background='#ffebee', foreground='#c62828')  # Çok geç
            tree.tag_configure('overdue_warning', background='#fff3e0', foreground='#ef6c00')   # Geç
            tree.tag_configure('overdue_normal', background='#f3e5f5', foreground='#7b1fa2')   # Normal gecikme
            tree.tag_configure('selected', background='#e3f2fd', foreground='#1976d2')
            
            # Modern TTK Scrollbar (same as customer registration)
            complete_scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, 
                                             command=tree.yview, style='Modern.Vertical.TScrollbar')
            tree.configure(yscrollcommand=complete_scrollbar.set)
            
            # Pack widgets for complete dialog (same as customer registration)
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            complete_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                try:
                    top, bottom = tree.yview()
                    scrollbar_height = complete_scrollbar_bg.winfo_height()
                    
                    if bottom - top >= 1.0:
                        # Tüm içerik görünüyorsa tam boy scrollbar göster
                        complete_scrollbar_bg.create_rectangle(
                            1, 0, 7, scrollbar_height,
                            fill='#e0e0e0', outline='', tags="thumb",
                            width=0
                        )
                        return
                        
                    thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                    thumb_y = int(scrollbar_height * top)
                    
                    complete_scrollbar_bg.create_rectangle(
                        1, thumb_y, 7, thumb_y + thumb_height,
                        fill='#c0c0c0', outline='', tags="thumb",
                        width=0
                    )
                except:
                    pass
            
            # Scrollbar drag variables for complete dialog
            complete_scrollbar_dragging = False
            
            def complete_smooth_scroll(event):
                try:
                    if tree.winfo_exists():
                        tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
                        update_complete_scrollbar()
                except tk.TclError:
                    pass
            
            def on_complete_scrollbar_click(event):
                nonlocal complete_scrollbar_dragging
                try:
                    top, bottom = tree.yview()
                    if bottom - top >= 1.0:
                        return
                    scrollbar_height = complete_scrollbar_bg.winfo_height()
                    click_position = max(0, min(1, event.y / scrollbar_height))
                    tree.yview_moveto(click_position)
                    update_complete_scrollbar()
                    complete_scrollbar_dragging = True
                except:
                    pass
            
            def on_complete_scrollbar_drag(event):
                nonlocal complete_scrollbar_dragging
                if not complete_scrollbar_dragging:
                    return
                try:
                    top, bottom = tree.yview()
                    if bottom - top >= 1.0:
                        return
                    scrollbar_height = complete_scrollbar_bg.winfo_height()
                    click_position = max(0, min(1, event.y / scrollbar_height))
                    tree.yview_moveto(click_position)
                    update_complete_scrollbar()
                except:
                    pass
            
            def on_complete_scrollbar_release(event):
                nonlocal complete_scrollbar_dragging
                complete_scrollbar_dragging = False
            
            def on_complete_scrollbar_enter(event):
                complete_scrollbar_bg.delete("thumb")
                try:
                    top, bottom = tree.yview()
                    if bottom - top >= 1.0:
                        return
                    scrollbar_height = complete_scrollbar_bg.winfo_height()
                    thumb_height = max(20, int(scrollbar_height * (bottom - top)))
                    thumb_y = int(scrollbar_height * top)
                    complete_scrollbar_bg.create_rectangle(
                        1, thumb_y, 7, thumb_y + thumb_height,
                        fill='#999999', outline='', tags="thumb",
                        width=0
                    )
                except:
                    pass
            
            def on_complete_scrollbar_leave(event):
                update_complete_scrollbar()
            
            def on_complete_tree_scroll(*args):
                update_complete_scrollbar()
            
            tree.configure(yscrollcommand=on_complete_tree_scroll)
            
            # Bind events for complete dialog
            tree.bind("<MouseWheel>", complete_smooth_scroll)
            complete_scrollbar_bg.bind("<Button-1>", on_complete_scrollbar_click)
            complete_scrollbar_bg.bind("<B1-Motion>", on_complete_scrollbar_drag)
            complete_scrollbar_bg.bind("<ButtonRelease-1>", on_complete_scrollbar_release)
            complete_scrollbar_bg.bind("<Enter>", on_complete_scrollbar_enter)
            complete_scrollbar_bg.bind("<Leave>", on_complete_scrollbar_leave)
            complete_scrollbar_bg.configure(takefocus=True)
            
            
            # Siparişleri modern görünümle listeye ekle
            for i, order in enumerate(overdue_orders):
                order_id = order[0]  # o.id
                customer_name = order[1]  # c.name
                product_name = order[2]  # o.product_name
                end_date = order[7]  # o.end_date
                status = order[8]  # o.status
                formatted_date = self.format_date_for_display(end_date)
                
                # Gecikme süresini hesapla
                try:
                    from datetime import datetime
                    # Farklı tarih formatlarını dene
                    end_date_obj = None
                    
                    # Format 1: DD.MM.YYYY
                    try:
                        end_date_obj = datetime.strptime(end_date, '%d.%m.%Y')
                    except:
                        # Format 2: YYYY-MM-DD (Database formatı)
                        try:
                            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                        except:
                            # Format 3: DD/MM/YYYY
                            try:
                                end_date_obj = datetime.strptime(end_date, '%d/%m/%Y')
                            except:
                                pass
                    
                    if end_date_obj:
                        today = datetime.now()
                        days_overdue = (today - end_date_obj).days
                        
                        if days_overdue > 30:
                            gecikme_text = f"{days_overdue} gün (Kritik)"
                            tag = 'overdue_critical'
                            status_icon = "🔴"
                        elif days_overdue > 7:
                            gecikme_text = f"{days_overdue} gün"
                            tag = 'overdue_warning'
                            status_icon = "🟡"
                        else:
                            gecikme_text = f"{days_overdue} gün"
                            tag = 'overdue_normal'
                            status_icon = "🟠"
                    else:
                        # Tarih parse edilemedi, database'den gelen raw değeri kullan
                        if len(order) > 10 and order[10] is not None:
                            try:
                                days_overdue = int(float(order[10]))  # Database'den gelen days_overdue
                                if days_overdue > 30:
                                    gecikme_text = f"{days_overdue} gün (Kritik)"
                                    tag = 'overdue_critical'
                                    status_icon = "🔴"
                                elif days_overdue > 7:
                                    gecikme_text = f"{days_overdue} gün"
                                    tag = 'overdue_warning'
                                    status_icon = "🟡"
                                else:
                                    gecikme_text = f"{days_overdue} gün"
                                    tag = 'overdue_normal'
                                    status_icon = "🟠"
                            except:
                                gecikme_text = "Hesaplanamadı"
                                tag = 'overdue_normal'
                                status_icon = "❓"
                        else:
                            gecikme_text = "Hesaplanamadı"
                            tag = 'overdue_normal'
                            status_icon = "❓"
                        
                except Exception as e:
                    gecikme_text = f"Hata: {str(e)}"
                    tag = 'overdue_normal'
                    status_icon = "❓"
                
                tree.insert('', 'end', values=(
                    status_icon,           # Durum ikonu
                    customer_name,         # Müşteri
                    product_name,          # Ürün
                    formatted_date,        # Bitiş tarihi
                    gecikme_text,          # Gecikme
                    status                 # Mevcut durum
                ), tags=(tag,))
            
            def confirm_action():
                # Seçilen siparişleri al
                selected_items = tree.selection()
                if not selected_items:
                    messagebox.showwarning("⚠️ Uyarı", "Lütfen tamamlanacak siparişleri seçin!")
                    return
                
                # Seçilen siparişlerin ID'lerini al
                selected_order_ids = []
                for item in selected_items:
                    values = tree.item(item)['values']
                    # Yeni düzen: status_icon, customer_name, product_name, end_date, gecikme, status
                    customer_name = values[1]  # Müşteri
                    product_name = values[2]   # Ürün
                    end_date = values[3]       # Bitiş tarihi
                    
                    # overdue_orders listesinden bu siparişi bul
                    for order in overdue_orders:
                        if (order[1] == customer_name and 
                            order[2] == product_name and 
                            self.format_date_for_display(order[7]) == end_date):
                            selected_order_ids.append(order[0])  # order_id
                            break
                
                if not selected_order_ids:
                    messagebox.showerror("❌ Hata", "Seçilen siparişler bulunamadı!")
                    return
                
                confirm_dialog.destroy()
                self._execute_complete_selected_overdue_orders(selected_order_ids)
            
            def cancel_action():
                confirm_dialog.destroy()
            
            # Modern buton kartı
            button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
            button_card.pack(fill=tk.X, pady=(0, 0))
            
            button_container = tk.Frame(button_card, bg='white')
            button_container.pack(pady=50)
            
            # Modern buton stilleri
            modern_button_style = {
                'font': ('Segoe UI', 12, 'bold'),
                'relief': 'flat',
                'borderwidth': 0,
                'padx': 40,
                'pady': 15,
                'cursor': 'hand2'
            }
            
            # Ana aksiyon butonu (Yeşil)
            complete_btn = tk.Button(button_container, 
                                   text="✅ Seçilen Siparişleri Tamamla",
                                   command=confirm_action,
                                   bg='#27ae60',
                                   fg='white',
                                   activebackground='#229954',
                                   activeforeground='white',
                                   **modern_button_style)
            complete_btn.pack(side=tk.LEFT, padx=15)
            
            # İptal butonu (Gri)
            cancel_btn = tk.Button(button_container,
                                 text="❌ İptal",
                                 command=cancel_action,
                                 bg='#95a5a6',
                                 fg='white',
                                 activebackground='#7f8c8d',
                                 activeforeground='white',
                                 **modern_button_style)
            cancel_btn.pack(side=tk.LEFT, padx=15)
            
            # Hover efektleri
            def on_enter(e):
                if e.widget == complete_btn:
                    e.widget['bg'] = '#229954'
                elif e.widget == cancel_btn:
                    e.widget['bg'] = '#7f8c8d'
            
            def on_leave(e):
                if e.widget == complete_btn:
                    e.widget['bg'] = '#27ae60'
                elif e.widget == cancel_btn:
                    e.widget['bg'] = '#95a5a6'
            
            complete_btn.bind('<Enter>', on_enter)
            complete_btn.bind('<Leave>', on_leave)
            cancel_btn.bind('<Enter>', on_enter)
            cancel_btn.bind('<Leave>', on_leave)
            
        except Exception as e:
            messagebox.showerror(
                "Hata",
                f"Gecikmiş siparişler kontrol edilirken hata oluştu:\n{str(e)}"
            )
    
    def _execute_complete_selected_overdue_orders(self, selected_order_ids):
        """Seçilen gecikmiş siparişleri tamamlar (onay sonrası)"""
        try:
            
            # Seçilen gecikmiş siparişleri tamamla
            result = self.db.complete_selected_overdue_orders(selected_order_ids)
            updated_count = result['updated_count']
            completed_orders = result['completed_orders']
            
            if updated_count > 0:
                # Modern başarı penceresi oluştur
                success_dialog = tk.Toplevel(self.root)
                success_dialog.title("🎉 İşlem Tamamlandı")
                success_dialog.geometry("800x600")
                success_dialog.transient(self.root)
                success_dialog.grab_set()
                success_dialog.configure(bg='#f8f9fa')
                
                # Pencereyi merkeze yerleştir
                success_dialog.update_idletasks()
                x = (success_dialog.winfo_screenwidth() // 2) - (800 // 2)
                y = (success_dialog.winfo_screenheight() // 2) - (600 // 2)
                success_dialog.geometry(f"800x600+{x}+{y}")
                
                # Ana container
                main_container = tk.Frame(success_dialog, bg='#f8f9fa')
                main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
                
                # Başarı kartı - Hero section
                success_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
                success_card.pack(fill=tk.X, pady=(0, 25))
                
                success_inner = tk.Frame(success_card, bg='white')
                success_inner.pack(fill=tk.X, padx=40, pady=30)
                
                # Büyük başarı ikonu ve mesajı
                tk.Label(success_inner,
                        text="🎉",
                        font=('Segoe UI', 48),
                        bg='white').pack()
                
                tk.Label(success_inner,
                        text="İşlem Başarıyla Tamamlandı!",
                        font=('Segoe UI', 24, 'bold'),
                        bg='white',
                        fg='#27ae60').pack(pady=(10, 5))
                
                tk.Label(success_inner,
                        text=f"{updated_count} adet sipariş durumu güncellendi",
                        font=('Segoe UI', 14),
                        bg='white',
                        fg='#7f8c8d').pack(pady=(0, 10))
                
                # İstatistik kartları
                stats_container = tk.Frame(success_inner, bg='white')
                stats_container.pack(fill=tk.X, pady=(15, 0))
                
                # Sol stat
                left_stat = tk.Frame(stats_container, bg='#e8f5e8', relief='flat', bd=0)
                left_stat.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
                
                tk.Label(left_stat,
                        text="✅ Tamamlanan",
                        font=('Segoe UI', 12, 'bold'),
                        bg='#e8f5e8',
                        fg='#27ae60').pack(pady=(15, 5))
                
                tk.Label(left_stat,
                        text=str(updated_count),
                        font=('Segoe UI', 20, 'bold'),
                        bg='#e8f5e8',
                        fg='#27ae60').pack(pady=(0, 15))
                
                # Sağ stat - Toplam süre
                right_stat = tk.Frame(stats_container, bg='#e3f2fd', relief='flat', bd=0)
                right_stat.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
                
                tk.Label(right_stat,
                        text="⏱️ İşlem Süresi",
                        font=('Segoe UI', 12, 'bold'),
                        bg='#e3f2fd',
                        fg='#1976d2').pack(pady=(15, 5))
                
                tk.Label(right_stat,
                        text="< 1 saniye",
                        font=('Segoe UI', 16, 'bold'),
                        bg='#e3f2fd',
                        fg='#1976d2').pack(pady=(0, 15))
                
                # Detaylar kartı
                details_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
                details_card.pack(fill=tk.BOTH, expand=True, pady=(0, 25))
                
                details_header = tk.Frame(details_card, bg='white')
                details_header.pack(fill=tk.X, padx=30, pady=(25, 0))
                
                tk.Label(details_header,
                        text="📋 Güncellenen Siparişler",
                        font=('Segoe UI', 18, 'bold'),
                        bg='white',
                        fg='#2c3e50').pack(anchor='w')
                
                tk.Label(details_header,
                        text="Aşağıdaki siparişlerin durumu 'Tamamlandı' olarak işaretlendi:",
                        font=('Segoe UI', 11),
                        bg='white',
                        fg='#7f8c8d').pack(anchor='w', pady=(5, 0))
                
                # Modern liste container
                list_container = tk.Frame(details_card, bg='white')
                list_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
                
                # Modern treeview
                columns = ('Müşteri', 'Ürün/Hizmet', 'Bitiş Tarihi', 'Durum')
                details_tree = ttk.Treeview(list_container, columns=columns, show='headings', height=8, style='Modern.Treeview')
                
                # Sütun başlıkları - sola hizalı
                details_tree.heading('Müşteri', text='Müşteri Adı', anchor='w')
                details_tree.heading('Ürün/Hizmet', text='Ürün/Hizmet', anchor='w')
                details_tree.heading('Bitiş Tarihi', text='Bitiş Tarihi', anchor='w')
                details_tree.heading('Durum', text='Yeni Durum', anchor='w')
                
                # Sütun genişlikleri ve hizalama
                details_tree.column('Müşteri', width=200, anchor='w')
                details_tree.column('Ürün/Hizmet', width=250, anchor='w')
                details_tree.column('Bitiş Tarihi', width=120, anchor='w')
                details_tree.column('Durum', width=120, anchor='w')
                
                # Başarı renk şeması
                details_tree.tag_configure('completed', background='#e8f5e8', foreground='#27ae60')
                details_tree.tag_configure('alternate', background='#f8f9fa')
                
                # Scrollbar
                details_scrollbar = ttk.Scrollbar(list_container, orient=tk.VERTICAL, command=details_tree.yview, style='Modern.Vertical.TScrollbar')
                details_tree.configure(yscrollcommand=details_scrollbar.set)
                
                details_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
                details_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                
                # Siparişleri listeye ekle
                for i, order in enumerate(completed_orders):
                    order_id = order[0]  # o.id
                    customer_name = order[1]  # c.name
                    product_name = order[2]  # o.product_name
                    end_date = order[3]  # o.end_date
                    status = order[4]  # o.status (Tamamlandı olmalı)
                    formatted_date = self.format_date_for_display(end_date)
                    
                    tag = 'completed' if i % 2 == 0 else 'alternate'
                    details_tree.insert('', 'end', values=(
                        customer_name,
                        product_name,
                        formatted_date,
                        "✅ Tamamlandı"
                    ), tags=(tag,))
                
                def close_dialog():
                    success_dialog.destroy()
                    # Siparişler listesini yenile
                    self.load_orders()
                    self.load_notifications()
                
                # Modern buton kartı
                button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
                button_card.pack(fill=tk.X)
                
                button_container = tk.Frame(button_card, bg='white')
                button_container.pack(pady=25)
                
                # Modern kapatma butonu
                close_btn = tk.Button(button_container,
                                    text="🎉 Harika! Kapat",
                                    command=close_dialog,
                                    font=('Segoe UI', 14, 'bold'),
                                    bg='#27ae60',
                                    fg='white',
                                    relief='flat',
                                    borderwidth=0,
                                    padx=50,
                                    pady=15,
                                    cursor='hand2',
                                    activebackground='#229954',
                                    activeforeground='white')
                close_btn.pack()
                
                # Hover efekti
                def on_hover_enter(e):
                    e.widget['bg'] = '#229954'
                def on_hover_leave(e):
                    e.widget['bg'] = '#27ae60'
                
                close_btn.bind('<Enter>', on_hover_enter)
                close_btn.bind('<Leave>', on_hover_leave)
                
                # Otomatik kayıt
                self.auto_save_data("Seçilen Gecikmiş Siparişler Tamamlandı")
                
            else:
                messagebox.showinfo(
                    "Bilgi",
                    "Seçilen siparişler zaten tamamlanmış veya gecikmiş değil."
                )
                
        except Exception as e:
            messagebox.showerror(
                "Hata",
                f"Seçilen siparişler tamamlanırken hata oluştu:\n{str(e)}"
            )
    
    def show_modern_date_picker(self, entry_widget, title="Tarih Seç"):
        """Modern takvim picker dialog'u"""
        from datetime import datetime, timedelta
        import calendar
        
        # Dialog oluştur - Büyük boyut
        dialog = tk.Toplevel(self.root)
        dialog.title(f"📅 {title}")
        dialog.geometry("520x800")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg='#f8f9fa')
        dialog.resizable(False, False)
        
        # Pencereyi merkeze yerleştir
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (520 // 2)
        y = (dialog.winfo_screenheight() // 2) - (800 // 2)
        dialog.geometry(f"520x800+{x}+{y}")
        
        # Mevcut tarihi al veya bugünü kullan
        current_date = datetime.now()
        try:
            current_text = entry_widget.get().strip()
            if current_text:
                current_date = datetime.strptime(current_text, "%d.%m.%Y")
        except:
            pass
        
        selected_date = [current_date]  # List kullanarak referans geçmek için
        
        # Ana container
        main_container = tk.Frame(dialog, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header card
        header_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        header_card.pack(fill=tk.X, pady=(0, 15))
        
        header_inner = tk.Frame(header_card, bg='white')
        header_inner.pack(fill=tk.X, padx=20, pady=15)
        
        # Ay/Yıl navigation
        nav_frame = tk.Frame(header_inner, bg='white')
        nav_frame.pack(fill=tk.X)
        
        def prev_month():
            if selected_date[0].month == 1:
                selected_date[0] = selected_date[0].replace(year=selected_date[0].year-1, month=12)
            else:
                selected_date[0] = selected_date[0].replace(month=selected_date[0].month-1)
            update_calendar()
        
        def next_month():
            if selected_date[0].month == 12:
                selected_date[0] = selected_date[0].replace(year=selected_date[0].year+1, month=1)
            else:
                selected_date[0] = selected_date[0].replace(month=selected_date[0].month+1)
            update_calendar()
        
        # Sol ok
        prev_btn = tk.Button(nav_frame, text="◀", command=prev_month,
                           font=('Segoe UI', 14, 'bold'), bg='#3498db', fg='white',
                           relief='flat', borderwidth=0, padx=15, pady=8,
                           cursor='hand2', activebackground='#2980b9')
        prev_btn.pack(side=tk.LEFT)
        
        # Ay/Yıl label
        month_label = tk.Label(nav_frame, text="", font=('Segoe UI', 16, 'bold'),
                             bg='white', fg='#2c3e50')
        month_label.pack(side=tk.LEFT, expand=True)
        
        # Sağ ok
        next_btn = tk.Button(nav_frame, text="▶", command=next_month,
                           font=('Segoe UI', 14, 'bold'), bg='#3498db', fg='white',
                           relief='flat', borderwidth=0, padx=15, pady=8,
                           cursor='hand2', activebackground='#2980b9')
        next_btn.pack(side=tk.RIGHT)
        
        # Takvim container
        calendar_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        calendar_card.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        calendar_inner = tk.Frame(calendar_card, bg='white')
        calendar_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Günler frame
        days_frame = tk.Frame(calendar_inner, bg='white')
        days_frame.pack(fill=tk.BOTH, expand=True)
        
        # Haftanın günleri başlıkları
        week_days = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Paz']
        for i, day in enumerate(week_days):
            day_label = tk.Label(days_frame, text=day, font=('Segoe UI', 11, 'bold'),
                               bg='#ecf0f1', fg='#34495e', width=5, height=2)
            day_label.grid(row=0, column=i, padx=1, pady=1, sticky='nsew')
        
        # Grid weight ayarları
        for i in range(7):
            days_frame.columnconfigure(i, weight=1)
        for i in range(7):  # 1 header + 6 hafta
            days_frame.rowconfigure(i, weight=1)
        
        def update_calendar():
            # Önceki gün butonlarını temizle
            for widget in days_frame.winfo_children():
                if int(widget.grid_info().get('row', 0)) > 0:
                    widget.destroy()
            
            # Ay/yıl label'ı güncelle
            month_names = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
                          'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
            month_label.config(text=f"{month_names[selected_date[0].month]} {selected_date[0].year}")
            
            # Ayın takvimini oluştur
            cal = calendar.monthcalendar(selected_date[0].year, selected_date[0].month)
            
            current_day = selected_date[0].day
            current_month = selected_date[0].month
            current_year = selected_date[0].year
            
            for week_num, week in enumerate(cal):
                for day_num, day in enumerate(week):
                    if day == 0:
                        # Boş hücre
                        empty_label = tk.Label(days_frame, text="", bg='white')
                        empty_label.grid(row=week_num+1, column=day_num, padx=1, pady=1, sticky='nsew')
                    else:
                        # Gün butonu
                        is_selected = (day == current_day)
                        is_today = (day == datetime.now().day and 
                                  current_month == datetime.now().month and 
                                  current_year == datetime.now().year)
                        
                        if is_selected:
                            bg_color = '#3498db'
                            fg_color = 'white'
                            font_style = ('Segoe UI', 11, 'bold')
                        elif is_today:
                            bg_color = '#e8f5e8'
                            fg_color = '#27ae60'
                            font_style = ('Segoe UI', 11, 'bold')
                        else:
                            bg_color = '#f8f9fa'
                            fg_color = '#2c3e50'
                            font_style = ('Segoe UI', 11)
                        
                        def select_day(d=day):
                            selected_date[0] = selected_date[0].replace(day=d)
                            update_calendar()
                        
                        day_btn = tk.Button(days_frame, text=str(day),
                                          command=select_day,
                                          font=font_style, bg=bg_color, fg=fg_color,
                                          relief='flat', borderwidth=0,
                                          cursor='hand2', width=5, height=2,
                                          activebackground='#2980b9', activeforeground='white')
                        day_btn.grid(row=week_num+1, column=day_num, padx=1, pady=1, sticky='nsew')
        
        # İlk takvimi oluştur
        update_calendar()
        
        # Buton kartı
        button_card = tk.Frame(main_container, bg='white', relief='flat', bd=0)
        button_card.pack(fill=tk.X, pady=(10, 0))
        
        # Açıklama metni
        info_frame = tk.Frame(button_card, bg='white')
        info_frame.pack(pady=(20, 10))
        
        tk.Label(info_frame,
                text="📌 Yukarıdan bir tarih seçin ve kaydedin",
                font=('Segoe UI', 12, 'bold'),
                bg='white',
                fg='#34495e').pack()
        
        button_container = tk.Frame(button_card, bg='white')
        button_container.pack(pady=(0, 20))
        
        def select_date():
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, selected_date[0].strftime("%d.%m.%Y"))
            dialog.destroy()
        
        def cancel_selection():
            dialog.destroy()
        
        # Büyük ve belirgin butonlar
        select_btn = tk.Button(button_container, text="💾 Tarihi Kaydet",
                             command=select_date,
                             font=('Segoe UI', 14, 'bold'),
                             bg='#27ae60', fg='white',
                             relief='flat', borderwidth=2,
                             padx=40, pady=15, cursor='hand2',
                             activebackground='#229954',
                             width=15, height=2)
        select_btn.pack(side=tk.LEFT, padx=15)
        
        cancel_btn = tk.Button(button_container, text="❌ İptal",
                             command=cancel_selection,
                             font=('Segoe UI', 14, 'bold'),
                             bg='#e74c3c', fg='white',
                             relief='flat', borderwidth=2,
                             padx=40, pady=15, cursor='hand2',
                             activebackground='#c0392b',
                             width=12, height=2)
        cancel_btn.pack(side=tk.LEFT, padx=15)
        
        # Hover efektleri
        def on_hover_enter(e):
            if e.widget == select_btn:
                e.widget['bg'] = '#229954'
            elif e.widget == cancel_btn:
                e.widget['bg'] = '#c0392b'
        
        def on_hover_leave(e):
            if e.widget == select_btn:
                e.widget['bg'] = '#27ae60'
            elif e.widget == cancel_btn:
                e.widget['bg'] = '#e74c3c'
        
        select_btn.bind('<Enter>', on_hover_enter)
        select_btn.bind('<Leave>', on_hover_leave)
        cancel_btn.bind('<Enter>', on_hover_enter)
        cancel_btn.bind('<Leave>', on_hover_leave)
    
 